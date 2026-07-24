import sqlite3
import random
import re
from datetime import timedelta
import csv
import json
import mimetypes
import math
import base64
import os
import shutil
import tempfile
from pathlib import Path
from zipfile import ZipFile
from functools import wraps
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
from django.core.mail import send_mail
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction, connection, connections
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, Http404
from django.forms.models import model_to_dict
from xml.etree.ElementTree import Element, SubElement, tostring
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    XLImage = None

from .models import (
    Empresa,
    Fornecedor,
    Funcionario,
    Frota,
    Combustivel,
    Lancamento,
    UsuarioSistema,
    BaixaFatura,
    Cliente,
    Vendedor,
    CategoriaArtigo,
    Armazem,
    Artigo,
    SerieDocumento,
    TemplateDocumento,
    CentroCusto,
    NivelAprovacao,
    DocumentoVenda,
    DocumentoVendaLinha,
    RecebimentoCliente,
    PasswordResetCode,
    ConfiguracaoSistema,
    ConfiguracaoFiscal,
    RevisaoFrota,
)

try:
    from .models import ItemCombustivel
except Exception:
    try:
        from .models import LancamentoCombustivel as ItemCombustivel
    except Exception:
        ItemCombustivel = None

try:
    from .models import ManutencaoFrota
except Exception:
    ManutencaoFrota = None




def _normalizar_nif_busca(valor):
    return ''.join(ch for ch in str(valor or '') if ch.isdigit())


def _resolver_fornecedor_por_busca(texto, request=None):
    texto = (texto or '').strip()
    if not texto:
        return None

    normalizado = _normalizar_nif_busca(texto)
    qs = _fornecedores_empresa_qs(request) if request is not None else Fornecedor.objects.all()

    exato = qs.filter(Q(nome__iexact=texto) | Q(nif__iexact=texto)).first()
    if exato:
        return exato

    if normalizado:
        for item in qs.only('id', 'nome', 'nif').order_by('nome'):
            if _normalizar_nif_busca(getattr(item, 'nif', '')) == normalizado:
                return item

    parcial = qs.filter(Q(nome__icontains=texto) | Q(nif__icontains=texto)).order_by('nome').first()
    if parcial:
        return parcial

    if normalizado:
        for item in qs.only('id', 'nome', 'nif').order_by('nome'):
            if normalizado in _normalizar_nif_busca(getattr(item, 'nif', '')):
                return item

    return None


def _resolver_cliente_por_busca(texto):
    texto = (texto or '').strip()
    if not texto:
        return None

    normalizado = _normalizar_nif_busca(texto)
    qs = Cliente.objects.all()

    exato = qs.filter(Q(nome__iexact=texto) | Q(nif__iexact=texto)).first()
    if exato:
        return exato

    if normalizado:
        for item in qs.only('id', 'nome', 'nif').order_by('nome'):
            if _normalizar_nif_busca(getattr(item, 'nif', '')) == normalizado:
                return item

    parcial = qs.filter(Q(nome__icontains=texto) | Q(nif__icontains=texto)).order_by('nome').first()
    if parcial:
        return parcial

    if normalizado:
        for item in qs.only('id', 'nome', 'nif').order_by('nome'):
            if normalizado in _normalizar_nif_busca(getattr(item, 'nif', '')):
                return item

    return None

def _is_ajax(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _to_decimal(value, default="0.00"):
    if value in [None, ""]:
        value = default
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _to_date(value):
    # Aceita tanto YYYY-MM-DD (input date) quanto DD/MM/YYYY digitado manualmente.
    if not value:
        return None
    if isinstance(value, date):
        return value
    value = str(value).strip()
    if not value:
        return None
    # Alguns browsers/dispositivos podem enviar um datetime completo
    # (ex.: "2026-02-01T00:00:00Z") em vez de apenas a data. Sem isto,
    # o valor não batia com nenhum dos formatos aceites abaixo e o filtro
    # era descartado silenciosamente, fazendo a pesquisa devolver resultados
    # sem limite (outros meses além do período pedido).
    value = value.split("T")[0].split(" ")[0]
    parsed = parse_date(value)
    if parsed:
        return parsed
    from datetime import datetime
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass
    return None
    return None


def _json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return {}


def _save_model_or_error(obj, mensagem_duplicado=None):
    try:
        obj.full_clean()
        obj.save()
        return None
    except IntegrityError:
        return mensagem_duplicado or "Já existe um registo com estes dados."
    except ValidationError as exc:
        if hasattr(exc, "message_dict"):
            partes = []
            for campo, erros in exc.message_dict.items():
                if isinstance(erros, (list, tuple)):
                    partes.extend([f"{campo}: {e}" for e in erros])
                else:
                    partes.append(f"{campo}: {erros}")
            return " ; ".join(partes) or "Dados inválidos."
        return "; ".join(exc.messages) or "Dados inválidos."
    except Exception as exc:
        return str(exc) or "Erro ao guardar registo."


def _resolver_frota_por_matricula(matricula, request=None):
    matricula = (matricula or "").strip()
    if not matricula:
        return None
    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota) if request is not None else Frota.objects.all()
    obj = qs.filter(matricula__iexact=matricula).first()
    if obj:
        return obj
    obj = Frota(matricula=matricula)
    if request is not None:
        _atribuir_empresa_ativa(request, obj)
    obj.save()
    return obj


def _normalizar_status_fatura(status):
    status = (status or "").strip().lower()
    mapa = {
        "aberta": "Em aberto",
        "em aberto": "Em aberto",
        "em_aberto": "Em aberto",
        "parcial": "Parcial",
        "paga": "Paga",
        "pago": "Paga",
        "fechada": "Paga",
        "fechado": "Paga",
        "quitada": "Paga",
        "quitado": "Paga",
        "a vencer": "Em aberto",
        "vencida": "Em aberto",
    }
    return mapa.get(status, "")


def _empresa_ativa(request):
    empresas = list(_empresas_usuario(request))
    if not empresas:
        return None

    empresa_id = request.session.get("empresa_ativa_id")
    if empresa_id:
        for empresa in empresas:
            if str(empresa.id) == str(empresa_id):
                return empresa
    return empresas[0]


def _fornecedores_empresa_qs(request):
    qs = Fornecedor.objects.all()
    empresa = _empresa_ativa(request)
    if empresa and hasattr(Fornecedor, "empresa_id"):
        qs = qs.filter(empresa=empresa)
    return qs


def _lancamentos_empresa_qs(request):
    return _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor").all())


def _filtrar_empresa_generico(request, qs, model=None, campo="empresa"):
    empresa = _empresa_ativa(request)
    model = model or getattr(qs, "model", None)
    if empresa and model is not None and hasattr(model, f"{campo}_id"):
        return qs.filter(**{campo: empresa})
    return qs


def _atribuir_empresa_ativa(request, obj, campo="empresa"):
    empresa = _empresa_ativa(request)
    if empresa and hasattr(obj, campo):
        setattr(obj, campo, empresa)
    return empresa


def _get_objeto_empresa_or_404(request, model, objeto_id, campo="empresa"):
    return get_object_or_404(_filtrar_empresa_generico(request, model.objects.all(), model=model, campo=campo), id=objeto_id)


def _get_fornecedor_empresa_or_404(request, fornecedor_id):
    return get_object_or_404(_fornecedores_empresa_qs(request), id=fornecedor_id)


def _get_lancamento_empresa_or_404(request, lancamento_id):
    return get_object_or_404(_filtrar_lancamentos_empresa(request, Lancamento.objects.all()), id=lancamento_id)


def _configuracao_sistema():
    try:
        if _table_exists("Elion_configuracaosistema"):
            return ConfiguracaoSistema.objects.order_by("id").first()
    except Exception:
        return None
    return None


def _empresas_usuario(request):
    try:
        user = getattr(request, "user", None)
        username = (getattr(user, "username", "") or "").strip().casefold()
        if username == "maia" or getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
            return Empresa.objects.all().order_by("nome")

        perfil = UsuarioSistema.objects.filter(user=request.user).first()
        if perfil:
            if getattr(perfil, "administrador_geral", False):
                return Empresa.objects.all().order_by("nome")
            if hasattr(perfil, "empresas"):
                return perfil.empresas.all().order_by("nome")
            if hasattr(perfil, "empresa") and getattr(perfil, "empresa_id", None):
                return Empresa.objects.filter(id=perfil.empresa_id).order_by("nome")
    except Exception:
        pass
    return Empresa.objects.all().order_by("nome")



ROTINAS_PERMISSOES = [
    ("fornecedor", "CRM", "Fornecedor CRM"),
    ("funcionario", "CRM", "Funcionário CRM"),
    ("frota", "CRM", "Frota CRM"),
    ("combustivel", "CRM", "Combustível CRM"),
    ("manutencao", "CRM", "Manutenção CRM"),
    ("revisao_frota", "Transporte", "Revisão de Frota"),

    ("lancamentos", "Fornecedores", "Lançamentos de Fatura"),
    ("consulta", "Fornecedores", "Consultar Pagamentos"),
    ("baixa_faturas", "Fornecedores", "Pagamentos a Fornecedor"),

    ("relatorio_manutencao", "Transporte", "Relatório de Manutenção"),
    ("relatorio_revisao", "Transporte", "Relatório de Revisão"),
    ("dashboard_revisao", "Transporte", "Dashboard Revisão de Frota"),
    ("relatorio_documentos", "Transporte", "Documento de Frota"),

    ("relatorios", "Relatórios", "Relatórios"),
    ("configuracoes", "Configurações", "Configurações"),
]

PERMISSOES_AGREGADAS = {
    "relatorios": [
        "relatorio_financeiro",
        "relatorio_fornecedor",
        "relatorio_faturas",
        "relatorio_manutencao",
        "relatorio_caixa",
        "relatorio_combustivel",
        "relatorio_documentos",
        "extrato_clientes",
        "extrato_fornecedores",
    ],
    "configuracoes": [
        "series",
        "iva_caixa",
        "saft",
        "inventario_existencias",
        "numeracao_series",
        "configuracao_iva",
        "impostos",
        "saft_inventario_ficheiros",
        "resumo_stock_legal",
        "artigos_stocks",
        "empresa",
        "usuario",
        "configuracoes_sistema",
        "exportacoes",
    ],
}

ACOES_PERMISSAO = ["view", "create", "edit", "delete", "export"]


def _todas_rotinas_permissoes():
    rotinas = [rotina for rotina, _, _ in ROTINAS_PERMISSOES]
    for itens in PERMISSOES_AGREGADAS.values():
        rotinas.extend(itens)
    return list(dict.fromkeys(rotinas))


def _permissoes_padrao_total():
    return {rotina: {acao: True for acao in ACOES_PERMISSAO} for rotina in _todas_rotinas_permissoes()}


def _permissoes_vazias():
    return {rotina: {acao: False for acao in ACOES_PERMISSAO} for rotina in _todas_rotinas_permissoes()}


def _perfil_usuario_sistema(request):
    try:
        return UsuarioSistema.objects.select_related("user").filter(user=request.user).first()
    except Exception:
        return None


def _permissoes_usuario(request):
    user = getattr(request, "user", None)
    username = (getattr(user, "username", "") or "").strip().casefold()
    if not user or not getattr(user, "is_authenticated", False):
        return {}
    if username == "maia" or getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return _permissoes_padrao_total()

    perfil = _perfil_usuario_sistema(request)
    if not perfil:
        return _permissoes_vazias()
    if getattr(perfil, "administrador_geral", False):
        return _permissoes_padrao_total()

    bruto = getattr(perfil, "permissoes_json", "") or ""
    if not bruto:
        return _permissoes_vazias()
    try:
        data = json.loads(bruto) or {}
    except Exception:
        data = {}
    base = _permissoes_vazias()
    for rotina, permissoes in (data or {}).items():
        if isinstance(permissoes, dict) and rotina in base:
            for acao in ACOES_PERMISSAO:
                if acao in permissoes:
                    base[rotina][acao] = bool(permissoes.get(acao))
    return base


def _usuario_tem_permissao(request, rotina, acao="view"):
    perms = _permissoes_usuario(request)
    return bool(perms.get(rotina, {}).get(acao, False))


def _permissoes_para_ui(permissoes):
    base = _permissoes_vazias()
    origem = permissoes or {}
    for rotina in base:
        atual = origem.get(rotina, {}) if isinstance(origem.get(rotina, {}), dict) else {}
        for acao in ACOES_PERMISSAO:
            base[rotina][acao] = bool(atual.get(acao, False))
    for rotina_agregada, itens in PERMISSOES_AGREGADAS.items():
        for acao in ACOES_PERMISSAO:
            base[rotina_agregada][acao] = any(bool(origem.get(item, {}).get(acao, False)) for item in itens)
    return {rotina: base.get(rotina, {acao: False for acao in ACOES_PERMISSAO}) for rotina, _, _ in ROTINAS_PERMISSOES}


def _normalizar_permissoes_payload(payload):
    base = _permissoes_vazias()
    if not isinstance(payload, dict):
        return base
    for rotina, permissoes in payload.items():
        if not isinstance(permissoes, dict):
            continue
        destinos = [rotina]
        if rotina in PERMISSOES_AGREGADAS:
            destinos.extend(PERMISSOES_AGREGADAS[rotina])
        for destino in destinos:
            if destino not in base:
                continue
            for acao in ACOES_PERMISSAO:
                if acao in permissoes:
                    base[destino][acao] = bool(permissoes.get(acao))
    return base


def _permission_denied_response(request, rotina, acao):
    mensagem = f"Sem permissão para {acao} na rotina {rotina}."
    if _is_ajax(request):
        return JsonResponse({"error": mensagem}, status=403)
    return HttpResponse(mensagem, status=403, content_type="text/plain; charset=utf-8")


def _require_perm(rotina, acao="view"):
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            if not _usuario_tem_permissao(request, rotina, acao):
                return _permission_denied_response(request, rotina, acao)
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


def _require_lancamentos_salvar_perm(func):
    @wraps(func)
    def wrapper(request, *args, **kwargs):
        data = _json_body(request)
        acao = "edit" if data.get("id") else "create"
        if not _usuario_tem_permissao(request, "lancamentos", acao):
            return _permission_denied_response(request, "lancamentos", acao)
        return func(request, *args, **kwargs)
    return wrapper


def _filtrar_lancamentos_empresa(request, qs):
    empresa = _empresa_ativa(request)
    if empresa and hasattr(Lancamento, "empresa_id"):
        return qs.filter(empresa=empresa)
    return qs


def _fornecedor_nome(obj):
    fornecedor = getattr(obj, "fornecedor", None)
    return getattr(fornecedor, "nome", "") if fornecedor else ""


def _fornecedor_nif(obj):
    fornecedor = getattr(obj, "fornecedor", None)
    return getattr(fornecedor, "nif", "") if fornecedor else ""


def _format_date_br(value):
    return value.strftime("%d/%m/%Y") if value else ""


def _total_pago_lancamento(obj):
    total_pago = getattr(obj, "total_pago", None)
    if total_pago not in [None, ""]:
        return _to_decimal(total_pago)

    total_field = getattr(obj, "total", None)
    if total_field not in [None, ""]:
        return _to_decimal(total_field)

    dinheiro = _to_decimal(getattr(obj, "dinheiro", 0))
    cartao = _to_decimal(getattr(obj, "cartao", 0))
    transferencia = _to_decimal(getattr(obj, "transferencia", 0))
    mbway = _to_decimal(getattr(obj, "mbway", 0))
    nota_credito = _to_decimal(getattr(obj, "nota_credito", 0))
    return dinheiro + cartao + transferencia + mbway - nota_credito


def _lancamento_somente_nota_credito(obj):
    dinheiro = _to_decimal(getattr(obj, "dinheiro", 0))
    cartao = _to_decimal(getattr(obj, "cartao", 0))
    transferencia = _to_decimal(getattr(obj, "transferencia", 0))
    mbway = _to_decimal(getattr(obj, "mbway", 0))
    nota_credito = _to_decimal(getattr(obj, "nota_credito", 0))
    return nota_credito > Decimal("0.00") and all(v == Decimal("0.00") for v in [dinheiro, cartao, transferencia, mbway])


def _total_pago_exibicao_lancamento(obj):
    valor_documento = _valor_documento_lancamento(obj)
    total_pago = _total_pago_lancamento(obj)
    if _lancamento_somente_nota_credito(obj) and valor_documento > Decimal("0.00"):
        return valor_documento
    return total_pago


def _valor_pago_exibicao_baixa(item, lancamento=None):
    nota_credito = _to_decimal(getattr(item, "nota_credito", 0))
    dinheiro = _to_decimal(getattr(item, "dinheiro", 0))
    cartao = _to_decimal(getattr(item, "cartao", 0))
    transferencia = _to_decimal(getattr(item, "transferencia", 0))
    mbway = _to_decimal(getattr(item, "mbway", 0))
    total_baixa = _to_decimal(getattr(item, "total_baixa", 0))
    if lancamento is not None:
        valor_documento = _valor_documento_lancamento(lancamento)
    else:
        valor_documento = Decimal("0.00")
    somente_credito = nota_credito > Decimal("0.00") and all(v == Decimal("0.00") for v in [dinheiro, cartao, transferencia, mbway])
    if somente_credito and valor_documento > Decimal("0.00"):
        return valor_documento
    return total_baixa


def _valor_documento_lancamento(obj):
    valor_fatura = _to_decimal(getattr(obj, "valor_fatura", 0))
    if valor_fatura > 0:
        return valor_fatura
    total_field = _to_decimal(getattr(obj, "total", 0))
    total_pago = _total_pago_lancamento(obj)
    saldo_aberto = _to_decimal(getattr(obj, "saldo_aberto", 0))
    candidatos = [valor_fatura, total_field, total_pago + saldo_aberto, total_pago]
    return max(candidatos)


def _saldo_documento_lancamento(obj):
    saldo_aberto = _to_decimal(getattr(obj, "saldo_aberto", 0))
    if saldo_aberto > 0:
        return saldo_aberto
    valor_documento = _valor_documento_lancamento(obj)
    total_pago = _total_pago_lancamento(obj)
    saldo_calc = valor_documento - total_pago
    return saldo_calc if saldo_calc > 0 else Decimal("0.00")


def _seguradora_frota(obj):
    if hasattr(obj, "seguradora"):
        return getattr(obj, "seguradora", "") or ""
    if hasattr(obj, "asseguradora"):
        return getattr(obj, "asseguradora", "") or ""
    return ""


def _recalcular_status_lancamento(lancamento):
    dinheiro = _to_decimal(getattr(lancamento, "dinheiro", 0))
    cartao = _to_decimal(getattr(lancamento, "cartao", 0))
    transferencia = _to_decimal(getattr(lancamento, "transferencia", 0))
    mbway = _to_decimal(getattr(lancamento, "mbway", 0))
    nota_credito = _to_decimal(getattr(lancamento, "nota_credito", 0))
    valor_fatura = _to_decimal(getattr(lancamento, "valor_fatura", 0))

    total_pago = dinheiro + cartao + transferencia + mbway - nota_credito
    saldo_aberto = valor_fatura - total_pago
    if saldo_aberto < 0:
        saldo_aberto = Decimal("0.00")

    lancamento.total_pago = total_pago
    if hasattr(lancamento, "total"):
        lancamento.total = total_pago

    possui_credito = nota_credito > Decimal("0.00")
    possui_pagamento = any(v > Decimal("0.00") for v in [dinheiro, cartao, transferencia, mbway])
    possui_movimento = possui_pagamento or possui_credito or total_pago != Decimal("0.00") or valor_fatura > Decimal("0.00")
    possui_vencimento = bool(getattr(lancamento, "data_vencimento", None))

    if not possui_vencimento:
        lancamento.saldo_aberto = Decimal("0.00")
        lancamento.status_pagamento = "Paga" if possui_movimento else "Paga"
        if hasattr(lancamento, "data_pagamento"):
            lancamento.data_pagamento = getattr(lancamento, "data_pagamento", None) or getattr(lancamento, "data_emissao", None) or timezone.localdate()
    else:
        lancamento.saldo_aberto = saldo_aberto
        if saldo_aberto == Decimal("0.00") or (possui_credito and not possui_pagamento):
            lancamento.status_pagamento = "Paga"
            lancamento.saldo_aberto = Decimal("0.00") if (possui_credito and not possui_pagamento) else saldo_aberto
            if hasattr(lancamento, "data_pagamento"):
                lancamento.data_pagamento = getattr(lancamento, "data_pagamento", None) or timezone.localdate()
        elif total_pago != Decimal("0.00"):
            lancamento.status_pagamento = "Parcial"
            if hasattr(lancamento, "data_pagamento"):
                lancamento.data_pagamento = None
        else:
            lancamento.status_pagamento = "Em aberto"
            if hasattr(lancamento, "data_pagamento"):
                lancamento.data_pagamento = None

    return lancamento


def _linhas_financeiras(qs):
    linhas = []
    for obj in qs:
        valor_fatura = _valor_documento_lancamento(obj)
        total_pago = _total_pago_lancamento(obj)
        saldo_aberto = _saldo_documento_lancamento(obj)
        linhas.append({
            "id": obj.id,
            "lancamento_id": obj.id,
            "data_emissao": _format_date_br(getattr(obj, "data_emissao", None)),
            "data_vencimento": _format_date_br(getattr(obj, "data_vencimento", None)),
            "data_pagamento": _format_date_br(getattr(obj, "data_pagamento", None)),
            "numero_fatura": getattr(obj, "numero_fatura", "") or "",
            "fornecedor": _fornecedor_nome(obj),
            "nif": _fornecedor_nif(obj),
            "valor_fatura": float(valor_fatura),
            "dinheiro": float(_to_decimal(getattr(obj, "dinheiro", 0))),
            "cartao": float(_to_decimal(getattr(obj, "cartao", 0))),
            "transferencia": float(_to_decimal(getattr(obj, "transferencia", 0))),
            "mbway": float(_to_decimal(getattr(obj, "mbway", 0))),
            "nota_credito": float(-_to_decimal(getattr(obj, "nota_credito", 0))),
            "total_pago": float(_total_pago_exibicao_lancamento(obj)),
            "total_pago_movimento": float(total_pago),
            "saldo_aberto": float(max(Decimal("0.00"), saldo_aberto)),
            "status_pagamento": getattr(obj, "status_pagamento", "") or "",
        })
    return linhas


def _totais_financeiros(qs):
    total_valor = Decimal("0.00")
    total_pago = Decimal("0.00")
    total_saldo = Decimal("0.00")
    total_nota_credito = Decimal("0.00")

    for obj in qs:
        valor_fatura = _valor_documento_lancamento(obj)
        pago = _total_pago_exibicao_lancamento(obj)
        saldo = _saldo_documento_lancamento(obj)
        total_valor += valor_fatura
        total_pago += pago
        total_saldo += max(Decimal("0.00"), saldo)
        total_nota_credito += _to_decimal(getattr(obj, "nota_credito", 0))

    return {
        "valor_fatura": float(total_valor),
        "total_pago": float(total_pago),
        "saldo_aberto": float(total_saldo),
        "nota_credito": float(-total_nota_credito),
    }


def _faturas_alerta(request):
    hoje = date.today()
    limite = hoje + timedelta(days=20)

    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    ).filter(
        data_vencimento__isnull=False
    ).exclude(
        status_pagamento__iexact="Paga"
    ).filter(
        data_vencimento__range=[hoje, limite]
    ).order_by("data_vencimento", "id")

    dados = []
    for item in qs:
        dias = (item.data_vencimento - hoje).days if item.data_vencimento else 0
        dados.append({
            "id": item.id,
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": _fornecedor_nome(item),
            "data": item.data_vencimento.strftime("%d/%m/%Y") if item.data_vencimento else "",
            "status": item.status_pagamento or "Em aberto",
            "dias": dias,
            "valor_fatura": float(_to_decimal(getattr(item, "valor_fatura", 0))),
            "total_pago": float(_total_pago_lancamento(item)),
            "saldo_aberto": float(_to_decimal(getattr(item, "saldo_aberto", 0))),
        })
    return dados


def _faturas_vencidas(request):
    hoje = date.today()

    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    ).filter(
        data_vencimento__isnull=False,
        data_vencimento__lt=hoje
    ).exclude(
        status_pagamento__iexact="Paga"
    ).order_by("data_vencimento", "id")

    dados = []
    for item in qs:
        dias_atraso = (hoje - item.data_vencimento).days if item.data_vencimento else 0
        dados.append({
            "id": item.id,
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": _fornecedor_nome(item),
            "data": item.data_vencimento.strftime("%d/%m/%Y") if item.data_vencimento else "",
            "status": item.status_pagamento or "Em aberto",
            "dias_atraso": dias_atraso,
            "valor_fatura": float(_to_decimal(getattr(item, "valor_fatura", 0))),
            "total_pago": float(_total_pago_lancamento(item)),
            "saldo_aberto": float(_to_decimal(getattr(item, "saldo_aberto", 0))),
        })
    return dados





def _button_config_file():
    base = Path(getattr(settings, "MEDIA_ROOT", Path(__file__).resolve().parent.parent / "media"))
    base.mkdir(parents=True, exist_ok=True)
    return base / "button_configs.json"


def _load_button_configs():
    path = _button_config_file()
    if not path.exists():
        return {"combustiveis": {}, "manutencao": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8") or "{}")
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}
    data.setdefault("combustiveis", {})
    data.setdefault("manutencao", {})
    return data


def _save_button_configs(data):
    path = _button_config_file()
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_button_key(nome):
    key = slugify((nome or '').strip(), allow_unicode=True)
    return key or 'item'


def _save_button_image(uploaded_file, key, grupo='combustiveis'):
    if not uploaded_file:
        return ''
    ext = Path(getattr(uploaded_file, 'name', '')).suffix.lower() or '.png'
    if ext not in {'.png', '.jpg', '.jpeg', '.webp'}:
        ext = '.png'
    rel_dir = Path('button_images') / grupo
    abs_dir = Path(settings.MEDIA_ROOT) / rel_dir
    abs_dir.mkdir(parents=True, exist_ok=True)
    filename = f'{key}{ext}'
    abs_path = abs_dir / filename
    with abs_path.open('wb+') as destino:
        for chunk in uploaded_file.chunks():
            destino.write(chunk)
    return f'{settings.MEDIA_URL}{rel_dir.as_posix()}/{filename}'


def _combustivel_imagem_url(nome, fallback=''):
    data = _load_button_configs()
    item = data.get('combustiveis', {}).get(_normalize_button_key(nome), {})
    return item.get('imagem_url') or fallback


def _manutencao_config():
    data = _load_button_configs()
    cfg = data.get('manutencao') or {}
    tipos = cfg.get('tipos') if isinstance(cfg, dict) else None
    if not isinstance(tipos, list):
        nome = cfg.get('nome') if isinstance(cfg, dict) else ''
        tipos = [{"key": _normalize_button_key(nome or 'Manutenção'), "nome": nome or 'Manutenção'}]
    tipos_normalizados = []
    for item in tipos:
        nome = (item or {}).get('nome', '').strip() or 'Manutenção'
        tipos_normalizados.append({
            'key': (item or {}).get('key') or _normalize_button_key(nome),
            'nome': nome,
        })
    if not tipos_normalizados:
        tipos_normalizados = [{'key': 'manutencao', 'nome': 'Manutenção'}]
    return {'tipos': tipos_normalizados}


def _ensure_combustiveis_padrao(request=None):
    qs = _filtrar_empresa_generico(request, Combustivel.objects.all(), model=Combustivel) if request is not None else Combustivel.objects.all()
    return [
        {**item, "imagem_url": _combustivel_imagem_url(item.get("nome", ""))}
        for item in qs.values("id", "nome").order_by("nome")
    ]

def _frota_alerta(request):
    hoje = date.today()
    limite = hoje + timedelta(days=30)

    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota).order_by("matricula")
    dados = []

    for item in qs:
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)

        dias_seguro = (seguro - hoje).days if seguro else None
        dias_inspecao = (inspecao - hoje).days if inspecao else None

        status_seguro = "OK"
        if seguro:
            if seguro < hoje:
                status_seguro = "Vencido"
            elif seguro <= limite:
                status_seguro = "A vencer"

        status_inspecao = "OK"
        if inspecao:
            if inspecao < hoje:
                status_inspecao = "Vencido"
            elif inspecao <= limite:
                status_inspecao = "A vencer"

        status_geral = "OK"
        if status_seguro == "Vencido" or status_inspecao == "Vencido":
            status_geral = "Vencido"
        elif status_seguro == "A vencer" or status_inspecao == "A vencer":
            status_geral = "A vencer"

        dados.append({
            "id": item.id,
            "matricula": getattr(item, "matricula", "") or "",
            "seguradora": _seguradora_frota(item),
            "seguro": seguro.strftime("%d/%m/%Y") if seguro else "",
            "inspecao": inspecao.strftime("%d/%m/%Y") if inspecao else "",
            "dias_seguro": dias_seguro,
            "dias_inspecao": dias_inspecao,
            "status_seguro": status_seguro,
            "status_inspecao": status_inspecao,
            "status": status_geral,
            "css_seguro": "danger" if status_seguro in ["Vencido", "A vencer"] else "success",
            "css_inspecao": "danger" if status_inspecao in ["Vencido", "A vencer"] else "success",
            "css_status": "danger" if status_geral in ["Vencido", "A vencer"] else "success",
        })

    return dados


def _km_rodados_frota(frota):
    if not frota:
        return 0
    total = 0
    try:
        if ItemCombustivel is not None:
            total += sum(int(getattr(item, "km_total", 0) or 0) for item in ItemCombustivel.objects.filter(frota=frota))
    except Exception:
        pass
    try:
        if ManutencaoFrota is not None:
            total += sum(int(getattr(item, "km_total", 0) or 0) for item in ManutencaoFrota.objects.filter(frota=frota))
    except Exception:
        pass
    return total


def _dashboard_revisao(request=None):
    linhas = []
    qs = RevisaoFrota.objects.select_related("frota").order_by("data_ultima_revisao", "id")
    empresa = _empresa_ativa(request) if request is not None else None
    if empresa:
        qs = qs.filter(frota__empresa=empresa)
    for item in qs[:8]:
        linhas.append({
            "matricula": getattr(getattr(item, "frota", None), "matricula", "") or "-",
            "marca": getattr(getattr(item, "frota", None), "marca", "") or "-",
            "modelo": getattr(getattr(item, "frota", None), "modelo", "") or "-",
            "data_ultima_revisao": item.data_ultima_revisao.strftime("%d/%m/%Y") if getattr(item, "data_ultima_revisao", None) else "-",
            "km_ultima_revisao": int(getattr(item, "km_ultima_revisao", 0) or 0),
            "km_rodados": int(getattr(item, "km_rodados", 0) or 0),
            "kms_previsao": int(getattr(item, "kms_previsao", 0) or 0),
            "km_para_revisao": int(getattr(item, "km_para_revisao", 0) or 0),
        })
    return linhas


def _render_dashboard(request, secao_ativa="dashboard"):
    config = _configuracao_sistema()
    context = {
        "secao_ativa": secao_ativa.replace("-section", ""),
        "empresa_ativa": _empresa_ativa(request),
        "empresas_usuario": _empresas_usuario(request),
        "alertas_faturas": _faturas_alerta(request),
        "faturas_vencidas": _faturas_vencidas(request),
        "alertas_frota": _frota_alerta(request),
        "combustiveis_menu": json.dumps(_ensure_combustiveis_padrao(request), ensure_ascii=False),
        "empresas_json": json.dumps(
            list(_empresas_usuario(request).values("id", "nome"))
        ),
        "frota_json": json.dumps(
            list(_filtrar_empresa_generico(request, Frota.objects.all(), model=Frota).values("id", "matricula", "marca", "modelo").order_by("matricula"))
        ),
        "permissoes_usuario_json": json.dumps(_permissoes_usuario(request), ensure_ascii=False),
        "rotinas_permissoes_json": json.dumps([{"key": key, "grupo": grupo, "label": label} for key, grupo, label in ROTINAS_PERMISSOES], ensure_ascii=False),
        "manutencao_botao_json": json.dumps(_manutencao_config(), ensure_ascii=False),
        "manutencao_tipos_json": json.dumps(_manutencao_config().get("tipos", []), ensure_ascii=False),
        "dashboard_revisao": _dashboard_revisao(request),
        "sistema_nome": getattr(config, "nome_sistema", "Elion Gestão") if config else "Elion Gestão",
        "logo_sistema_url": getattr(getattr(config, "logo_login", None), "url", "") if config and getattr(config, "logo_login", None) else "",
        "logo_paineis_url": getattr(getattr(config, "logo_paineis", None), "url", "") if config and getattr(config, "logo_paineis", None) else "",
    }
    return render(request, "Elion/dashboard.html", context)


def login_view(request):
    if request.user.is_authenticated:
        return redirect("/painel/")

    if request.method == "POST":
        username = (
            request.POST.get("username")
            or request.POST.get("nome de usuário")
            or request.POST.get("usuario")
            or request.POST.get("email")
            or ""
        ).strip()
        password = (request.POST.get("password") or request.POST.get("senha") or "").strip()

        user = None
        tentativas = []
        if username:
            tentativas.extend([username, username.lower(), username.upper(), username.title()])
            tentativas.extend(
                list(
                    User.objects.filter(
                        Q(username__iexact=username) | Q(email__iexact=username)
                    ).values_list("username", flat=True)
                )
            )
            try:
                perfis = UsuarioSistema.objects.filter(
                    Q(nome__iexact=username)
                    | Q(email_recuperacao__iexact=username)
                    | Q(contato__iexact=username)
                ).select_related("user")
                tentativas.extend([item.user.username for item in perfis if getattr(item, "user", None)])
            except Exception:
                pass

        vistos = set()
        for candidato in tentativas:
            if not candidato or candidato in vistos:
                continue
            vistos.add(candidato)
            user = authenticate(request, username=candidato, password=password)
            if user:
                break

        if user and getattr(user, "is_active", False):
            login(request, user)
            empresa = _empresa_ativa(request)
            if empresa:
                request.session["empresa_ativa_id"] = empresa.id
            return redirect(request.GET.get("next") or "/painel/")
        messages.error(request, "Utilizador ou senha inválidos.")

    config = _configuracao_sistema()
    return render(request, "Elion/login.html", {
        "sistema_nome": getattr(config, "nome_sistema", "Elion Gestão") if config else "Elion Gestão",
        "sistema_subtitulo": getattr(config, "subtitulo_login", "Faturação, financeiro, frota e combustível num único sistema") if config else "Faturação, financeiro, frota e combustível num único sistema",
        "logo_login": getattr(getattr(config, "logo_login", None), "url", "") if config and getattr(config, "logo_login", None) else "",
    })


@login_required
def logout_view(request):
    logout(request)
    return redirect("/")




BACKUP_MODEL_KEYS = [
    "Empresa", "Fornecedor", "Funcionario", "Frota", "Combustivel", "Lancamento",
    "BaixaFatura", "Cliente", "Vendedor", "CategoriaArtigo", "Armazem", "Artigo",
    "SerieDocumento", "TemplateDocumento", "CentroCusto", "NivelAprovacao",
    "DocumentoVenda", "DocumentoVendaLinha", "RecebimentoCliente", "ConfiguracaoFiscal",
    "RevisaoFrota", "LancamentoCombustivel", "ManutencaoFrota", "UsuarioSistema",
]


def _serialize_backup_queryset(model_label, queryset):
    rows = []
    if queryset is None:
        return rows
    for obj in queryset:
        field_names = []
        for field in obj._meta.fields:
            if field.auto_created or field.name == "id":
                continue
            field_names.append(field.name)
        rows.append({
            "model": model_label,
            "pk": obj.pk,
            "fields": model_to_dict(obj, fields=field_names),
        })
    return rows


def _backup_payload_empresa(request):
    empresa = _empresa_ativa(request)
    payload = {
        "meta": {
            "empresa_id": getattr(empresa, "id", None),
            "empresa_nome": getattr(empresa, "nome", "") or "",
            "gerado_em": timezone.localtime().isoformat(),
        },
        "modelos": {key: [] for key in BACKUP_MODEL_KEYS},
    }

    empresas_qs = Empresa.objects.filter(id=empresa.id) if empresa else Empresa.objects.all()
    payload["modelos"]["Empresa"] = _serialize_backup_queryset("Elion.empresa", empresas_qs)
    payload["modelos"]["Fornecedor"] = _serialize_backup_queryset("Elion.fornecedor", _fornecedores_empresa_qs(request))
    payload["modelos"]["Funcionario"] = _serialize_backup_queryset("Elion.funcionario", _filtrar_empresa_generico(request, Funcionario.objects.all(), model=Funcionario))
    payload["modelos"]["Frota"] = _serialize_backup_queryset("Elion.frota", _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota))
    payload["modelos"]["Combustivel"] = _serialize_backup_queryset("Elion.combustivel", _filtrar_empresa_generico(request, Combustivel.objects.all(), model=Combustivel))
    payload["modelos"]["Lancamento"] = _serialize_backup_queryset("Elion.lancamento", _lancamentos_empresa_qs(request))
    payload["modelos"]["BaixaFatura"] = _serialize_backup_queryset("Elion.baixafatura", _filtrar_empresa_generico(request, BaixaFatura.objects.select_related("lancamento"), model=BaixaFatura))
    payload["modelos"]["ConfiguracaoFiscal"] = _serialize_backup_queryset("Elion.configuracaofiscal", _filtrar_empresa_generico(request, ConfiguracaoFiscal.objects.all(), model=ConfiguracaoFiscal))
    payload["modelos"]["RevisaoFrota"] = _serialize_backup_queryset("Elion.revisaofrota", RevisaoFrota.objects.filter(frota__empresa=empresa) if empresa else RevisaoFrota.objects.all())
    payload["modelos"]["LancamentoCombustivel"] = _serialize_backup_queryset("Elion.lancamentocombustivel", ItemCombustivel.objects.filter(lancamento__empresa=empresa) if (empresa and ItemCombustivel is not None) else (ItemCombustivel.objects.all() if ItemCombustivel is not None else None))
    payload["modelos"]["ManutencaoFrota"] = _serialize_backup_queryset("Elion.manutencaofrota", ManutencaoFrota.objects.filter(lancamento__empresa=empresa) if (empresa and ManutencaoFrota is not None) else (ManutencaoFrota.objects.all() if ManutencaoFrota is not None else None))
    payload["modelos"]["UsuarioSistema"] = _serialize_backup_queryset("Elion.usuariosistema", _filtrar_empresa_generico(request, UsuarioSistema.objects.select_related("user"), model=UsuarioSistema))
    return payload


def _merge_backup_payload(payload):
    inserted = {
        "Empresa": 0, "Fornecedor": 0, "Funcionario": 0, "Frota": 0, "Combustivel": 0,
        "Lancamento": 0, "BaixaFatura": 0, "ConfiguracaoFiscal": 0,
        "RevisaoFrota": 0, "LancamentoCombustivel": 0, "ManutencaoFrota": 0, "UsuarioSistema": 0,
    }
    model_map = {
        "Empresa": Empresa,
        "Fornecedor": Fornecedor,
        "Funcionario": Funcionario,
        "Frota": Frota,
        "Combustivel": Combustivel,
        "Lancamento": Lancamento,
        "BaixaFatura": BaixaFatura,
        "ConfiguracaoFiscal": ConfiguracaoFiscal,
        "RevisaoFrota": RevisaoFrota,
        "LancamentoCombustivel": ItemCombustivel,
        "ManutencaoFrota": ManutencaoFrota,
        "UsuarioSistema": UsuarioSistema,
    }
    order = ["Empresa", "Fornecedor", "Funcionario", "Frota", "Combustivel", "ConfiguracaoFiscal", "Lancamento", "BaixaFatura", "RevisaoFrota", "LancamentoCombustivel", "ManutencaoFrota", "UsuarioSistema"]
    for key in order:
        model = model_map.get(key)
        if model is None:
            continue
        for item in ((payload.get("modelos") or {}).get(key) or []):
            pk = item.get("pk")
            if pk and model.objects.filter(pk=pk).exists():
                continue
            fields = dict(item.get("fields") or {})
            if key == "Lancamento" and "nota_credito" not in fields:
                fields["nota_credito"] = "0.00"
            fields.pop("empresas", None)
            obj = model(pk=pk, **fields)
            obj.save(force_insert=True)
            inserted[key] += 1
    return inserted


def configuracoes_sistema_view(request):
    config = _configuracao_sistema() or ConfiguracaoSistema()
    if request.method == "POST":
        config.nome_sistema = (request.POST.get("nome_sistema") or "Elion Gestão").strip()
        config.subtitulo_login = (request.POST.get("subtitulo_login") or "").strip()
        config.smtp_host = (request.POST.get("smtp_host") or "").strip()
        config.smtp_porta = int(request.POST.get("smtp_porta") or 587)
        config.smtp_user = (request.POST.get("smtp_user") or "").strip()
        config.smtp_password = (request.POST.get("smtp_password") or "").strip()
        config.sms_provider = (request.POST.get("sms_provider") or "").strip()
        config.sms_token = (request.POST.get("sms_token") or "").strip()
        if request.POST.get("remover_logo_login") == "1":
            config.logo_login = None
        if request.FILES.get("logo_login"):
            config.logo_login = request.FILES["logo_login"]
        if request.FILES.get("logo_paineis"):
            config.logo_paineis = request.FILES["logo_paineis"]
        config.save()
        messages.success(request, "Configurações do sistema guardadas com sucesso.")
        return redirect("configuracoes_sistema")
    empresa = _empresa_ativa(request)
    return render(request, "Elion/configuracoes_sistema.html", {"config": config, "empresa_ativa": empresa})


@login_required
def exportar_backup_empresa(request):
    empresa = _empresa_ativa(request)
    nome = (getattr(empresa, "nome", "empresa") or "empresa").strip().lower()
    nome = re.sub(r"[^a-z0-9]+", "-", nome).strip("-") or "empresa"
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    payload = _backup_payload_empresa(request)
    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="backup_{nome}_{stamp}.zip"'
    with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        with ZipFile(tmp_path, "w") as zf:
            zf.writestr("README_BACKUP.txt", "Backup por empresa para importação por mesclagem. IDs já existentes são ignorados e só entra o que falta.")
            zf.writestr(f"backup_{nome}_{stamp}.json", json.dumps(payload, ensure_ascii=False, indent=2, default=str))
        response.write(tmp_path.read_bytes())
    finally:
        tmp_path.unlink(missing_ok=True)
    return response


@login_required
@require_POST
@transaction.atomic
def importar_backup_empresa(request):
    ficheiro = request.FILES.get("backup_file")
    if not ficheiro:
        messages.error(request, "Selecione um ficheiro de backup .zip ou .json.")
        return redirect("configuracoes_sistema")

    suffix = Path(ficheiro.name).suffix.lower()
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        uploaded = tmpdir / ficheiro.name
        with uploaded.open("wb") as dest:
            for chunk in ficheiro.chunks():
                dest.write(chunk)

        payload = None
        if suffix == ".zip":
            with ZipFile(uploaded, "r") as zf:
                json_members = [n for n in zf.namelist() if n.lower().endswith(".json")]
                if not json_members:
                    messages.error(request, "O ZIP não contém um backup JSON compatível para mesclagem.")
                    return redirect("configuracoes_sistema")
                payload = json.loads(zf.read(json_members[0]))
        elif suffix == ".json":
            payload = json.loads(uploaded.read_text(encoding="utf-8"))
        else:
            messages.error(request, "Formato inválido. Use .zip ou .json exportado pelo sistema.")
            return redirect("configuracoes_sistema")

        inserted = _merge_backup_payload(payload)

    total_inseridos = sum(inserted.values())
    messages.success(request, f"Backup mesclado com sucesso. Registos novos importados: {total_inseridos}. Lançamentos novos: {inserted.get('Lancamento', 0)}.")
    return redirect("configuracoes_sistema")

    suffix = Path(ficheiro.name).suffix.lower()
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        uploaded = tmpdir / ficheiro.name
        with uploaded.open("wb") as dest:
            for chunk in ficheiro.chunks():
                dest.write(chunk)

        source_db = None
        if suffix == ".zip":
            with ZipFile(uploaded, "r") as zf:
                sqlite_members = [n for n in zf.namelist() if n.lower().endswith(".sqlite3") or n.lower().endswith("db.sqlite3")]
                if not sqlite_members:
                    messages.error(request, "O ZIP não contém um ficheiro db.sqlite3.")
                    return redirect("configuracoes_sistema")
                member = sqlite_members[0]
                zf.extract(member, path=tmpdir)
                source_db = tmpdir / member
        elif suffix == ".sqlite3":
            source_db = uploaded
        else:
            messages.error(request, "Formato inválido. Use .zip ou .sqlite3.")
            return redirect("configuracoes_sistema")

        db_path = Path(settings.DATABASES["default"]["NAME"])
        if not db_path.is_absolute():
            db_path = Path(settings.BASE_DIR) / db_path
        backup_path = db_path.with_name(f"{db_path.stem}_antes_importacao_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}{db_path.suffix}")
        connections.close_all()
        shutil.copy2(db_path, backup_path)
        shutil.copy2(source_db, db_path)

    messages.success(request, f"Backup importado com sucesso. Cópia anterior guardada em {backup_path.name}.")
    return redirect("configuracoes_sistema")


@login_required
def dashboard_view(request):
    secao = (request.GET.get("secao") or "dashboard").strip()
    return _render_dashboard(request, secao)


@login_required
@require_POST
def trocar_empresa_ativa(request):
    data = _json_body(request)
    empresa_id = data.get("empresa_id")
    if empresa_id:
        request.session["empresa_ativa_id"] = str(empresa_id)
    return JsonResponse({"ok": True})


@login_required
@require_GET
def fornecedores_list(request):
    linhas = list(_fornecedores_empresa_qs(request).values("id", "nif", "nome", "iban", "contato", "responsavel", "email", "morada", "conselho", "caixa_postal").order_by("nome"))
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def fornecedores_salvar(request):
    data = _json_body(request)
    obj = _get_fornecedor_empresa_or_404(request, data["id"]) if data.get("id") else Fornecedor()
    empresa = _empresa_ativa(request)
    if hasattr(obj, "empresa") and empresa:
        obj.empresa = empresa
    obj.nif = (data.get("nif", "") or "").strip()
    obj.nome = (data.get("nome", "") or "").strip()
    obj.iban = (data.get("iban", "") or "").strip()
    obj.contato = (data.get("contato", "") or "").strip()
    obj.responsavel = (data.get("responsavel", "") or "").strip()
    obj.email = (data.get("email", "") or "").strip() or None
    obj.morada = (data.get("morada", "") or "").strip()
    obj.conselho = (data.get("conselho", "") or "").strip()
    obj.caixa_postal = (data.get("caixa_postal", "") or "").strip()
    erro = _save_model_or_error(obj, "Já existe um fornecedor com este contribuinte.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return JsonResponse({"ok": True, "id": obj.id, "nome": obj.nome, "nif": obj.nif, "iban": obj.iban, "contato": obj.contato, "responsavel": obj.responsavel, "email": obj.email or "", "morada": obj.morada, "conselho": obj.conselho, "caixa_postal": obj.caixa_postal})


@login_required
@require_GET
def fornecedores_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = _fornecedores_empresa_qs(request)

    # Por omissão a pesquisa devolve TODOS os fornecedores correspondentes
    # (sem limitar/ocultar registos). Só aplicamos um limite quando o
    # chamador pede explicitamente (ex.: listas de sugestões/autocomplete),
    # através do parâmetro "limit".
    limit_param = (request.GET.get("limit") or "").strip()
    limite = int(limit_param) if limit_param.isdigit() else None

    if q:
        normalizado = _normalizar_nif_busca(q)
        campos = ("id", "nome", "nif", "contato", "responsavel", "email", "iban", "morada", "conselho", "caixa_postal")
        filtro = (
            Q(nome__icontains=q)
            | Q(nif__icontains=q)
            | Q(contato__icontains=q)
            | Q(responsavel__icontains=q)
            | Q(email__icontains=q)
            | Q(iban__icontains=q)
            | Q(morada__icontains=q)
            | Q(conselho__icontains=q)
            | Q(caixa_postal__icontains=q)
        )
        if q.isdigit():
            filtro |= Q(id=int(q))
        base_qs = qs.filter(filtro).order_by("nome").values(*campos)
        linhas = list(base_qs[:limite] if limite else base_qs)

        if normalizado:
            vistos = {item["id"] for item in linhas}
            for item in qs.order_by("nome").values(*campos):
                if item["id"] in vistos:
                    continue
                if normalizado in _normalizar_nif_busca(item.get("nif", "")):
                    linhas.append(item)
                    vistos.add(item["id"])
                if limite and len(linhas) >= limite:
                    break

        termo = q.casefold()
        linhas.sort(key=lambda item: (
            0 if ((item.get("nome", "").casefold() == termo) or (item.get("nif", "").casefold() == termo) or (_normalizar_nif_busca(item.get("nif", "")) == normalizado and normalizado)) else 1,
            item.get("nome", "").casefold(),
        ))
    else:
        campos = ("id", "nome", "nif", "contato", "responsavel", "email", "iban", "morada", "conselho", "caixa_postal")
        base_qs = qs.order_by("nome").values(*campos)
        linhas = list(base_qs[:limite] if limite else base_qs)

    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def funcionarios_list(request):
    linhas = list(
        _filtrar_empresa_generico(request, Funcionario.objects.all(), model=Funcionario).values("id", "nome", "contato", "email").order_by("nome")
    )
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def funcionarios_salvar(request):
    data = _json_body(request)
    obj = _get_objeto_empresa_or_404(request, Funcionario, data["id"]) if data.get("id") else Funcionario()
    _atribuir_empresa_ativa(request, obj)
    obj.nome = data.get("nome", "")
    obj.contato = data.get("contato", "")
    obj.email = data.get("email", "")
    erro = _save_model_or_error(obj)
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def frota_list(request):
    linhas = []
    for item in _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota).order_by("matricula"):
        linhas.append({
            "id": item.id,
            "matricula": getattr(item, "matricula", "") or "",
            "marca": getattr(item, "marca", "") or "",
            "modelo": getattr(item, "modelo", "") or "",
            "seguro": item.seguro.strftime("%Y-%m-%d") if getattr(item, "seguro", None) else "",
            "seguradora": _seguradora_frota(item),
            "inspecao": item.inspecao.strftime("%Y-%m-%d") if getattr(item, "inspecao", None) else "",
        })
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def frota_salvar(request):
    data = _json_body(request)
    obj = _get_objeto_empresa_or_404(request, Frota, data["id"]) if data.get("id") else Frota()
    _atribuir_empresa_ativa(request, obj)
    obj.matricula = data.get("matricula", "")
    obj.marca = data.get("marca", "")
    obj.modelo = data.get("modelo", "")
    obj.seguro = _to_date(data.get("seguro"))
    obj.inspecao = _to_date(data.get("inspecao"))
    if hasattr(obj, "seguradora"):
        obj.seguradora = data.get("seguradora", "")
    elif hasattr(obj, "asseguradora"):
        obj.asseguradora = data.get("seguradora", "")
    erro = _save_model_or_error(obj, "Já existe uma viatura com esta matrícula.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return JsonResponse({"ok": True, "id": obj.id, "marca": obj.marca, "modelo": obj.modelo})


@login_required
@require_GET
def frota_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)

    if q:
        qs = qs.filter(Q(matricula__icontains=q) | Q(marca__icontains=q) | Q(modelo__icontains=q))

    linhas = list(qs.order_by("matricula").values("id", "matricula", "marca", "modelo")[:20])
    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def combustiveis_list(request):
    linhas = list(_filtrar_empresa_generico(request, Combustivel.objects.all(), model=Combustivel).order_by("nome").values("id", "nome"))
    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def combustiveis_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = _filtrar_empresa_generico(request, Combustivel.objects.all(), model=Combustivel)
    if q:
        qs = qs.filter(nome__icontains=q)
    linhas = list(qs.order_by("nome").values("id", "nome")[:20])
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def combustiveis_salvar(request):
    data = request.POST if request.POST else _json_body(request)
    obj = _get_objeto_empresa_or_404(request, Combustivel, data["id"]) if data.get("id") else Combustivel()
    _atribuir_empresa_ativa(request, obj)
    obj.nome = (data.get("nome") or "").strip()
    erro = _save_model_or_error(obj, "Já existe um combustível com esse nome.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return JsonResponse({"ok": True, "id": obj.id, "nome": obj.nome})




@login_required
@require_GET
def manutencao_botao_config(request):
    cfg = _manutencao_config()
    return JsonResponse({"tipos": cfg.get("tipos", [])})


@login_required
@require_POST
def manutencao_botao_salvar(request):
    data = request.POST if request.POST else _json_body(request)
    nome = (data.get('nome') or '').strip() or 'Manutenção'
    key = (data.get('key') or '').strip() or _normalize_button_key(nome)
    configs = _load_button_configs()
    cfg = _manutencao_config()
    tipos = [item for item in cfg.get('tipos', []) if item.get('key') != key]
    tipos.append({'key': key, 'nome': nome})
    tipos = sorted(tipos, key=lambda item: (item.get('nome') or '').lower())
    configs['manutencao'] = {'tipos': tipos}
    _save_button_configs(configs)
    return JsonResponse({"ok": True, "tipos": tipos, "key": key, "nome": nome})


@login_required
@require_POST
def manutencao_botao_excluir(request, key):
    configs = _load_button_configs()
    cfg = _manutencao_config()
    tipos = [item for item in cfg.get('tipos', []) if item.get('key') != key]
    if not tipos:
        tipos = [{'key': 'manutencao', 'nome': 'Manutenção'}]
    configs['manutencao'] = {'tipos': tipos}
    _save_button_configs(configs)
    return JsonResponse({'ok': True, 'tipos': tipos})

def _criar_estrutura_inicial_empresa(empresa):
    if not empresa:
        return
    try:
        if _table_exists("Elion_configuracaofiscal"):
            ConfiguracaoFiscal.objects.get_or_create(empresa=empresa)
    except Exception:
        pass

    if _table_exists("Elion_seriedocumento"):
        defaults_series = [
            ("FT", "Faturas"),
            ("FS", "Faturas Simplificadas"),
            ("FR", "Faturas-Recibo"),
            ("OR", "Orçamentos"),
        ]
        for codigo, descricao in defaults_series:
            try:
                SerieDocumento.objects.get_or_create(empresa=empresa, codigo=codigo, ano=timezone.localdate().year, defaults={
                    "descricao": descricao,
                    "tipo_documento": codigo,
                    "prefixo_documento": codigo,
                    "usar_ano": True,
                    "separador": "/",
                    "casas_numero": 4,
                    "proximo_numero": 1,
                    "ativa": True,
                })
            except Exception:
                continue


@login_required
@require_GET
def empresas_list(request):
    linhas = list(
        Empresa.objects.values("id", "nif", "nome", "cidade", "contato").order_by("nome")
    )
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def empresas_salvar(request):
    obj = get_object_or_404(Empresa, id=request.POST.get("id")) if request.POST.get("id") else Empresa()
    obj.nif = request.POST.get("nif", "")
    obj.nome = request.POST.get("nome", "")
    obj.morada = request.POST.get("morada", "")
    obj.caixa_postal = request.POST.get("caixa_postal", "")
    obj.cidade = request.POST.get("cidade", "")
    obj.contato = request.POST.get("contato", "")
    obj.email = request.POST.get("email", "")
    if request.POST.get("remover_logo") == "1":
        obj.logo = None
    if request.FILES.get("logo"):
        obj.logo = request.FILES["logo"]
    empresa_nova = obj.pk is None
    erro = _save_model_or_error(obj, "Já existe uma empresa com este contribuinte.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    if empresa_nova:
        _criar_estrutura_inicial_empresa(obj)
    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def usuarios_list(request):
    linhas = []
    for item in UsuarioSistema.objects.select_related("user").all().order_by("nome"):
        linhas.append({
            "id": item.id,
            "nome": getattr(item, "nome", ""),
            "contato": getattr(item, "contato", ""),
            "email_recuperacao": getattr(item, "email_recuperacao", ""),
            "username": item.user.username if getattr(item, "user", None) else "",
            "empresa": item.empresas.first().id if hasattr(item, "empresas") and item.empresas.exists() else "",
            "empresa_nome": item.empresas.first().nome if hasattr(item, "empresas") and item.empresas.exists() else "",
            "administrador_geral": bool(getattr(item, "administrador_geral", False)),
            "permissoes": _permissoes_para_ui(_normalizar_permissoes_payload(getattr(item, "get_permissoes", lambda: {})())),
        })
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def usuarios_salvar(request):
    data = _json_body(request)
    try:
        if data.get("id"):
            perfil = get_object_or_404(UsuarioSistema, id=data["id"])
            user = perfil.user
        else:
            user = User()
            perfil = None
        username = (data.get("username") or "").strip()
        if not username:
            return JsonResponse({"error": "Informe o utilizador."}, status=400)
        existente = User.objects.filter(username__iexact=username)
        if user.pk:
            existente = existente.exclude(pk=user.pk)
        if existente.exists():
            return JsonResponse({"error": "Já existe um utilizador com esse nome."}, status=400)
        user.username = username
        if data.get("password"):
            user.set_password(data["password"])
        elif not user.pk:
            user.set_password("123456")
        username_norm = username.casefold()
        administrador_geral = True if username_norm == "maia" else bool(data.get("administrador_geral"))
        user.is_staff = administrador_geral or bool(getattr(user, "is_superuser", False))
        if username_norm == "maia":
            user.is_superuser = True
        user.save()
        if perfil is None:
            perfil, _ = UsuarioSistema.objects.get_or_create(user=user)
        perfil.nome = data.get("nome", "")
        perfil.contato = data.get("contato", "")
        perfil.email_recuperacao = data.get("email_recuperacao", "") or None
        perfil.administrador_geral = True if username_norm == "maia" else administrador_geral
        username_norm = username.casefold()
        if perfil.administrador_geral or username_norm == "maia":
            perfil.permissoes_json = json.dumps(_permissoes_padrao_total(), ensure_ascii=False)
        else:
            perfil.permissoes_json = json.dumps(_normalizar_permissoes_payload(data.get("permissoes") or {}), ensure_ascii=False)
        erro = _save_model_or_error(perfil)
        if erro:
            return JsonResponse({"error": erro}, status=400)
        if hasattr(perfil, "empresas"):
            if perfil.administrador_geral:
                perfil.empresas.set(Empresa.objects.all())
            else:
                empresa_id = data.get("empresa")
                if empresa_id:
                    perfil.empresas.set(Empresa.objects.filter(id=empresa_id))
                else:
                    perfil.empresas.clear()
        elif hasattr(perfil, "empresa_id"):
            perfil.empresa_id = data.get("empresa") or None
            erro = _save_model_or_error(perfil)
            if erro:
                return JsonResponse({"error": erro}, status=400)
        return JsonResponse({"ok": True, "id": perfil.id})
    except Exception as exc:
        return JsonResponse({"error": str(exc) or "Erro ao guardar utilizador."}, status=400)


@login_required
@require_POST
def fornecedor_excluir(request, id):
    _get_fornecedor_empresa_or_404(request, id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def funcionario_excluir(request, id):
    _get_objeto_empresa_or_404(request, Funcionario, id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def frota_excluir(request, id):
    _get_objeto_empresa_or_404(request, Frota, id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def combustivel_excluir(request, id):
    _get_objeto_empresa_or_404(request, Combustivel, id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def empresa_excluir(request, id):
    get_object_or_404(Empresa, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def usuario_excluir(request, id):
    perfil = get_object_or_404(UsuarioSistema, id=id)
    user = getattr(perfil, "user", None)
    if user and (user.username or "").strip().casefold() == "maia":
        return JsonResponse({"error": "O utilizador Maia é protegido e não pode ser excluído."}, status=403)
    perfil.delete()
    if user:
        user.delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def usuario_alterar_senha(request, id=None):
    data = _json_body(request)
    alvo = request.user

    if id:
        perfil = get_object_or_404(UsuarioSistema, id=id)
        alvo = perfil.user
        if request.user != alvo and not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({"error": "Sem permissão para alterar a senha deste utilizador."}, status=403)

    senha = (data.get("password") or "").strip()
    if not senha:
        return JsonResponse({"error": "Informe a nova senha."}, status=400)

    alvo.set_password(senha)
    alvo.save(update_fields=["password"])
    return JsonResponse({"ok": True})


@login_required
@require_POST
@transaction.atomic
def lancamentos_salvar(request):
    data = _json_body(request)

    if data.get("id"):
        obj = _get_lancamento_empresa_or_404(request, data["id"])
    else:
        obj = Lancamento()

    empresa = _empresa_ativa(request)
    if hasattr(obj, "empresa") and empresa:
        obj.empresa = empresa

    fornecedor_id = data.get("fornecedor_id")
    fornecedor_busca = (data.get("fornecedor_busca") or "").strip()
    fornecedor_obj = _fornecedores_empresa_qs(request).filter(id=fornecedor_id).first() if fornecedor_id else None
    if not fornecedor_obj and fornecedor_busca:
        fornecedor_obj = _resolver_fornecedor_por_busca(fornecedor_busca, request=request)
    if not fornecedor_obj:
        return JsonResponse({"error": "Selecione um fornecedor válido antes de salvar."}, status=400)
    obj.fornecedor = fornecedor_obj

    obj.data_emissao = _to_date(data.get("data_emissao")) or date.today()
    if not (2000 <= obj.data_emissao.year <= 2100):
        return JsonResponse({"error": f"Data de emissão inválida ({obj.data_emissao.strftime('%d/%m/%Y')}). Verifique o ano digitado."}, status=400)
    obj.numero_fatura = data.get("numero_fatura", "")
    obj.dinheiro = _to_decimal(data.get("dinheiro"))
    obj.cartao = _to_decimal(data.get("cartao"))
    obj.valor_fatura = _to_decimal(data.get("valor_fatura"))
    if hasattr(obj, "nota_credito"):
        obj.nota_credito = _to_decimal(data.get("nota_credito"))
    obj.data_vencimento = _to_date(data.get("data_vencimento"))
    if obj.data_vencimento and not (2000 <= obj.data_vencimento.year <= 2100):
        return JsonResponse({"error": f"Data de vencimento inválida ({obj.data_vencimento.strftime('%d/%m/%Y')}). Verifique o ano digitado."}, status=400)

    if hasattr(obj, "parcelas"):
        try:
            obj.parcelas = int(data.get("parcelas") or 1)
        except Exception:
            obj.parcelas = 1

    if hasattr(obj, "valor_parcela"):
        obj.valor_parcela = _to_decimal(data.get("valor_parcela"))

    if hasattr(obj, "transferencia"):
        obj.transferencia = _to_decimal(data.get("transferencia") or "0")

    if hasattr(obj, "mbway"):
        obj.mbway = _to_decimal(data.get("mbway") or "0")

    rotina_repetivel = bool(data.get("itens_combustivel")) or bool(data.get("itens_manutencao")) or bool(data.get("itens_revisao"))
    if not rotina_repetivel and obj.numero_fatura and obj.data_emissao and obj.fornecedor_id:
        duplicado_qs = _lancamentos_empresa_qs(request).filter(
            numero_fatura=obj.numero_fatura,
            data_emissao=obj.data_emissao,
            fornecedor_id=obj.fornecedor_id,
        )
        if obj.pk:
            duplicado_qs = duplicado_qs.exclude(pk=obj.pk)
        if duplicado_qs.exists():
            return JsonResponse({"error": "Já existe um lançamento com este número de fatura, fornecedor e data."}, status=400)

    _recalcular_status_lancamento(obj)
    erro = _save_model_or_error(obj)
    if erro:
        return JsonResponse({"error": erro}, status=400)

    if ItemCombustivel is not None:
        ItemCombustivel.objects.filter(lancamento=obj).delete()
        for item in data.get("itens_combustivel", []):
            combustivel_id = item.get("combustivel_id") or None
            if not combustivel_id:
                continue
            frota_id = item.get("frota_id") or None
            if not frota_id and item.get("matricula"):
                frota = _resolver_frota_por_matricula(item.get("matricula"), request=request)
                frota_id = getattr(frota, "id", None)
            ItemCombustivel.objects.create(
                lancamento=obj,
                combustivel_id=combustivel_id,
                frota_id=frota_id,
                km_inicio=_to_decimal(item.get("km_inicio")),
                km_final=_to_decimal(item.get("km_final")),
                km_total=_to_decimal(item.get("km_total")),
                litro=_to_decimal(item.get("litro")),
                valor_litro=_to_decimal(item.get("valor_litro")),
                valor_total=_to_decimal(item.get("valor_total")),
            )

    if ManutencaoFrota is not None:
        ManutencaoFrota.objects.filter(lancamento=obj).delete()
        for item in data.get("itens_manutencao", []):
            frota_id = item.get("frota_id") or None
            if not frota_id and item.get("matricula"):
                frota = _resolver_frota_por_matricula(item.get("matricula"), request=request)
                frota_id = getattr(frota, "id", None)
            kwargs = {
                "lancamento": obj,
                "frota_id": frota_id,
                "descricao": (item.get("descricao") or "Manutenção").strip() or "Manutenção",
                "km_inicio": _to_decimal(item.get("km_inicio")),
                "km_final": _to_decimal(item.get("km_final")),
                "km_total": _to_decimal(item.get("km_total")),
                "valor": _to_decimal(item.get("valor")),
                "observacao": item.get("observacao", ""),
            }
            ManutencaoFrota.objects.create(**kwargs)

    RevisaoFrota.objects.filter(lancamento=obj).delete()
    for item in data.get("itens_revisao", []):
        frota_id = item.get("frota_id") or None
        if not frota_id and item.get("matricula"):
            frota = _resolver_frota_por_matricula(item.get("matricula"), request=request)
            frota_id = getattr(frota, "id", None)
        if not frota_id:
            continue
        frota = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota).filter(id=frota_id).first()
        km_total = int(_to_decimal(item.get("km_total")))
        km_rodados = _km_rodados_frota(frota) + km_total
        RevisaoFrota.objects.create(
            lancamento=obj,
            frota_id=frota_id,
            funcionario_id=item.get("funcionario_id") or None,
            data_ultima_revisao=_to_date(item.get("data_ultima_revisao")) or obj.data_emissao or timezone.localdate(),
            km_ultima_revisao=int(_to_decimal(item.get("km_final") or item.get("km_ultima_revisao"))),
            km_rodados=km_rodados,
            kms_previsao=int(_to_decimal(item.get("kms_previsao"))),
            observacao=item.get("observacao", ""),
        )

    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def revisao_frota_list(request):
    matricula = (request.GET.get("matricula") or "").strip()
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    qs = RevisaoFrota.objects.select_related("frota", "funcionario").all()
    empresa = _empresa_ativa(request)
    if empresa:
        qs = qs.filter(frota__empresa=empresa)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)
    if data_inicio:
        qs = qs.filter(data_ultima_revisao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_ultima_revisao__lte=data_fim)
    linhas=[]
    for item in qs.order_by('-data_ultima_revisao','-id'):
        linhas.append({
            'id': item.id,
            'frota_id': item.frota_id or '',
            'matricula': getattr(getattr(item,'frota',None),'matricula','') or '',
            'marca': getattr(getattr(item,'frota',None),'marca','') or '',
            'modelo': getattr(getattr(item,'frota',None),'modelo','') or '',
            'data_ultima_revisao': item.data_ultima_revisao.strftime('%Y-%m-%d') if item.data_ultima_revisao else '',
            'km_ultima_revisao': int(item.km_ultima_revisao or 0),
            'km_rodados': int(item.km_rodados or 0),
            'kms_previsao': int(item.kms_previsao or 0),
            'km_para_revisao': int(item.km_para_revisao or 0),
            'funcionario_id': item.funcionario_id or '',
            'funcionario': getattr(getattr(item,'funcionario',None),'nome','') or '',
            'observacao': item.observacao or '',
        })
    return JsonResponse({'linhas': linhas})


@login_required
@require_POST
def revisao_frota_salvar(request):
    data = _json_body(request)
    obj = get_object_or_404(RevisaoFrota.objects.filter(frota__empresa=_empresa_ativa(request)), id=data['id']) if data.get('id') else RevisaoFrota()
    frota_id = data.get('frota_id') or None
    if not frota_id and data.get('matricula'):
        frota = _resolver_frota_por_matricula(data.get('matricula'), request=request)
        frota_id = getattr(frota,'id',None)
    if not frota_id:
        return JsonResponse({'error':'Selecione uma matrícula válida.'}, status=400)
    frota = get_object_or_404(_filtrar_empresa_generico(request, Frota.objects.all(), model=Frota), id=frota_id)
    obj.frota = frota
    obj.data_ultima_revisao = _to_date(data.get('data_ultima_revisao')) or timezone.localdate()
    obj.km_ultima_revisao = int(data.get('km_ultima_revisao') or 0)
    km_rodados = data.get('km_rodados')
    obj.km_rodados = int(km_rodados) if str(km_rodados).strip() not in ['', 'None'] else _km_rodados_frota(frota)
    obj.kms_previsao = int(data.get('kms_previsao') or 0)
    funcionario_id = data.get('funcionario_id') or None
    if funcionario_id:
        funcionario_id = get_object_or_404(_filtrar_empresa_generico(request, Funcionario.objects.all(), model=Funcionario), id=funcionario_id).id
    obj.funcionario_id = funcionario_id
    obj.observacao = data.get('observacao','')
    erro = _save_model_or_error(obj)
    if erro:
        return JsonResponse({'error': erro}, status=400)
    return JsonResponse({'ok': True, 'id': obj.id, 'km_rodados': obj.km_rodados, 'km_para_revisao': obj.km_para_revisao})


@login_required
@require_POST
def revisao_frota_excluir(request, id):
    get_object_or_404(RevisaoFrota.objects.filter(frota__empresa=_empresa_ativa(request)), id=id).delete()
    return JsonResponse({'ok': True})


@login_required
@require_GET
def lancamento_detalhe(request, id):
    obj = _get_lancamento_empresa_or_404(request, id)

    itens_combustivel = []
    if ItemCombustivel is not None:
        for item in ItemCombustivel.objects.filter(lancamento=obj).select_related("combustivel", "frota"):
            itens_combustivel.append({
                "combustivel_id": item.combustivel_id,
                "combustivel_nome": getattr(item.combustivel, "nome", ""),
                "frota_id": item.frota_id,
                "matricula": getattr(item.frota, "matricula", "") if getattr(item, "frota", None) else "",
                "km_inicio": float(_to_decimal(item.km_inicio)),
                "km_final": float(_to_decimal(item.km_final)),
                "km_total": float(_to_decimal(item.km_total)),
                "litro": float(_to_decimal(item.litro)),
                "valor_litro": float(_to_decimal(item.valor_litro)),
                "valor_total": float(_to_decimal(item.valor_total)),
            })

    itens_manutencao = []
    if ManutencaoFrota is not None:
        for item in ManutencaoFrota.objects.filter(lancamento=obj).select_related("frota"):
            itens_manutencao.append({
                "frota_id": item.frota_id,
                "descricao": getattr(item, "descricao", "Manutenção") or "Manutenção",
                "matricula": getattr(item.frota, "matricula", "") if getattr(item, "frota", None) else "",
                "km_inicio": float(_to_decimal(item.km_inicio)),
                "km_final": float(_to_decimal(item.km_final)),
                "km_total": float(_to_decimal(item.km_total)),
                "valor": float(_to_decimal(item.valor)),
                "observacao": getattr(item, "observacao", ""),
            })

    itens_revisao = []
    for item in RevisaoFrota.objects.filter(lancamento=obj).select_related("frota", "funcionario"):
        itens_revisao.append({
            "frota_id": item.frota_id,
            "matricula": getattr(item.frota, "matricula", "") if getattr(item, "frota", None) else "",
            "marca": getattr(item.frota, "marca", "") if getattr(item, "frota", None) else "",
            "modelo": getattr(item.frota, "modelo", "") if getattr(item, "frota", None) else "",
            "data_ultima_revisao": item.data_ultima_revisao.strftime("%Y-%m-%d") if item.data_ultima_revisao else "",
            "km_ultima_revisao": int(item.km_ultima_revisao or 0),
            "km_rodados": int(item.km_rodados or 0),
            "kms_previsao": int(item.kms_previsao or 0),
            "km_para_revisao": int(item.km_para_revisao or 0),
            "funcionario_id": item.funcionario_id or '',
            "funcionario": getattr(getattr(item, "funcionario", None), "nome", "") or '',
            "observacao": item.observacao or '',
        })

    return JsonResponse({
        "id": obj.id,
        "data_emissao": obj.data_emissao.strftime("%Y-%m-%d") if obj.data_emissao else "",
        "numero_fatura": obj.numero_fatura or "",
        "fornecedor_id": obj.fornecedor_id,
        "fornecedor": _fornecedor_nome(obj),
        "dinheiro": float(_to_decimal(getattr(obj, "dinheiro", 0))),
        "cartao": float(_to_decimal(getattr(obj, "cartao", 0))),
        "transferencia": float(_to_decimal(getattr(obj, "transferencia", 0))),
        "mbway": float(_to_decimal(getattr(obj, "mbway", 0))),
        "valor_fatura": float(_to_decimal(getattr(obj, "valor_fatura", 0))),
        "nota_credito": float(_to_decimal(getattr(obj, "nota_credito", 0))),
        "parcelas": getattr(obj, "parcelas", 1),
        "valor_parcela": float(_to_decimal(getattr(obj, "valor_parcela", 0))),
        "data_vencimento": obj.data_vencimento.strftime("%Y-%m-%d") if obj.data_vencimento else "",
        "status_pagamento": obj.status_pagamento or "",
        "total_pago": float(_total_pago_exibicao_lancamento(obj)),
        "saldo_aberto": float(_to_decimal(getattr(obj, "saldo_aberto", 0))),
        "itens_combustivel": itens_combustivel,
        "itens_manutencao": itens_manutencao,
        "itens_revisao": itens_revisao,
    })


@login_required
@require_POST
def lancamento_excluir(request, id):
    obj = _get_lancamento_empresa_or_404(request, id)
    obj.delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
@transaction.atomic
def lancamento_baixa(request, id):
    obj = _get_lancamento_empresa_or_404(request, id)
    data = _json_body(request)

    obj.dinheiro = _to_decimal(getattr(obj, "dinheiro", 0)) + _to_decimal(data.get("dinheiro"))
    obj.cartao = _to_decimal(getattr(obj, "cartao", 0)) + _to_decimal(data.get("cartao"))

    if hasattr(obj, "mbway"):
        obj.mbway = _to_decimal(getattr(obj, "mbway", 0)) + _to_decimal(data.get("mbway"))

    if hasattr(obj, "transferencia"):
        obj.transferencia = _to_decimal(getattr(obj, "transferencia", 0)) + _to_decimal(data.get("transferencia"))

    if hasattr(obj, "nota_credito"):
        obj.nota_credito = _to_decimal(getattr(obj, "nota_credito", 0)) + _to_decimal(data.get("nota_credito"))

    proxima_data = _to_date(data.get("proxima_data_vencimento"))
    if proxima_data and hasattr(obj, "data_vencimento"):
        obj.data_vencimento = proxima_data

    if hasattr(obj, "observacao_baixa"):
        obj.observacao_baixa = data.get("observacao", "")

    _recalcular_status_lancamento(obj)
    erro = _save_model_or_error(obj)
    if erro:
        return JsonResponse({"error": erro}, status=400)

    total_baixa = _to_decimal(data.get("total_baixa"))
    if total_baixa > Decimal("0.00") and _baixa_table_available():
        try:
            BaixaFatura.objects.create(
                empresa=getattr(obj, "empresa", None),
                lancamento=obj,
                usuario=request.user,
                data_baixa=timezone.localdate(),
                fornecedor_snapshot=_fornecedor_nome(obj),
                numero_fatura_snapshot=obj.numero_fatura or "",
                dinheiro=_to_decimal(data.get("dinheiro")),
                cartao=_to_decimal(data.get("cartao")),
                transferencia=_to_decimal(data.get("transferencia")),
                mbway=_to_decimal(data.get("mbway")),
                nota_credito=_to_decimal(data.get("nota_credito")),
                total_baixa=total_baixa,
                saldo_resultante=_to_decimal(obj.saldo_aberto),
                observacao=data.get("observacao", ""),
            )
        except (OperationalError, ProgrammingError):
            pass

    return JsonResponse({
        "ok": True,
        "status_pagamento": obj.status_pagamento,
        "total_pago": float(_total_pago_exibicao_lancamento(obj)),
        "saldo_aberto": float(_to_decimal(obj.saldo_aberto)),
    })


@login_required
@require_GET


@login_required
@require_GET
def faturas_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor").all())

    if q:
        filtro = Q(numero_fatura__icontains=q) | Q(fornecedor__nome__icontains=q) | Q(fornecedor__nif__icontains=q)
        if q.isdigit():
            filtro |= Q(id=int(q))
        qs = qs.filter(filtro)

    linhas = []
    vistos = set()
    for item in qs.order_by("-data_emissao", "-id")[:20]:
        numero = getattr(item, "numero_fatura", "") or f"Lançamento #{item.id}"
        chave = (numero or "").strip().lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        linhas.append({
            "id": item.id,
            "numero_fatura": numero,
            "fornecedor": _fornecedor_nome(item),
            "nif": _fornecedor_nif(item),
        })
    return JsonResponse({"linhas": linhas})

def consulta_lancamentos(request):
    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    )

    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    ordem = (request.GET.get("ordem") or "data").strip().lower()

    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)

    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)

    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)

    if ordem == "id":
        qs = qs.order_by("-id")
    else:
        qs = qs.order_by("data_emissao", "id")

    return JsonResponse({"linhas": _linhas_financeiras(qs), "ordem": ordem})


def _filtrar_lancamentos_financeiros(request, qs=None):
    qs = _filtrar_lancamentos_empresa(
        request,
        qs or Lancamento.objects.select_related("fornecedor", "empresa").all()
    )

    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    status = _normalizar_status_fatura(request.GET.get("status"))
    periodo_tipo = (request.GET.get("periodo_tipo") or "emissao").strip().lower()

    campo_data = "data_emissao"
    campo_ordenacao = ["data_emissao", "id"]
    if periodo_tipo == "vencimento":
        campo_data = "data_vencimento"
        campo_ordenacao = ["data_vencimento", "data_emissao", "id"]
    elif periodo_tipo == "pagamento":
        campo_data = "data_pagamento"
        campo_ordenacao = ["data_pagamento", "data_emissao", "id"]
    elif periodo_tipo == "id":
        campo_data = "data_emissao"
        campo_ordenacao = ["-id"]

    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    if data_inicio:
        qs = qs.filter(**{f"{campo_data}__gte": data_inicio})
    if data_fim:
        qs = qs.filter(**{f"{campo_data}__lte": data_fim})

    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)

    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)

    if status:
        qs = qs.filter(status_pagamento__iexact=status)

    return qs.order_by(*campo_ordenacao), {
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "fornecedor": fornecedor,
        "fatura": fatura,
        "status": status,
        "periodo_tipo": periodo_tipo,
    }


@login_required
def relatorio_baixa_faturas(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-faturas")

    qs, filtros = _filtrar_lancamentos_financeiros(request)
    qs = qs.filter(
        Q(numero_fatura__isnull=False) & ~Q(numero_fatura="") | Q(valor_fatura__gt=0) | Q(data_vencimento__isnull=False)
    ).distinct()

    return JsonResponse({
        "cabecalho": "BAIXA DE FATURAS",
        "linhas": _linhas_financeiras(qs),
        "totais": _totais_financeiros(qs),
        "filtros": filtros,
    })


@login_required
def relatorio_financeiro(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-financeiro")

    qs, filtros = _filtrar_lancamentos_financeiros(request)
    return JsonResponse({
        "cabecalho": "RELATÓRIO FINANCEIRO",
        "linhas": _linhas_financeiras(qs),
        "totais": _totais_financeiros(qs),
        "filtros": filtros,
    })


def _baixa_table_available():
    try:
        return BaixaFatura._meta.db_table in connection.introspection.table_names()
    except Exception:
        return False


def _linhas_relatorio_baixa(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()

    linhas = []
    total_baixado = Decimal("0.00")
    total_em_aberto = Decimal("0.00")
    total_dinheiro = Decimal("0.00")
    total_cartao = Decimal("0.00")
    total_transferencia = Decimal("0.00")
    total_mbway = Decimal("0.00")
    total_nota_credito = Decimal("0.00")
    previsao_vencimento_20_dias = Decimal("0.00")

    if _baixa_table_available():
        try:
            qs = BaixaFatura.objects.select_related("empresa", "usuario", "lancamento", "lancamento__fornecedor")
            empresa = _empresa_ativa(request)
            if empresa is not None:
                qs = qs.filter(empresa=empresa)
            if data_inicio:
                qs = qs.filter(data_baixa__gte=data_inicio)
            if data_fim:
                qs = qs.filter(data_baixa__lte=data_fim)
            if fornecedor:
                qs = qs.filter(
                    Q(fornecedor_snapshot__icontains=fornecedor)
                    | Q(lancamento__fornecedor__nome__icontains=fornecedor)
                    | Q(lancamento__fornecedor__nif__icontains=fornecedor)
                )
            if fatura:
                qs = qs.filter(
                    Q(numero_fatura_snapshot__icontains=fatura)
                    | Q(lancamento__numero_fatura__icontains=fatura)
                )

            for item in qs.order_by('-data_baixa', '-id'):
                lanc = getattr(item, 'lancamento', None)
                valor_total = _to_decimal(getattr(lanc, 'valor_fatura', 0) if lanc else 0)
                valor_pago = _valor_pago_exibicao_baixa(item, lanc)
                saldo = _to_decimal(getattr(item, 'saldo_resultante', 0))
                dinheiro = _to_decimal(getattr(item, 'dinheiro', 0))
                cartao = _to_decimal(getattr(item, 'cartao', 0))
                transferencia = _to_decimal(getattr(item, 'transferencia', 0))
                mbway = _to_decimal(getattr(item, 'mbway', 0))
                nota_credito = _to_decimal(getattr(item, 'nota_credito', 0))
                data_vencimento = getattr(lanc, 'data_vencimento', None) if lanc else None
                total_baixado += valor_pago
                total_em_aberto += saldo
                total_dinheiro += dinheiro
                total_cartao += cartao
                total_transferencia += transferencia
                total_mbway += mbway
                total_nota_credito += nota_credito
                if saldo > Decimal('0.00') and data_vencimento and timezone.localdate() <= data_vencimento <= (timezone.localdate() + timedelta(days=20)):
                    previsao_vencimento_20_dias += saldo
                linhas.append({
                    'id': getattr(lanc, 'id', None) or item.id,
                    'lancamento_id': getattr(lanc, 'id', None) or item.id,
                    'fornecedor': item.fornecedor_snapshot or _fornecedor_nome(lanc),
                    'numero_fatura': item.numero_fatura_snapshot or getattr(lanc, 'numero_fatura', '') or '',
                    'data_baixa': _format_date_br(item.data_baixa),
                    'forma_pagamento': _forma_pagamento_baixa(item),
                    'usuario': getattr(item.usuario, 'username', '') if getattr(item, 'usuario', None) else 'Maia',
                    'valor_total': float(valor_total),
                    'dinheiro': float(dinheiro),
                    'cartao': float(cartao),
                    'transferencia': float(transferencia),
                    'mbway': float(mbway),
                    'nota_credito': float(-nota_credito),
                    'valor_baixado': float(valor_pago),
                    'saldo_resultante': float(saldo),
                    'em_aberto': float(saldo),
                    'data_vencimento': _format_date_br(data_vencimento),
                })
        except (OperationalError, ProgrammingError):
            pass

    if not linhas:
        qs, _ = _filtrar_lancamentos_financeiros(request)
        qs = qs.filter(Q(status_pagamento__iexact='Parcial') | Q(status_pagamento__iexact='Paga'))
        for item in qs.order_by('-data_pagamento', '-data_emissao', '-id'):
            valor_total = _to_decimal(getattr(item, 'valor_fatura', 0))
            valor_pago = _total_pago_exibicao_lancamento(item)
            nota_credito = _to_decimal(getattr(item, 'nota_credito', 0))
            if valor_pago == Decimal('0.00') and nota_credito <= Decimal('0.00'):
                continue
            saldo = _to_decimal(getattr(item, 'saldo_aberto', 0))
            dinheiro = _to_decimal(getattr(item, 'dinheiro', 0))
            cartao = _to_decimal(getattr(item, 'cartao', 0))
            transferencia = _to_decimal(getattr(item, 'transferencia', 0))
            mbway = _to_decimal(getattr(item, 'mbway', 0))
            nota_credito = _to_decimal(getattr(item, 'nota_credito', 0))
            data_vencimento = getattr(item, 'data_vencimento', None)
            total_baixado += valor_pago
            total_em_aberto += saldo
            total_dinheiro += dinheiro
            total_cartao += cartao
            total_transferencia += transferencia
            total_mbway += mbway
            total_nota_credito += nota_credito
            if saldo > Decimal('0.00') and data_vencimento and timezone.localdate() <= data_vencimento <= (timezone.localdate() + timedelta(days=20)):
                previsao_vencimento_20_dias += saldo
            linhas.append({
                'id': item.id,
                'lancamento_id': item.id,
                'fornecedor': _fornecedor_nome(item),
                'numero_fatura': getattr(item, 'numero_fatura', '') or '',
                'data_baixa': item.data_pagamento.strftime('%Y-%m-%d') if getattr(item, 'data_pagamento', None) else (item.data_emissao.strftime('%Y-%m-%d') if getattr(item, 'data_emissao', None) else ''),
                'forma_pagamento': _forma_pagamento_baixa(item),
                'usuario': getattr(getattr(item, 'criado_por', None), 'username', '') or 'Maia',
                'valor_total': float(valor_pago if valor_total <= Decimal('0.00') else valor_total),
                'dinheiro': float(dinheiro),
                'cartao': float(cartao),
                'transferencia': float(transferencia),
                'mbway': float(mbway),
                'nota_credito': float(-nota_credito),
                'valor_baixado': float(valor_pago),
                'saldo_resultante': float(saldo),
                'em_aberto': float(saldo),
                'data_vencimento': _format_date_br(data_vencimento),
            })

    filtros = {
        'data_inicio': data_inicio.strftime('%Y-%m-%d') if data_inicio else '',
        'data_fim': data_fim.strftime('%Y-%m-%d') if data_fim else '',
        'fornecedor': fornecedor,
        'fatura': fatura,
    }
    totais = {
        'total_baixado': float(total_baixado),
        'total_em_aberto': float(total_em_aberto),
        'total_dinheiro': float(total_dinheiro),
        'total_cartao': float(total_cartao),
        'total_transferencia': float(total_transferencia),
        'total_mbway': float(total_mbway),
        'total_nota_credito': float(-total_nota_credito),
        'previsao_vencimento_20_dias': float(previsao_vencimento_20_dias),
    }
    return linhas, totais, filtros



def _forma_pagamento_baixa(item):
    formas = []
    if _to_decimal(getattr(item, "dinheiro", 0)) > 0:
        formas.append("Dinheiro")
    if _to_decimal(getattr(item, "cartao", 0)) > 0:
        formas.append("Cartão")
    if _to_decimal(getattr(item, "transferencia", 0)) > 0:
        formas.append("Transferência")
    if _to_decimal(getattr(item, "mbway", 0)) > 0:
        formas.append("MBWay")
    if _to_decimal(getattr(item, "nota_credito", 0)) > 0:
        formas.append("Nota de Crédito")
    return ", ".join(formas)


def _safe_div(numerador, denominador, casas="0.00"):
    numerador = _to_decimal(numerador)
    denominador = _to_decimal(denominador)
    if denominador == 0:
        return Decimal(casas)
    return numerador / denominador


def _aplicar_estilo_excel(ws):
    preenchimento = PatternFill("solid", fgColor="D9EAF7")
    borda = Border(
        left=Side(style="thin", color="D5DCE5"),
        right=Side(style="thin", color="D5DCE5"),
        top=Side(style="thin", color="D5DCE5"),
        bottom=Side(style="thin", color="D5DCE5"),
    )
    for celula in ws[1]:
        celula.font = Font(bold=True, color="1E3A5F")
        celula.fill = preenchimento
        celula.border = borda
        celula.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for celula in row:
            celula.border = borda
            celula.alignment = Alignment(vertical="top")


def _auto_ajustar_colunas(ws):
    for idx in range(1, ws.max_column + 1):
        letra = get_column_letter(idx)
        maior = 0
        for row in ws.iter_rows(min_col=idx, max_col=idx):
            cell = row[0]
            if getattr(cell, 'value', None) is None:
                continue
            tamanho = len(str(cell.value))
            if tamanho > maior:
                maior = tamanho
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 40)


def _empresa_relatorio_info(request):
    empresa = _empresa_ativa(request)
    config = _configuracao_sistema()
    logo_empresa = getattr(getattr(empresa, "logo", None), "path", "") if empresa else ""
    logo_sistema = getattr(getattr(config, "logo_login", None), "path", "") if config else ""
    return {
        "nome": getattr(empresa, "nome", "") or "Elion Gestão",
        "logo_path": logo_empresa or logo_sistema or "",
        "logo_sistema_path": logo_sistema or "",
        "sistema_nome": getattr(config, "nome_sistema", "Elion Gestão") if config else "Elion Gestão",
    }


def _aplicar_cabecalho_relatorio_excel(ws, request, titulo, subtitulo=""):
    info = _empresa_relatorio_info(request)
    ultima_coluna = max(ws.max_column, 8)
    ws.insert_rows(1, 5)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ultima_coluna)
    ws.cell(1, 2, info["nome"])
    ws.cell(1, 2).font = Font(size=15, bold=True, color="163A63")
    ws.cell(1, 2).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ultima_coluna)
    ws.cell(2, 2, titulo)
    ws.cell(2, 2).font = Font(size=12, bold=True, color="1F4E78")
    ws.cell(2, 2).alignment = Alignment(horizontal="left", vertical="center")

    if subtitulo:
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=ultima_coluna)
        ws.cell(3, 2, subtitulo)
        ws.cell(3, 2).font = Font(size=10, italic=True, color="475569")
        ws.cell(3, 2).alignment = Alignment(horizontal="left", vertical="center")

    if XLImage and info.get("logo_path") and os.path.exists(info["logo_path"]):
        try:
            imagem = XLImage(info["logo_path"])
            imagem.height = 56
            imagem.width = 56
            ws.add_image(imagem, "A1")
            ws.row_dimensions[1].height = 44
            ws.row_dimensions[2].height = 22
        except Exception:
            pass


def _estilizar_bloco_totais_excel(ws, linha, valores):
    for coluna, valor in enumerate(valores, start=1):
        cell = ws.cell(linha, coluna, value=valor)
        cell.font = Font(bold=True, color="163A63")
        cell.fill = PatternFill("solid", fgColor="EEF6FF")
        cell.border = Border(
            left=Side(style="thin", color="D5DCE5"),
            right=Side(style="thin", color="D5DCE5"),
            top=Side(style="thin", color="D5DCE5"),
            bottom=Side(style="thin", color="D5DCE5"),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _formatar_valor_pdf(valor):
    if valor is None:
        return ''
    if isinstance(valor, Decimal):
        return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    if isinstance(valor, float):
        return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return str(valor)


def _exportar_tabela_pdf(request, titulo, cabecalhos, linhas, nome_arquivo, subtitulo=''):
    info = _empresa_relatorio_info(request)
    if not str(nome_arquivo).lower().endswith('.pdf'):
        nome_arquivo = f"{nome_arquivo}.pdf"
    total_chars = sum(max(len(str(c)), 8) for c in cabecalhos)
    usar_landscape = len(cabecalhos) > 7 or total_chars > 80

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Paragraph, Table, TableStyle, Image as RLImage

        pagesize = landscape(A4) if usar_landscape else portrait(A4)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

        doc = SimpleDocTemplate(response, pagesize=pagesize, leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
        styles = getSampleStyleSheet()
        story = []
        if info.get('logo_path') and os.path.exists(info['logo_path']):
            try:
                logo = RLImage(info['logo_path'])
                logo.drawHeight = 20 * mm
                logo.drawWidth = 20 * mm
                story.append(logo)
                story.append(Spacer(1, 4))
            except Exception:
                pass
        story.append(Paragraph(info.get('nome') or 'Elion Gestão', styles['Title']))
        story.append(Spacer(1, 4))
        story.append(Paragraph(titulo, styles['Heading2']))
        if subtitulo:
            story.append(Paragraph(subtitulo, styles['Normal']))
        story.append(Spacer(1, 6))

        tabela = [cabecalhos] + [[_formatar_valor_pdf(v) for v in linha] for linha in linhas]
        larg_disp = pagesize[0] - doc.leftMargin - doc.rightMargin
        pesos = [max(len(str(c)), 8) for c in cabecalhos]
        total_pesos = sum(pesos) or 1
        col_widths = [larg_disp * (peso / total_pesos) for peso in pesos]
        tb = Table(tabela, repeatRows=1, colWidths=col_widths)
        tb.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9EAF7')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#163A63')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.35, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(tb)
        doc.build(story)
        return response
    except ModuleNotFoundError:
        pass
    except Exception:
        pass

    try:
        from fpdf import FPDF

        orientacao = 'L' if usar_landscape else 'P'
        pdf = FPDF(orientation=orientacao, unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.set_margins(10, 10, 10)
        pdf.add_page()

        logo_path = info.get('logo_path') or ''
        if logo_path and os.path.exists(logo_path):
            try:
                pdf.image(logo_path, x=10, y=10, w=24)
                pdf.set_y(36)
            except Exception:
                pdf.set_y(10)

        pdf.set_font('Helvetica', 'B', 16)
        pdf.cell(0, 8, (info.get('nome') or 'Elion Gestão')[:90], ln=1)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 7, titulo[:110], ln=1)
        if subtitulo:
            pdf.set_font('Helvetica', '', 9)
            pdf.multi_cell(0, 5, subtitulo[:220])
        pdf.ln(2)

        largura_util = pdf.w - pdf.l_margin - pdf.r_margin
        pesos = [max(len(str(c)), 8) for c in cabecalhos]
        total_pesos = sum(pesos) or 1
        larguras = [largura_util * (peso / total_pesos) for peso in pesos]
        min_largura = 18 if orientacao == 'L' else 16
        larguras = [max(l, min_largura) for l in larguras]
        excesso = sum(larguras) - largura_util
        if excesso > 0:
            fator = largura_util / sum(larguras)
            larguras = [l * fator for l in larguras]

        pdf.set_fill_color(217, 234, 247)
        pdf.set_text_color(22, 58, 99)
        pdf.set_draw_color(203, 213, 225)
        pdf.set_font('Helvetica', 'B', 8)
        for cab, largura in zip(cabecalhos, larguras):
            pdf.cell(largura, 8, str(cab)[:40], border=1, align='C', fill=True)
        pdf.ln(8)

        pdf.set_text_color(15, 23, 42)
        pdf.set_font('Helvetica', '', 7)
        for linha in linhas:
            y_inicio = pdf.get_y()
            alturas = []
            valores = [_formatar_valor_pdf(v) for v in linha]
            for largura, valor in zip(larguras, valores):
                texto = str(valor)
                linhas_estimadas = max(1, (len(texto) // max(1, int(largura / 2.6))) + 1)
                alturas.append(max(6, linhas_estimadas * 4))
            altura = min(max(alturas), 24)
            if y_inicio + altura > pdf.h - pdf.b_margin:
                pdf.add_page()
                pdf.set_fill_color(217, 234, 247)
                pdf.set_text_color(22, 58, 99)
                pdf.set_font('Helvetica', 'B', 8)
                for cab, largura in zip(cabecalhos, larguras):
                    pdf.cell(largura, 8, str(cab)[:40], border=1, align='C', fill=True)
                pdf.ln(8)
                pdf.set_text_color(15, 23, 42)
                pdf.set_font('Helvetica', '', 7)
                y_inicio = pdf.get_y()
            x = pdf.l_margin
            for largura, valor in zip(larguras, valores):
                pdf.set_xy(x, y_inicio)
                pdf.multi_cell(largura, 4, str(valor), border=1)
                pdf.set_xy(x + largura, y_inicio)
                x += largura
            pdf.set_y(y_inicio + altura)

        pdf_bytes = pdf.output(dest='S')
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin-1', errors='ignore')
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        return response
    except Exception:
        def _pdf_escape(value):
            return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

        def _latin_text(value):
            return _pdf_escape(str(value).encode('latin-1', errors='replace').decode('latin-1'))

        def _gerar_pdf_simples(linhas_texto):
            width = 842 if usar_landscape else 595
            height = 595 if usar_landscape else 842
            margin = 36
            line_height = 14
            font_size = 9
            linhas_por_pagina = 48
            page_count = max(1, math.ceil(len(linhas_texto) / linhas_por_pagina))
            objetos = []

            def add_obj(data):
                objetos.append(data)
                return len(objetos)

            font_id = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
            bold_id = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
            page_ids = []
            content_ids = []

            for page_index in range(page_count):
                chunk = linhas_texto[page_index * linhas_por_pagina:(page_index + 1) * linhas_por_pagina]
                comandos = ["BT"]
                y = height - margin
                for idx, linha_txt in enumerate(chunk):
                    is_header = idx < 3
                    font_name = 'F2' if is_header else 'F1'
                    tamanho = 15 if idx == 0 else (11 if idx in (1, 2) else font_size)
                    if idx == 3:
                        y -= 4
                    comandos.append(f"/{font_name} {tamanho} Tf")
                    comandos.append(f"1 0 0 1 {margin} {int(y)} Tm")
                    comandos.append(f"({_latin_text(linha_txt)}) Tj")
                    y -= line_height if idx >= 3 else (18 if idx == 0 else 15)
                comandos.append("ET")
                stream_text = "\n".join(comandos)
                stream_bytes = stream_text.encode('latin-1', errors='replace')
                content_id = add_obj(f"<< /Length {len(stream_bytes)} >>\nstream\n{stream_text}\nendstream")
                content_ids.append(content_id)
                page_ids.append(add_obj("PAGE_PLACEHOLDER"))

            kids = ' '.join(f'{pid} 0 R' for pid in page_ids)
            pages_id = add_obj(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>")
            for i, pid in enumerate(page_ids):
                objetos[pid - 1] = f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 {font_id} 0 R /F2 {bold_id} 0 R >> >> /Contents {content_ids[i]} 0 R >>"
            catalog_id = add_obj(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

            pdf = bytearray(b'%PDF-1.4\n%\xe2\xe3\xcf\xd3\n')
            offsets = [0]
            for i, obj in enumerate(objetos, start=1):
                offsets.append(len(pdf))
                pdf.extend(f"{i} 0 obj\n{obj}\nendobj\n".encode('latin-1', errors='replace'))
            xref_pos = len(pdf)
            pdf.extend(f"xref\n0 {len(objetos) + 1}\n".encode('latin-1'))
            pdf.extend(b"0000000000 65535 f \n")
            for off in offsets[1:]:
                pdf.extend(f"{off:010d} 00000 n \n".encode('latin-1'))
            pdf.extend(f"trailer\n<< /Size {len(objetos) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode('latin-1'))
            return bytes(pdf)

        titulo_empresa = info.get('nome') or 'Elion Gestão'
        linhas_texto = [titulo_empresa, titulo, subtitulo or '', '']
        linhas_texto.append(' | '.join(str(c) for c in cabecalhos))
        linhas_texto.append('-' * min(180, max(24, sum(len(str(c)) for c in cabecalhos) + 3 * len(cabecalhos))))
        for linha in linhas:
            valores = [_formatar_valor_pdf(v) for v in linha]
            linhas_texto.append(' | '.join(str(v) for v in valores))
        pdf_bytes = _gerar_pdf_simples(linhas_texto)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        return response


def _imprimir_tabela_html(request, titulo, cabecalhos, linhas, subtitulo=''):
    info = _empresa_relatorio_info(request)
    head = ''.join(f'<th>{c}</th>' for c in cabecalhos)

    def _eh_linha_totais(linha):
        return any(
            isinstance(v, str) and v.strip().lower().startswith('tota')
            for v in linha
        )

    def _eh_rotulo(v):
        return v is None or v == '' or (isinstance(v, str) and v.strip().lower().startswith('tota'))

    linha_totais = linhas[-1] if linhas and _eh_linha_totais(linhas[-1]) else None

    cards_html = ''
    if linha_totais:
        cartoes = []
        for idx, cabecalho in enumerate(cabecalhos):
            if idx >= len(linha_totais):
                break
            valor = linha_totais[idx]
            if _eh_rotulo(valor):
                continue
            cartoes.append((cabecalho, valor))
        if cartoes:
            itens = []
            for i, (rotulo, valor) in enumerate(cartoes):
                destaque = ' cartao-destaque' if i == len(cartoes) - 1 else ''
                itens.append(
                    f"<div class='cartao{destaque}'><span class='cartao-rotulo'>{rotulo}</span>"
                    f"<span class='cartao-valor'>{_formatar_valor_pdf(valor)}</span></div>"
                )
            cards_html = f"<div class='cartoes-totais'>{''.join(itens)}</div>"

    body_rows = []
    for linha in linhas:
        classe = ' class="linha-total"' if _eh_linha_totais(linha) else ''
        celulas = ''.join(f'<td>{_formatar_valor_pdf(v)}</td>' for v in linha)
        body_rows.append(f'<tr{classe}>{celulas}</tr>')
    body = ''.join(body_rows)

    logo_html = ''
    logo_path = info.get('logo_path') or ''
    if logo_path and os.path.exists(logo_path):
        try:
            mime = mimetypes.guess_type(logo_path)[0] or 'image/png'
            data = base64.b64encode(Path(logo_path).read_bytes()).decode('ascii')
            logo_html = f"<img src='data:{mime};base64,{data}' alt='Logo' class='logo'>"
        except Exception:
            logo_html = ''

    agora = timezone.localtime().strftime('%d/%m/%Y às %H:%M')
    subtitulo_html = f"<p class='subtitulo'>{subtitulo}</p>" if subtitulo else ''
    total_registos = max(len(linhas) - (1 if linhas and _eh_linha_totais(linhas[-1]) else 0), 0)

    html = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{titulo}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', Arial, Helvetica, sans-serif;
    margin: 0;
    padding: 14mm 12mm;
    color: #0F172A;
    font-size: 12px;
  }}
  .cabecalho {{
    display: flex;
    align-items: center;
    gap: 14px;
    border-bottom: 3px solid #163A63;
    padding-bottom: 12px;
    margin-bottom: 6px;
  }}
  .logo {{
    height: 60px;
    max-width: 150px;
    object-fit: contain;
  }}
  .cabecalho-textos h1 {{
    font-size: 19px;
    margin: 0 0 2px;
    color: #163A63;
    letter-spacing: .2px;
  }}
  .cabecalho-textos h2 {{
    font-size: 14px;
    margin: 0;
    color: #1F4E78;
    font-weight: 600;
  }}
  .subtitulo {{
    margin: 6px 0 0;
    color: #475569;
    font-size: 11.5px;
    font-style: italic;
  }}
  .meta-linha {{
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    margin: 10px 0 14px;
    font-size: 10.5px;
    color: #475569;
  }}
  .meta-linha span strong {{ color: #1F4E78; }}
  .cartoes-totais {{
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin: 0 0 16px;
  }}
  .cartao {{
    flex: 1 1 140px;
    background: #F8FAFC;
    border: 1px solid #CBD5E1;
    border-radius: 8px;
    padding: 10px 14px;
    display: flex;
    flex-direction: column;
    gap: 4px;
  }}
  .cartao-rotulo {{
    font-size: 9.5px;
    text-transform: uppercase;
    letter-spacing: .3px;
    color: #64748B;
    font-weight: 600;
  }}
  .cartao-valor {{
    font-size: 16px;
    font-weight: 700;
    color: #163A63;
  }}
  .cartao.cartao-destaque {{
    background: #163A63;
    border-color: #163A63;
  }}
  .cartao.cartao-destaque .cartao-rotulo {{ color: #BFDBFE; }}
  .cartao.cartao-destaque .cartao-valor {{ color: #fff; font-size: 19px; }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 11px;
  }}
  thead th {{
    background: #D9EAF7;
    color: #163A63;
    text-align: left;
    padding: 8px 9px;
    font-weight: 700;
    border: 1px solid #CBD5E1;
    text-transform: uppercase;
    font-size: 9.5px;
    letter-spacing: .3px;
  }}
  tbody td {{
    padding: 6px 9px;
    border: 1px solid #CBD5E1;
    vertical-align: top;
  }}
  tbody tr:nth-child(even) {{ background: #F8FAFC; }}
  tbody tr:hover {{ background: #F1F5F9; }}
  tr.linha-total td {{
    background: #EEF6FF !important;
    color: #163A63;
    font-weight: 700;
    border-top: 2px solid #1F4E78;
  }}
  .rodape {{
    margin-top: 16px;
    padding-top: 8px;
    border-top: 1px solid #CBD5E1;
    display: flex;
    justify-content: space-between;
    font-size: 9.5px;
    color: #94A3B8;
  }}
  .botao-imprimir {{
    position: fixed;
    top: 14px;
    right: 14px;
    background: #163A63;
    color: #fff;
    border: none;
    padding: 9px 16px;
    border-radius: 6px;
    font-size: 12px;
    cursor: pointer;
    box-shadow: 0 2px 6px rgba(0,0,0,.2);
  }}
  .botao-imprimir:hover {{ background: #1F4E78; }}
  @page {{ margin: 12mm; }}
  @media print {{
    body {{ padding: 0; }}
    .botao-imprimir {{ display: none; }}
    thead {{ display: table-header-group; }}
    tr {{ page-break-inside: avoid; }}
  }}
</style>
</head>
<body onload="window.print()">
  <button class="botao-imprimir" onclick="window.print()">Imprimir novamente</button>
  <div class="cabecalho">
    {logo_html}
    <div class="cabecalho-textos">
      <h1>{info.get('nome') or 'Elion Gestão'}</h1>
      <h2>{titulo}</h2>
    </div>
  </div>
  {subtitulo_html}
  <div class="meta-linha">
    <span>Total de registos: <strong>{total_registos}</strong></span>
    <span>Gerado em: <strong>{agora}</strong></span>
  </div>
  {cards_html}
  <table>
    <thead><tr>{head}</tr></thead>
    <tbody>{body}</tbody>
  </table>
  <div class="rodape">
    <span>{info.get('sistema_nome') or 'Elion Gestão'}</span>
    <span>Documento gerado automaticamente para fins de consulta interna.</span>
  </div>
</body>
</html>"""
    return HttpResponse(html)


def _fornecedores_combustivel_fallback_ids(fornecedor_texto="", request=None):
    # Fallback apenas para postos/fornecedores claramente ligados a combustível.
    # Mantemos a lista restrita para não misturar despesas que não são abastecimentos.
    termos_fixos = {
        'superbeja', 'bribeja', 'galp', 'bp', 'repsol', 'cepsa', 'prio'
    }
    nifs_fixos = {'504499211', '514192020'}
    ids = set()
    texto = (fornecedor_texto or '').strip().lower()
    for forn in (_fornecedores_empresa_qs(request) if request is not None else Fornecedor.objects.all()).only('id', 'nome', 'nif'):
        nome = (getattr(forn, 'nome', '') or '').lower()
        nif = (getattr(forn, 'nif', '') or '')
        eh_combustivel = nif in nifs_fixos or any(term in nome for term in termos_fixos)
        if not eh_combustivel:
            continue
        if texto and not (texto in nome or texto in nif):
            continue
        ids.add(forn.id)
    return ids



def _lancamento_eh_combustivel(lancamento):
    # Considera combustível apenas quando houver item de combustível gravado
    # com um tipo de combustível efetivamente selecionado.
    if ItemCombustivel is None:
        return False
    try:
        return ItemCombustivel.objects.filter(
            lancamento=lancamento,
            combustivel_id__isnull=False,
        ).exists()
    except Exception:
        return False

def _linhas_relatorio_combustivel(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    combustivel = (request.GET.get("combustivel") or "").strip()

    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    total_valor = Decimal("0.00")
    total_litros = Decimal("0.00")
    total_km = Decimal("0.00")
    resumo_map = {}
    linhas = []

    def _combustivel_match(nome):
        if not combustivel:
            return True
        return combustivel.casefold() in (nome or '').casefold()

    def _matricula_match(valor):
        if not matricula:
            return True
        return matricula.casefold() in (valor or '').casefold()

    def _fornecedor_match(lanc):
        if not fornecedor:
            return True
        termo = fornecedor.casefold()
        return termo in (_fornecedor_nome(lanc) or '').casefold() or termo in (_fornecedor_nif(lanc) or '').casefold()

    def _add_row(lanc, matric='-', comb='', km_inicio=0, km_final=0, km_total=0, litros=0, valor_litro=0, valor_total=None):
        nonlocal total_valor, total_litros, total_km, linhas, resumo_map
        litros = _to_decimal(litros)
        valor_total = _to_decimal(valor_total)
        if valor_total <= 0:
            valor_total = _to_decimal(getattr(lanc, 'total', 0) or getattr(lanc, 'valor_fatura', 0))
        valor_litro = _to_decimal(valor_litro)
        if valor_litro <= 0 and litros > 0:
            valor_litro = _safe_div(valor_total, litros, '0.000')
        km_inicio = _to_decimal(km_inicio)
        km_final = _to_decimal(km_final)
        km_total = _to_decimal(km_total)
        if km_total <= 0 and km_final > km_inicio:
            km_total = km_final - km_inicio
        media_km_l = _safe_div(km_total, litros)
        media_euro_km = _safe_div(valor_total, km_total, "0.000")

        total_valor += valor_total
        total_litros += litros
        total_km += km_total

        chave_matricula = matric or '-'
        acumulado = resumo_map.setdefault(chave_matricula, {
            'matricula': chave_matricula,
            'abastecimentos': 0,
            'litros': Decimal('0.00'),
            'km': Decimal('0.00'),
            'valor': Decimal('0.00'),
        })
        acumulado['abastecimentos'] += 1
        acumulado['litros'] += litros
        acumulado['km'] += km_total
        acumulado['valor'] += valor_total

        linhas.append({
            'id': getattr(lanc, 'id', None),
            'item_id': None,
            'lancamento_id': getattr(lanc, 'id', None),
            'data': lanc.data_emissao.strftime('%d/%m/%y') if getattr(lanc, 'data_emissao', None) else '',
            'empresa': getattr(getattr(lanc, 'empresa', None), 'nome', '') or '',
            'fornecedor': _fornecedor_nome(lanc) if lanc else '',
            'fatura': getattr(lanc, 'numero_fatura', '') or '',
            'matricula': chave_matricula,
            'combustivel': comb or '',
            'km_inicio': float(km_inicio),
            'km_final': float(km_final),
            'km_total': float(km_total),
            'km': float(km_total),
            'litros': float(litros),
            'valor_litro': float(valor_litro),
            'valor_total': float(valor_total),
            'valor': float(valor_total),
            'preco_medio_litro': float(valor_litro),
            'media_km_l': float(media_km_l),
            'media_euro_km': float(media_euro_km),
        })

    empresa = _empresa_ativa(request) if hasattr(Lancamento, 'empresa_id') else None

    if ItemCombustivel is not None:
        try:
            itens_qs = ItemCombustivel.objects.select_related(
                'lancamento', 'lancamento__fornecedor', 'lancamento__empresa', 'frota', 'combustivel'
            )
            if empresa:
                itens_qs = itens_qs.filter(lancamento__empresa=empresa)
            if data_inicio:
                itens_qs = itens_qs.filter(lancamento__data_emissao__gte=data_inicio)
            if data_fim:
                itens_qs = itens_qs.filter(lancamento__data_emissao__lte=data_fim)
            if fornecedor:
                itens_qs = itens_qs.filter(
                    Q(lancamento__fornecedor__nome__icontains=fornecedor)
                    | Q(lancamento__fornecedor__nif__icontains=fornecedor)
                )
            if matricula:
                itens_qs = itens_qs.filter(Q(frota__matricula__icontains=matricula) | Q(frota__isnull=True))
            if combustivel == '__nenhum__':
                itens_qs = itens_qs.filter(Q(combustivel__isnull=True) | Q(combustivel__nome__isnull=True) | Q(combustivel__nome__exact=''))
            elif combustivel:
                itens_qs = itens_qs.filter(combustivel__nome__icontains=combustivel)
            if request.GET.get('fatura'):
                itens_qs = itens_qs.filter(lancamento__numero_fatura__icontains=(request.GET.get('fatura') or '').strip())

            for item in itens_qs.order_by('lancamento__data_emissao', 'lancamento__id', 'id'):
                lanc = getattr(item, 'lancamento', None)
                if not lanc:
                    continue
                item_id = getattr(item, 'id', None)
                _add_row(
                    lanc,
                    matric=getattr(getattr(item, 'frota', None), 'matricula', '') or '-',
                    comb=getattr(getattr(item, 'combustivel', None), 'nome', '') or '',
                    km_inicio=getattr(item, 'km_inicio', 0),
                    km_final=getattr(item, 'km_final', 0),
                    km_total=getattr(item, 'km_total', 0),
                    litros=getattr(item, 'litro', 0),
                    valor_litro=getattr(item, 'valor_litro', 0),
                    valor_total=getattr(item, 'valor_total', None),
                )
                linhas[-1]['id'] = item_id
                linhas[-1]['item_id'] = item_id
        except Exception:
            linhas = []
            resumo_map = {}
            total_valor = Decimal('0.00')
            total_litros = Decimal('0.00')
            total_km = Decimal('0.00')

    linhas.sort(key=lambda l: (
        _to_date(l.get("data")) or date.min,
        l.get("lancamento_id") or 0,
        l.get("item_id") or l.get("id") or 0,
    ))

    resumo_matriculas = []
    for item in resumo_map.values():
        litros = item['litros']
        km = item['km']
        valor = item['valor']
        resumo_matriculas.append({
            'matricula': item['matricula'],
            'abastecimentos': item['abastecimentos'],
            'litros': float(litros),
            'km': float(km),
            'valor': float(valor),
            'media_km_l': float(_safe_div(km, litros)),
            'media_euro_km': float(_safe_div(valor, km, '0.000')),
            'preco_medio_litro': float(_safe_div(valor, litros, '0.000')),
        })

    resumo_matriculas.sort(key=lambda x: (x.get('matricula') or ''))
    resumo = {
        'valor_total': float(total_valor),
        'litros_total': float(total_litros),
        'km_total': float(total_km),
        'media_km_l': float(_safe_div(total_km, total_litros)),
        'media_euro_km': float(_safe_div(total_valor, total_km, '0.000')),
        'preco_medio_litro': float(_safe_div(total_valor, total_litros, '0.000')),
    }
    return linhas, resumo_matriculas, resumo



@login_required
def relatorio_combustivel(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-combustivel")

    linhas, resumo_matriculas, resumo = _linhas_relatorio_combustivel(request)
    return JsonResponse({
        "linhas": linhas,
        "resumo_matriculas": resumo_matriculas,
        "resumo": resumo,
    })


@login_required
def relatorio_documentos(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-documentos")

    matricula = (request.GET.get("matricula") or "").strip()
    status = (request.GET.get("status") or "").strip().lower()

    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)

    hoje = date.today()
    limite = hoje + timedelta(days=30)

    linhas = []
    for item in qs.order_by("matricula"):
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)

        estado = "OK"
        if seguro and seguro < hoje:
            estado = "Vencido"
        elif inspecao and inspecao < hoje:
            estado = "Vencido"
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = "A vencer"

        if status:
            if status == "ok" and estado != "OK":
                continue
            if status == "a vencer" and estado != "A vencer":
                continue
            if status == "vencido" and estado != "Vencido":
                continue

        def _dias_info(data_ref):
            if not data_ref:
                return None, "-"
            dias = (data_ref - hoje).days
            if dias < 0:
                return dias, f"Vencido há {abs(dias)} dia(s)"
            if dias == 0:
                return dias, "Vence hoje"
            return dias, f"{dias} dia(s)"
        dias_seguro, seguro_restante = _dias_info(seguro)
        dias_inspecao, inspecao_restante = _dias_info(inspecao)
        linhas.append({
            "matricula": item.matricula,
            "seguradora": _seguradora_frota(item) or "-",
            "seguro": seguro.strftime("%d/%m/%Y") if seguro else "-",
            "inspecao": inspecao.strftime("%d/%m/%Y") if inspecao else "-",
            "status": estado,
            "dias_seguro": dias_seguro,
            "dias_inspecao": dias_inspecao,
            "seguro_restante": seguro_restante,
            "inspecao_restante": inspecao_restante,
        })

    return JsonResponse({"linhas": linhas})


@login_required
def exportar_financeiro_excel(request):
    qs, _ = _filtrar_lancamentos_financeiros(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Financeiro"
    ws.append(["ID", "Data Emissão", "Vencimento", "Pagamento", "Fatura", "Fornecedor", "NIF", "Valor Fatura", "Dinheiro", "Cartão", "Transferência", "MBWay", "Nota de Crédito", "Total Pago", "Saldo", "Status"])

    for item in qs:
        ws.append([
            item.id,
            item.data_emissao.strftime("%d/%m/%Y") if getattr(item, "data_emissao", None) else "",
            item.data_vencimento.strftime("%d/%m/%Y") if getattr(item, "data_vencimento", None) else "",
            item.data_pagamento.strftime("%d/%m/%Y") if getattr(item, "data_pagamento", None) else "",
            getattr(item, "numero_fatura", "") or "",
            _fornecedor_nome(item),
            _fornecedor_nif(item),
            float(_to_decimal(getattr(item, "valor_fatura", 0))),
            float(_to_decimal(getattr(item, "dinheiro", 0))),
            float(_to_decimal(getattr(item, "cartao", 0))),
            float(_to_decimal(getattr(item, "transferencia", 0))),
            float(_to_decimal(getattr(item, "mbway", 0))),
            float(-_to_decimal(getattr(item, "nota_credito", 0))),
            float(_total_pago_exibicao_lancamento(item)),
            float(_to_decimal(getattr(item, "saldo_aberto", 0))),
            getattr(item, "status_pagamento", "") or "",
        ])

    _aplicar_cabecalho_relatorio_excel(ws, request, "Relatório Financeiro", "Mapa financeiro com identificação da empresa")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    for row in ws.iter_rows(min_row=7, min_col=8, max_col=15):
        for cell in row:
            cell.number_format = '#,##0.00 [$€-pt-PT]'

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_financeiro.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_faturas_excel(request):
    linhas, totais, _ = _linhas_relatorio_baixa(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Baixa de Faturas"
    ws.append(["ID Lançamento", "Fornecedor", "Número Fatura", "Data Baixa", "Forma de Pagamento", "Utilizador", "Valor Total", "Pago em Dinheiro", "Pago em Cartão", "Pago em Transferência", "Pago em MBWay", "Nota de Crédito", "Valor Pago", "Valor em Aberto"])

    for item in linhas:
        ws.append([
            item.get("lancamento_id", item.get("id", "")),
            item.get("fornecedor", ""),
            item.get("numero_fatura", ""),
            item.get("data_baixa", ""),
            item.get("forma_pagamento", ""),
            item.get("usuario", ""),
            item.get("valor_total", 0),
            item.get("dinheiro", 0),
            item.get("cartao", 0),
            item.get("transferencia", 0),
            item.get("mbway", 0),
            item.get("nota_credito", 0),
            item.get("valor_baixado", 0),
            item.get("saldo_resultante", item.get("em_aberto", 0)),
        ])

    linha_total = ws.max_row + 2
    _estilizar_bloco_totais_excel(ws, linha_total, ["", "", "", "", "", "Totais", "", float(totais.get('total_dinheiro', 0)), float(totais.get('total_cartao', 0)), float(totais.get('total_transferencia', 0)), float(totais.get('total_mbway', 0)), float(totais.get('total_nota_credito', 0)), float(totais.get('total_baixado', 0)), float(totais.get('total_em_aberto', 0))])
    _aplicar_cabecalho_relatorio_excel(ws, request, "Relatório de Baixa de Faturas", "Resumo profissional das baixas efectuadas")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    for row in ws.iter_rows(min_row=7, min_col=7, max_col=14):
        for cell in row:
            cell.number_format = '#,##0.00 [$€-pt-PT]'

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_baixa_faturas.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_revisao_excel(request):
    cab, linhas = _dados_relatorio_revisao(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Revisao Frota'
    ws.append(cab)
    for linha in linhas:
        ws.append(linha)
    _aplicar_cabecalho_relatorio_excel(ws, request, 'Relatório de Revisão de Frota', 'Mapa de revisão de frota com identificação da empresa')
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_revisao_frota.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_rotina_manutencao_excel(request):
    cab, linhas = _dados_rotina_manutencao(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rotina Manutencao'
    ws.append(cab)
    for linha in linhas:
        ws.append(linha)
    _aplicar_cabecalho_relatorio_excel(ws, request, 'Rotina de Manutenção', 'Listagem dos tipos de manutenção configurados')
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rotina_manutencao.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_frota_excel(request):
    cab, linhas = _dados_relatorio_frota(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Frota'
    ws.append(cab)
    for linha in linhas:
        ws.append(linha)
    _aplicar_cabecalho_relatorio_excel(ws, request, 'Relatório de Frota', 'Situação documental da frota')
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_frota.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_manutencao_excel(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()

    qs = ManutencaoFrota.objects.select_related("lancamento", "lancamento__empresa", "frota").all()
    empresa = _empresa_ativa(request)
    if empresa:
        qs = qs.filter(lancamento__empresa=empresa)
    if data_inicio:
        qs = qs.filter(lancamento__data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(lancamento__data_emissao__lte=data_fim)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)

    wb = Workbook()
    ws = wb.active
    ws.title = "Manutencao"
    ws.append(["ID", "Data", "Descrição", "Matrícula", "KM Início", "KM Final", "KM Total", "Valor", "Observação"])

    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    total_valor = Decimal("0.00")
    for item in qs.order_by("-lancamento__data_emissao", "-id"):
        lanc = getattr(item, "lancamento", None)
        valor = _to_decimal(getattr(item, "valor", 0))
        total_valor += valor
        ws.append([
            item.id,
            lanc.data_emissao.strftime("%d/%m/%Y") if getattr(lanc, "data_emissao", None) else "",
            getattr(item, "descricao", "Manutenção") or "Manutenção",
            getattr(getattr(item, "frota", None), "matricula", "") or "-",
            getattr(item, "km_inicio", 0),
            getattr(item, "km_final", 0),
            getattr(item, "km_total", 0),
            float(valor),
            getattr(item, "observacao", "") or "",
        ])

    linha_total = ws.max_row + 2
    _estilizar_bloco_totais_excel(ws, linha_total, ["", "", "", "", "", "Total", float(total_valor), ""])
    _aplicar_cabecalho_relatorio_excel(ws, request, "Relatório de Manutenção", "Mapa de manutenção com identificação da empresa")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    for row in ws.iter_rows(min_row=7, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = '#,##0.00 [$€-pt-PT]'

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_manutencao.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_combustivel_excel(request):
    linhas, resumo_matriculas, resumo = _linhas_relatorio_combustivel(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Combustível"
    ws.append(["Data", "Matrícula", "Fornecedor", "Litros", "KM", "Valor (€)", "KM/L", "€/KM", "Preço/L (€)"])

    for item in linhas:
        ws.append([
            item.get("data", ""),
            item.get("matricula", ""),
            item.get("fornecedor", ""),
            float(item.get("litros", 0) or 0),
            float(item.get("km", 0) or 0),
            float(item.get("valor", 0) or 0),
            float(item.get("media_km_l", 0) or 0),
            float(item.get("media_euro_km", 0) or 0),
            float(item.get("preco_medio_litro", 0) or 0),
        ])

    linha_total = ws.max_row + 2
    _estilizar_bloco_totais_excel(ws, linha_total, [
        "", "", "Totais",
        float(resumo.get("litros_total", 0) or 0),
        float(resumo.get("km_total", 0) or 0),
        float(resumo.get("valor_total", 0) or 0),
        float(resumo.get("media_km_l", 0) or 0),
        float(resumo.get("media_euro_km", 0) or 0),
        float(resumo.get("preco_medio_litro", 0) or 0),
    ])
    _aplicar_cabecalho_relatorio_excel(ws, request, "Relatório de Combustível", "Resumo de abastecimentos por período")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    for row in ws.iter_rows(min_row=7, min_col=6, max_col=9):
        for cell in row:
            cell.number_format = '#,##0.00 [$€-pt-PT]'

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_combustivel.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_documentos_excel(request):
    matricula = (request.GET.get("matricula") or "").strip()
    status = (request.GET.get("status") or "").strip().lower()

    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)

    hoje = date.today()
    limite = hoje + timedelta(days=30)

    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    ws.append(["Matrícula", "Seguradora", "Seguro", "Dias p/ Seguro", "Inspeção", "Dias p/ Inspeção", "Estado"])

    for item in qs.order_by("matricula"):
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)

        estado = "OK"
        if seguro and seguro < hoje:
            estado = "Vencido"
        elif inspecao and inspecao < hoje:
            estado = "Vencido"
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = "A vencer"

        if status:
            if status == "ok" and estado != "OK":
                continue
            if status == "a vencer" and estado != "A vencer":
                continue
            if status == "vencido" and estado != "Vencido":
                continue

        dias_seguro = (seguro - hoje).days if seguro else None
        dias_inspecao = (inspecao - hoje).days if inspecao else None
        ws.append([
            getattr(item, "matricula", "") or "",
            _seguradora_frota(item) or "",
            seguro.strftime("%d/%m/%Y") if seguro else "",
            dias_seguro if dias_seguro is not None else "",
            inspecao.strftime("%d/%m/%Y") if inspecao else "",
            dias_inspecao if dias_inspecao is not None else "",
            estado,
        ])

    _aplicar_cabecalho_relatorio_excel(ws, request, "Relatório de Documentos de Frota", "Situação de seguro e inspeção")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_documentos.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_caixa_excel(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    ordem = (request.GET.get("ordem") or "data").strip().lower()
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor").all())
    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)
    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)
    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)
    qs = qs.order_by("-id") if ordem == "id" else qs.order_by("data_emissao", "id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Folha de Caixa"

    headers = ["ID", "Data Emissão", "Documento/Fatura", "Fornecedor", "Dinheiro (€)", "Cartão (€)", "Nota de Crédito (€)", "Valor Total (€)"]
    ws.append(headers)

    total_dinheiro = Decimal("0.00")
    total_cartao = Decimal("0.00")
    total_nota_credito = Decimal("0.00")
    total_geral = Decimal("0.00")

    for item in qs:
        dinheiro = _to_decimal(getattr(item, "dinheiro", 0))
        cartao = _to_decimal(getattr(item, "cartao", 0))
        transferencia = _to_decimal(getattr(item, "transferencia", 0))
        mbway = _to_decimal(getattr(item, "mbway", 0))
        nota_credito = _to_decimal(getattr(item, "nota_credito", 0))
        total = dinheiro + cartao + transferencia + mbway - nota_credito
        total_dinheiro += dinheiro
        total_cartao += cartao
        total_nota_credito += nota_credito
        total_geral += total
        ws.append([
            item.id,
            item.data_emissao.strftime("%d/%m/%Y") if item.data_emissao else "",
            item.numero_fatura or f"Lançamento #{item.id}",
            _fornecedor_nome(item),
            float(dinheiro),
            float(cartao),
            float(-nota_credito),
            float(total),
        ])

    _aplicar_cabecalho_relatorio_excel(ws, request, "Folha de Caixa", "Resumo contabilístico com total por lançamento")
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    linha_totais = ws.max_row + 2
    _estilizar_bloco_totais_excel(ws, linha_totais, ["Total Dinheiro", float(total_dinheiro), "Total Cartão", float(total_cartao), "Total Nota Crédito", float(-total_nota_credito), "Total Geral", float(total_geral), ""])
    for col in [2,4,6,8]:
        ws.cell(linha_totais, col).number_format = '#,##0.00 [$€-pt-PT]'

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="folha_caixa_profissional.xlsx"'
    wb.save(response)
    return response


@login_required
def relatorio_caixa(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-caixa")

    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    ordem = (request.GET.get("ordem") or "data").strip().lower()
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor", "empresa").all())
    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)
    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)
    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)
    qs = qs.order_by("-id") if ordem == "id" else qs.order_by("data_emissao", "id")

    linhas = []
    totais = {"dinheiro": 0.0, "cartao": 0.0, "nota_credito": 0.0, "total": 0.0, "quantidade_faturas": 0}
    for item in qs:
        dinheiro = _to_decimal(getattr(item, "dinheiro", 0))
        cartao = _to_decimal(getattr(item, "cartao", 0))
        transferencia = _to_decimal(getattr(item, "transferencia", 0))
        mbway = _to_decimal(getattr(item, "mbway", 0))
        nota_credito = _to_decimal(getattr(item, "nota_credito", 0))
        total = dinheiro + cartao + transferencia + mbway - nota_credito
        totais["dinheiro"] += float(dinheiro)
        totais["cartao"] += float(cartao)
        totais["nota_credito"] += float(-nota_credito)
        totais["total"] += float(total)
        totais["quantidade_faturas"] += 1
        linhas.append({
            "id": item.id,
            "data_emissao": item.data_emissao.strftime("%d/%m/%Y") if getattr(item, "data_emissao", None) else "",
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": _fornecedor_nome(item),
            "dinheiro": float(dinheiro),
            "cartao": float(cartao),
            "nota_credito": float(-nota_credito),
            "total": float(total),
        })
    resposta = JsonResponse({"linhas": linhas, "totais": totais})
    resposta["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return resposta



def _agrupar_gastos_por_fornecedor(qs, fornecedor_filtro=""):
    fornecedor_filtro = (fornecedor_filtro or "").strip()
    resumo = {}
    total_faturas = 0

    total_valor = Decimal("0.00")
    total_pago = Decimal("0.00")
    total_nota_credito = Decimal("0.00")
    total_aberto = Decimal("0.00")
    total_vencido = Decimal("0.00")
    hoje = date.today()

    for item in qs.order_by('fornecedor__nome', 'data_emissao', 'id'):
        fornecedor_nome = _fornecedor_nome(item) or 'Sem fornecedor'
        fornecedor_nif = _fornecedor_nif(item) or ''
        if fornecedor_filtro:
            filtro_norm = fornecedor_filtro.casefold()
            if filtro_norm not in fornecedor_nome.casefold() and filtro_norm not in fornecedor_nif.casefold():
                continue

        fornecedor_obj = getattr(item, 'fornecedor', None)
        chave = getattr(fornecedor_obj, 'id', None) or f"nome::{fornecedor_nome.lower()}::{fornecedor_nif}"
        if chave not in resumo:
            resumo[chave] = {
                'fornecedor': fornecedor_nome,
                'nif': fornecedor_nif,
                'quantidade_faturas': 0,
                'primeira_data_raw': None,
                'ultima_data_raw': None,
                'valor_total': Decimal('0.00'),
                'nota_credito': Decimal('0.00'),
                'total_pago': Decimal('0.00'),
                'saldo_aberto': Decimal('0.00'),
                'saldo_vencido': Decimal('0.00'),
            }

        row = resumo[chave]
        valor_documento = _valor_documento_lancamento(item)
        pago = _total_pago_exibicao_lancamento(item)
        saldo = _saldo_documento_lancamento(item)
        vencimento = getattr(item, 'data_vencimento', None)
        saldo_vencido = saldo if (saldo > 0 and vencimento and vencimento < hoje) else Decimal('0.00')

        row['quantidade_faturas'] += 1
        nota_credito_item = _to_decimal(getattr(item, 'nota_credito', 0))
        row['valor_total'] += valor_documento - nota_credito_item
        row['nota_credito'] += nota_credito_item
        row['total_pago'] += pago
        row['saldo_aberto'] += saldo
        row['saldo_vencido'] += saldo_vencido

        data_emissao = getattr(item, 'data_emissao', None)
        if data_emissao:
            if row['primeira_data_raw'] is None or data_emissao < row['primeira_data_raw']:
                row['primeira_data_raw'] = data_emissao
            if row['ultima_data_raw'] is None or data_emissao > row['ultima_data_raw']:
                row['ultima_data_raw'] = data_emissao

        total_faturas += 1
        total_valor += valor_documento - nota_credito_item
        total_nota_credito += nota_credito_item
        total_pago += pago
        total_aberto += saldo
        total_vencido += saldo_vencido

    linhas = []
    for row in resumo.values():
        if row['saldo_aberto'] <= Decimal('0.00') and row['total_pago'] > Decimal('0.00'):
            status_consolidado = 'Pago'
        elif row['saldo_vencido'] > Decimal('0.00'):
            status_consolidado = 'Vencido'
        elif row['saldo_aberto'] > Decimal('0.00') and row['total_pago'] > Decimal('0.00'):
            status_consolidado = 'Parcial'
        elif row['saldo_aberto'] > Decimal('0.00'):
            status_consolidado = 'Em aberto'
        else:
            status_consolidado = 'Sem movimento'
        linhas.append({
            'fornecedor': row['fornecedor'],
            'nif': row['nif'],
            'quantidade_faturas': row['quantidade_faturas'],
            'primeira_emissao': row['primeira_data_raw'].strftime('%d/%m/%Y') if row['primeira_data_raw'] else '',
            'ultima_emissao': row['ultima_data_raw'].strftime('%d/%m/%Y') if row['ultima_data_raw'] else '',
            'valor_total': float(row['valor_total']),
            'nota_credito': float(-row['nota_credito']),
            'total_pago': float(row['total_pago']),
            'saldo_aberto': float(row['saldo_aberto']),
            'saldo_vencido': float(row['saldo_vencido']),
            'status_consolidado': status_consolidado,
        })

    linhas.sort(key=lambda x: (-(x.get('valor_total') or 0), (x.get('fornecedor') or '').lower()))
    totais = {
        'total_fornecedores': len(linhas),
        'total_faturas': total_faturas,
        'valor_total': float(total_valor),
        'nota_credito': float(-total_nota_credito),
        'total_pago': float(total_pago),
        'saldo_aberto': float(total_aberto),
        'saldo_vencido': float(total_vencido),
    }
    return linhas, totais


@login_required
def relatorio_financeiro_fornecedor(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-fornecedor")

    qs, filtros = _filtrar_lancamentos_financeiros(request)
    linhas, totais = _agrupar_gastos_por_fornecedor(qs, request.GET.get("fornecedor"))
    return JsonResponse({
        'cabecalho': 'GASTOS POR FORNECEDOR',
        'linhas': linhas,
        'totais': totais,
        'filtros': filtros,
    })


@login_required
def exportar_financeiro_fornecedor_excel(request):
    qs, _ = _filtrar_lancamentos_financeiros(request)
    linhas, totais = _agrupar_gastos_por_fornecedor(qs, request.GET.get("fornecedor"))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gastos por Fornecedor'
    ws.append(['Fornecedor', 'NIF', 'Quantidade de Faturas', 'Primeira Emissão', 'Última Emissão', 'Valor Total (€)', 'Nota de Crédito (€)', 'Total Pago (€)', 'Em Aberto (€)', 'Vencido (€)', 'Status Consolidado'])

    for row in linhas:
        ws.append([
            row['fornecedor'],
            row['nif'],
            row['quantidade_faturas'],
            row['primeira_emissao'],
            row['ultima_emissao'],
            row['valor_total'],
            row.get('nota_credito', 0),
            row['total_pago'],
            row['saldo_aberto'],
            row['saldo_vencido'],
            row['status_consolidado'],
        ])

    info = _empresa_relatorio_info(request)
    ws.insert_rows(1, 4)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)
    ws.cell(1, 2, info["nome"])
    ws.cell(1, 2).font = Font(size=16, bold=True, color="163A63")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=11)
    ws.cell(2, 2, "Gastos por Fornecedor")
    ws.cell(2, 2).font = Font(size=12, bold=True, color="1F4E78")
    borda = Border(left=Side(style="thin", color="C9D4E4"), right=Side(style="thin", color="C9D4E4"), top=Side(style="thin", color="C9D4E4"), bottom=Side(style="thin", color="C9D4E4"))
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor="0B3A78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = borda
        for col in [6,7,8,9,10]:
            row[col-1].number_format = '#,##0.00 [$€-pt-PT]'
    total_row = ws.max_row + 2
    ws.cell(total_row, 1, 'Total Fornecedores').font = Font(bold=True)
    ws.cell(total_row, 2, totais['total_fornecedores']).font = Font(bold=True)
    ws.cell(total_row, 3, 'Total Faturas').font = Font(bold=True)
    ws.cell(total_row, 4, totais['total_faturas']).font = Font(bold=True)
    ws.cell(total_row, 5, 'Valor Total').font = Font(bold=True)
    ws.cell(total_row, 6, totais['valor_total']).number_format = '#,##0.00 [$€-pt-PT]'
    ws.cell(total_row, 7, totais.get('nota_credito', 0)).number_format = '#,##0.00 [$€-pt-PT]'
    ws.cell(total_row, 8, totais['total_pago']).number_format = '#,##0.00 [$€-pt-PT]'
    ws.cell(total_row, 9, totais['saldo_aberto']).number_format = '#,##0.00 [$€-pt-PT]'
    ws.cell(total_row, 10, totais['saldo_vencido']).number_format = '#,##0.00 [$€-pt-PT]'
    for c in range(1,12):
        ws.cell(total_row, c).border = borda
        ws.column_dimensions[get_column_letter(c)].width = [28,18,18,16,16,18,18,18,18,18,18][c-1]
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="gastos_por_fornecedor.xlsx"'
    wb.save(response)
    return response


@login_required
def relatorio_faturas(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-faturas")

    linhas, totais, filtros = _linhas_relatorio_baixa(request)
    return JsonResponse({
        "cabecalho": "RELATÓRIO DE BAIXA DE FATURAS",
        "linhas": linhas,
        "totais": totais,
        "filtros": filtros,
    })


@login_required
def relatorio_frota(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-frota")

    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()

    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)

    linhas = []
    for item in qs.order_by("matricula"):
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)
        if data_inicio and not ((seguro and seguro >= data_inicio) or (inspecao and inspecao >= data_inicio)):
            continue
        if data_fim and not ((seguro and seguro <= data_fim) or (inspecao and inspecao <= data_fim)):
            continue

        hoje = date.today()
        limite = hoje + timedelta(days=30)
        estado = "OK"
        if (seguro and seguro < hoje) or (inspecao and inspecao < hoje):
            estado = "Vencido"
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = "A vencer"

        linhas.append({
            "matricula": item.matricula,
            "seguro": seguro.strftime("%Y-%m-%d") if seguro else "",
            "inspecao": inspecao.strftime("%Y-%m-%d") if inspecao else "",
            "status": estado,
        })
    return JsonResponse({"linhas": linhas})


@login_required
def relatorio_manutencao(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-manutencao")

    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()

    qs = ManutencaoFrota.objects.select_related("lancamento", "lancamento__empresa", "frota").all()
    empresa = _empresa_ativa(request)
    if empresa:
        qs = qs.filter(lancamento__empresa=empresa)
    if data_inicio:
        qs = qs.filter(lancamento__data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(lancamento__data_emissao__lte=data_fim)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)

    linhas = []
    for item in qs.order_by("-lancamento__data_emissao", "-id"):
        lanc = getattr(item, "lancamento", None)
        linhas.append({
            "id": item.id,
            "data": lanc.data_emissao.strftime("%Y-%m-%d") if getattr(lanc, "data_emissao", None) else "",
            "descricao": getattr(item, "descricao", "Manutenção") or "Manutenção",
            "matricula": getattr(getattr(item, "frota", None), "matricula", "") or "",
            "km_inicio": getattr(item, "km_inicio", 0),
            "km_final": getattr(item, "km_final", 0),
            "km_total": getattr(item, "km_total", 0),
            "valor": float(_to_decimal(getattr(item, "valor", 0))),
            "observacao": getattr(item, "observacao", "") or "",
        })
    return JsonResponse({"linhas": linhas})


def _dados_relatorio_financeiro(request):
    qs, _ = _filtrar_lancamentos_financeiros(request)
    cab = ["ID", "Data Emissão", "Vencimento", "Pagamento", "Fatura", "Fornecedor", "NIF", "Valor Fatura", "Dinheiro", "Cartão", "Transferência", "MBWay", "Nota de Crédito", "Total Pago", "Saldo", "Status"]
    linhas = []
    for item in qs:
        linhas.append([
            item.id,
            item.data_emissao.strftime("%d/%m/%Y") if getattr(item, "data_emissao", None) else "",
            item.data_vencimento.strftime("%d/%m/%Y") if getattr(item, "data_vencimento", None) else "",
            item.data_pagamento.strftime("%d/%m/%Y") if getattr(item, "data_pagamento", None) else "",
            getattr(item, "numero_fatura", "") or "",
            _fornecedor_nome(item), _fornecedor_nif(item),
            float(_to_decimal(getattr(item, "valor_fatura", 0))),
            float(_to_decimal(getattr(item, "dinheiro", 0))),
            float(_to_decimal(getattr(item, "cartao", 0))),
            float(_to_decimal(getattr(item, "transferencia", 0))),
            float(_to_decimal(getattr(item, "mbway", 0))),
            float(-_to_decimal(getattr(item, "nota_credito", 0))),
            float(_total_pago_exibicao_lancamento(item)),
            float(_to_decimal(getattr(item, "saldo_aberto", 0))),
            getattr(item, "status_pagamento", "") or "",
        ])
    return cab, linhas


def _dados_relatorio_fornecedor(request):
    qs, _ = _filtrar_lancamentos_financeiros(request)
    cab = ["Fornecedor", "NIF", "Fatura", "Emissão", "Vencimento", "Valor", "Nota de Crédito", "Pago", "Saldo", "Status"]
    linhas = []
    for item in qs:
        linhas.append([
            _fornecedor_nome(item), _fornecedor_nif(item), getattr(item, 'numero_fatura', '') or '',
            item.data_emissao.strftime('%d/%m/%Y') if getattr(item, 'data_emissao', None) else '',
            item.data_vencimento.strftime('%d/%m/%Y') if getattr(item, 'data_vencimento', None) else '',
            float(_to_decimal(getattr(item, 'valor_fatura', 0))), float(-_to_decimal(getattr(item, 'nota_credito', 0))), float(_total_pago_exibicao_lancamento(item)),
            float(_to_decimal(getattr(item, 'saldo_aberto', 0))), getattr(item, 'status_pagamento', '') or ''
        ])
    return cab, linhas


def _dados_relatorio_faturas(request):
    linhas, totais, _ = _linhas_relatorio_baixa(request)
    cab = ["ID Lançamento", "Fornecedor", "Número Fatura", "Data Baixa", "Forma de Pagamento", "Utilizador", "Valor Total", "Pago Dinheiro", "Pago Cartão", "Pago Transferência", "Pago MBWay", "Nota de Crédito", "Valor Pago", "Em Aberto"]
    rows = [[item.get("lancamento_id", item.get("id", "")), item.get("fornecedor", ""), item.get("numero_fatura", ""), item.get("data_baixa", ""), item.get("forma_pagamento", ""), item.get("usuario", ""), item.get("valor_total", 0), item.get("dinheiro", 0), item.get("cartao", 0), item.get("transferencia", 0), item.get("mbway", 0), item.get("nota_credito", 0), item.get("valor_baixado", 0), item.get("saldo_resultante", item.get("em_aberto", 0))] for item in linhas]
    rows.append(['', '', '', '', '', 'Totais', '', totais.get('total_dinheiro', 0), totais.get('total_cartao', 0), totais.get('total_transferencia', 0), totais.get('total_mbway', 0), totais.get('total_nota_credito', 0), totais.get('total_baixado', 0), totais.get('total_em_aberto', 0)])
    return cab, rows


def _dados_relatorio_frota(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()
    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)
    hoje = date.today()
    limite = hoje + timedelta(days=30)
    cab = ["Matrícula", "Seguro", "Inspeção", "Estado"]
    rows = []
    for item in qs.order_by('matricula'):
        seguro = getattr(item, 'seguro', None)
        inspecao = getattr(item, 'inspecao', None)
        if data_inicio and not ((seguro and seguro >= data_inicio) or (inspecao and inspecao >= data_inicio)):
            continue
        if data_fim and not ((seguro and seguro <= data_fim) or (inspecao and inspecao <= data_fim)):
            continue
        estado = 'OK'
        if (seguro and seguro < hoje) or (inspecao and inspecao < hoje):
            estado = 'Vencido'
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = 'A vencer'
        rows.append([item.matricula, seguro.strftime('%d/%m/%Y') if seguro else '', inspecao.strftime('%d/%m/%Y') if inspecao else '', estado])
    return cab, rows


def _dados_relatorio_manutencao(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    matricula = (request.GET.get("matricula") or "").strip()
    qs = ManutencaoFrota.objects.select_related("lancamento", "lancamento__empresa", "frota").all()
    empresa = _empresa_ativa(request)
    if empresa:
        qs = qs.filter(lancamento__empresa=empresa)
    if data_inicio:
        qs = qs.filter(lancamento__data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(lancamento__data_emissao__lte=data_fim)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)
    cab = ["ID", "Data", "Descrição", "Matrícula", "KM Início", "KM Final", "KM Total", "Valor", "Observação"]
    rows = []
    total = Decimal('0.00')
    for item in qs.order_by('-lancamento__data_emissao', '-id'):
        lanc = getattr(item, 'lancamento', None)
        valor = _to_decimal(getattr(item, 'valor', 0))
        total += valor
        rows.append([item.id, lanc.data_emissao.strftime('%d/%m/%Y') if getattr(lanc, 'data_emissao', None) else '', getattr(item, 'descricao', 'Manutenção') or 'Manutenção', getattr(getattr(item, 'frota', None), 'matricula', '') or '-', getattr(item, 'km_inicio', 0), getattr(item, 'km_final', 0), getattr(item, 'km_total', 0), float(valor), getattr(item, 'observacao', '') or ''])
    rows.append(['', '', '', '', '', 'Total', '', float(total), ''])
    return cab, rows


def _dados_relatorio_caixa(request):
    data_inicio = _to_date(request.GET.get("data_inicio") or request.GET.get("data_de") or request.GET.get("inicio"))
    data_fim = _to_date(request.GET.get("data_fim") or request.GET.get("data_ate") or request.GET.get("fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    ordem = (request.GET.get("ordem") or "data").strip().lower()
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related('fornecedor').all())
    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)
    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)
    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)
    qs = qs.order_by('-id') if ordem == 'id' else qs.order_by('data_emissao', 'id')
    cab = ["ID", "Data Emissão", "Documento/Fatura", "Fornecedor", "Dinheiro (€)", "Cartão (€)", "Nota de Crédito (€)", "Valor Total (€)"]
    rows = []
    td = Decimal('0.00')
    tc = Decimal('0.00')
    tnc = Decimal('0.00')
    tg = Decimal('0.00')
    for item in qs:
        dinheiro = _to_decimal(getattr(item, 'dinheiro', 0))
        cartao = _to_decimal(getattr(item, 'cartao', 0))
        transferencia = _to_decimal(getattr(item, 'transferencia', 0))
        mbway = _to_decimal(getattr(item, 'mbway', 0))
        nota_credito = _to_decimal(getattr(item, 'nota_credito', 0))
        total = dinheiro + cartao + transferencia + mbway - nota_credito
        td += dinheiro
        tc += cartao
        tnc += nota_credito
        tg += total
        rows.append([item.id, item.data_emissao.strftime('%d/%m/%Y') if item.data_emissao else '', item.numero_fatura or f'Lançamento #{item.id}', _fornecedor_nome(item), float(dinheiro), float(cartao), float(-nota_credito), float(total)])
    rows.append(['', '', '', 'Totais', float(td), float(tc), float(-tnc), float(tg)])
    return cab, rows


def _dados_relatorio_combustivel(request):
    linhas, _, resumo = _linhas_relatorio_combustivel(request)
    cab = ["Data", "Matrícula", "Fornecedor", "Litros", "KM", "Valor (€)", "KM/L", "€/KM", "Preço/L (€)"]
    rows = [[item.get('data', ''), item.get('matricula', ''), item.get('fornecedor', ''), float(item.get('litros', 0) or 0), float(item.get('km', 0) or 0), float(item.get('valor', 0) or 0), float(item.get('media_km_l', 0) or 0), float(item.get('media_euro_km', 0) or 0), float(item.get('preco_medio_litro', 0) or 0)] for item in linhas]
    rows.append(['', '', 'Totais', float(resumo.get('litros_total', 0) or 0), float(resumo.get('km_total', 0) or 0), float(resumo.get('valor_total', 0) or 0), float(resumo.get('media_km_l', 0) or 0), float(resumo.get('media_euro_km', 0) or 0), float(resumo.get('preco_medio_litro', 0) or 0)])
    return cab, rows


def _dados_relatorio_documentos(request):
    matricula = (request.GET.get('matricula') or '').strip()
    status = (request.GET.get('status') or '').strip().lower()
    qs = _filtrar_empresa_generico(request, Frota.objects.all(), model=Frota)
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)
    hoje = date.today()
    limite = hoje + timedelta(days=30)
    cab = ["Matrícula", "Seguradora", "Seguro", "Dias p/ Seguro", "Inspeção", "Dias p/ Inspeção", "Estado"]
    rows = []
    for item in qs.order_by('matricula'):
        seguro = getattr(item, 'seguro', None)
        inspecao = getattr(item, 'inspecao', None)
        estado = 'OK'
        if seguro and seguro < hoje:
            estado = 'Vencido'
        elif inspecao and inspecao < hoje:
            estado = 'Vencido'
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = 'A vencer'
        if status:
            if status == 'ok' and estado != 'OK':
                continue
            if status == 'a vencer' and estado != 'A vencer':
                continue
            if status == 'vencido' and estado != 'Vencido':
                continue
        dias_seguro = (seguro - hoje).days if seguro else ''
        dias_inspecao = (inspecao - hoje).days if inspecao else ''
        rows.append([getattr(item, 'matricula', '') or '', _seguradora_frota(item) or '', seguro.strftime('%d/%m/%Y') if seguro else '', dias_seguro, inspecao.strftime('%d/%m/%Y') if inspecao else '', dias_inspecao, estado])
    return cab, rows


def _dados_relatorio_revisao(request):
    matricula = (request.GET.get('matricula') or '').strip()
    data_inicio = _to_date(request.GET.get('data_inicio'))
    data_fim = _to_date(request.GET.get('data_fim'))
    qs = RevisaoFrota.objects.select_related('frota', 'funcionario').all()
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)
    if data_inicio:
        qs = qs.filter(data_ultima_revisao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_ultima_revisao__lte=data_fim)
    cab = ['ID', 'Matrícula', 'Marca', 'Modelo', 'Data da Última Revisão', 'KMs da Última Revisão', 'KMs Rodados', 'KMs de Previsão', 'KM para Fazer a Revisão', 'Funcionário', 'Observação']
    rows = []
    for item in qs.order_by('-data_ultima_revisao', '-id'):
        rows.append([
            item.id,
            getattr(getattr(item, 'frota', None), 'matricula', '') or '',
            getattr(getattr(item, 'frota', None), 'marca', '') or '',
            getattr(getattr(item, 'frota', None), 'modelo', '') or '',
            item.data_ultima_revisao.strftime('%d/%m/%Y') if item.data_ultima_revisao else '',
            int(item.km_ultima_revisao or 0),
            int(item.km_rodados or 0),
            int(item.kms_previsao or 0),
            int(item.km_para_revisao or 0),
            getattr(getattr(item, 'funcionario', None), 'nome', '') or '',
            item.observacao or '',
        ])
    return cab, rows


def _dados_rotina_manutencao(request):
    cfg = _manutencao_config()
    cab = ['Chave', 'Tipo de manutenção']
    rows = [[item.get('key', ''), item.get('nome', '')] for item in cfg.get('tipos', [])]
    return cab, rows


def _dados_exportacao(tipo, request):
    mapa = {
        'financeiro': ('Relatório Financeiro', 'Mapa financeiro com identificação da empresa', 'relatorio_financeiro.pdf', _dados_relatorio_financeiro),
        'fornecedor': ('Relatório por Fornecedor', 'Análise financeira por fornecedor', 'relatorio_fornecedor.pdf', _dados_relatorio_fornecedor),
        'faturas': ('Relatório de Baixa de Faturas', 'Resumo profissional das baixas efectuadas', 'relatorio_baixa_faturas.pdf', _dados_relatorio_faturas),
        'frota': ('Relatório de Frota', 'Situação documental da frota', 'relatorio_frota.pdf', _dados_relatorio_frota),
        'manutencao': ('Relatório de Manutenção', 'Mapa de manutenção com identificação da empresa', 'relatorio_manutencao.pdf', _dados_relatorio_manutencao),
        'revisao': ('Relatório de Revisão de Frota', 'Mapa de revisão de frota com identificação da empresa', 'relatorio_revisao_frota.pdf', _dados_relatorio_revisao),
        'caixa': ('Folha de Caixa', 'Resumo contabilístico com total por lançamento', 'folha_caixa.pdf', _dados_relatorio_caixa),
        'combustivel': ('Relatório de Combustível', 'Resumo de abastecimentos por período', 'relatorio_combustivel.pdf', _dados_relatorio_combustivel),
        'documentos': ('Relatório de Documentos de Frota', 'Situação de seguro e inspeção', 'relatorio_documentos.pdf', _dados_relatorio_documentos),
        'rotina_manutencao': ('Rotina de Manutenção', 'Listagem dos tipos de manutenção configurados', 'rotina_manutencao.pdf', _dados_rotina_manutencao),
    }
    return mapa[tipo]


@login_required
def exportar_pdf_generico(request, tipo):
    titulo, subtitulo, nome_arquivo, func = _dados_exportacao(tipo, request)
    cab, linhas = func(request)
    return _exportar_tabela_pdf(request, titulo, cab, linhas, nome_arquivo, subtitulo)


@login_required
def imprimir_generico(request, tipo):
    titulo, subtitulo, _, func = _dados_exportacao(tipo, request)
    cab, linhas = func(request)
    return _imprimir_tabela_html(request, titulo, cab, linhas, subtitulo)


# ==== Permission wrappers ====
fornecedores_list = _require_perm("fornecedor", "view")(fornecedores_list)
fornecedores_salvar = _require_perm("fornecedor", "edit")(fornecedores_salvar)
fornecedor_excluir = _require_perm("fornecedor", "delete")(fornecedor_excluir)
funcionarios_list = _require_perm("funcionario", "view")(funcionarios_list)
funcionarios_salvar = _require_perm("funcionario", "edit")(funcionarios_salvar)
funcionario_excluir = _require_perm("funcionario", "delete")(funcionario_excluir)
frota_list = _require_perm("frota", "view")(frota_list)
frota_salvar = _require_perm("frota", "edit")(frota_salvar)
frota_excluir = _require_perm("frota", "delete")(frota_excluir)
combustiveis_list = _require_perm("combustivel", "view")(combustiveis_list)
combustiveis_salvar = _require_perm("combustivel", "edit")(combustiveis_salvar)
combustivel_excluir = _require_perm("combustivel", "delete")(combustivel_excluir)
manutencao_botao_config = _require_perm("manutencao", "view")(manutencao_botao_config)
manutencao_botao_salvar = _require_perm("manutencao", "edit")(manutencao_botao_salvar)
manutencao_botao_excluir = _require_perm("manutencao", "delete")(manutencao_botao_excluir)
empresas_list = _require_perm("empresa", "view")(empresas_list)
empresas_salvar = _require_perm("empresa", "edit")(empresas_salvar)
empresa_excluir = _require_perm("empresa", "delete")(empresa_excluir)
usuarios_list = _require_perm("usuario", "view")(usuarios_list)
usuarios_salvar = _require_perm("usuario", "edit")(usuarios_salvar)
usuario_excluir = _require_perm("usuario", "delete")(usuario_excluir)
usuario_alterar_senha = _require_perm("usuario", "edit")(usuario_alterar_senha)
lancamentos_salvar = _require_lancamentos_salvar_perm(lancamentos_salvar)
lancamento_detalhe = _require_perm("lancamentos", "view")(lancamento_detalhe)
lancamento_excluir = _require_perm("lancamentos", "delete")(lancamento_excluir)
lancamento_baixa = _require_perm("baixa_faturas", "edit")(lancamento_baixa)
consulta_lancamentos = _require_perm("consulta", "view")(consulta_lancamentos)
relatorio_financeiro = _require_perm("relatorio_financeiro", "view")(relatorio_financeiro)
relatorio_financeiro_fornecedor = _require_perm("relatorio_fornecedor", "view")(relatorio_financeiro_fornecedor)
exportar_financeiro_excel = _require_perm("relatorio_financeiro", "export")(exportar_financeiro_excel)
exportar_financeiro_fornecedor_excel = _require_perm("relatorio_fornecedor", "export")(exportar_financeiro_fornecedor_excel)
exportar_pdf_generico = _require_perm('relatorios', 'export')(exportar_pdf_generico)
imprimir_generico = _require_perm('relatorios', 'export')(imprimir_generico)
exportar_rotina_manutencao_excel = _require_perm('manutencao', 'export')(exportar_rotina_manutencao_excel)
relatorio_faturas = _require_perm("relatorio_faturas", "view")(relatorio_faturas)
exportar_faturas_excel = _require_perm("relatorio_faturas", "export")(exportar_faturas_excel)
relatorio_manutencao = _require_perm("relatorio_manutencao", "view")(relatorio_manutencao)
exportar_manutencao_excel = _require_perm("relatorio_manutencao", "export")(exportar_manutencao_excel)
relatorio_caixa = _require_perm("relatorio_caixa", "view")(relatorio_caixa)
exportar_caixa_excel = _require_perm("relatorio_caixa", "export")(exportar_caixa_excel)
relatorio_combustivel = _require_perm("relatorio_combustivel", "view")(relatorio_combustivel)
exportar_combustivel_excel = _require_perm("relatorio_combustivel", "export")(exportar_combustivel_excel)
relatorio_documentos = _require_perm("relatorio_documentos", "view")(relatorio_documentos)
exportar_documentos_excel = _require_perm("relatorio_documentos", "export")(exportar_documentos_excel)


# ==== ERP unificado: clientes, artigos, faturação e PDFs ====

def _table_exists(table_name):
    try:
        return table_name in connection.introspection.table_names()
    except Exception:
        return False


def _json_ok(data=None):
    payload = {"ok": True}
    if data:
        payload.update(data)
    return JsonResponse(payload)


def _parse_json_request(request):
    body = _json_body(request)
    if isinstance(body, dict):
        return body
    return {}


def _preview_numero_serie(serie):
    if not serie:
        return ""
    prefixo = (getattr(serie, "prefixo_documento", "") or getattr(serie, "codigo", "") or "DOC").strip()
    casas = max(int(getattr(serie, "casas_numero", 4) or 4), 1)
    seq = f"{int(getattr(serie, 'proximo_numero', 1) or 1):0{casas}d}"
    if getattr(serie, "usar_ano", True):
        sep = getattr(serie, "separador", "/") or "/"
        return f"{prefixo} {getattr(serie, 'ano', timezone.localdate().year)}{sep}{seq}"
    return f"{prefixo} {seq}"


@login_required
@require_GET
def clientes_list(request):
    if not _table_exists("Elion_cliente"):
        return JsonResponse({"linhas": []})
    linhas = []
    for item in _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente).order_by("nome"):
        linhas.append({"id": item.id, "nif": item.nif, "nome": item.nome, "email": item.email or "", "contato": item.telefone or "", "morada": item.morada or ""})
    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def clientes_busca(request):
    if not _table_exists("Elion_cliente"):
        return JsonResponse({"linhas": []})
    q = (request.GET.get("q") or "").strip()
    qs = _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente).order_by("nome")
    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(nif__icontains=q))
    linhas = []
    for item in qs[:12]:
        linhas.append({"id": item.id, "nif": item.nif, "nome": item.nome, "email": item.email or "", "contato": item.telefone or "", "morada": item.morada or ""})
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def clientes_salvar(request):
    data = _parse_json_request(request)
    obj = _get_objeto_empresa_or_404(request, Cliente, data["id"]) if data.get("id") else Cliente()
    _atribuir_empresa_ativa(request, obj)
    nif = (data.get("nif", "") or "").strip()
    nome = (data.get("nome", "") or "").strip()
    existente = _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente)
    if obj.pk:
        existente = existente.exclude(pk=obj.pk)
    duplicado = None
    if nif:
        duplicado = existente.filter(nif__iexact=nif).first()
    if not duplicado and nome:
        duplicado = existente.filter(nome__iexact=nome).first()
    if duplicado:
        return JsonResponse({"error": f"Cliente já cadastrado: {duplicado.nome} ({duplicado.nif}).", "duplicado_id": duplicado.id}, status=400)
    obj.nif = nif
    obj.nome = nome
    obj.morada = data.get("morada", "")
    obj.cidade = data.get("cidade", "")
    obj.telefone = data.get("telefone") or data.get("contato", "")
    obj.email = data.get("email") or None
    obj.limite_credito = _to_decimal(data.get("limite_credito"))
    erro = _save_model_or_error(obj, "Já existe um cliente com este contribuinte.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return _json_ok({"id": obj.id, "nome": obj.nome, "nif": obj.nif})


@login_required
@require_POST
def cliente_excluir(request, id):
    if not _table_exists("Elion_cliente"):
        raise Http404
    _get_objeto_empresa_or_404(request, Cliente, id).delete()
    return _json_ok()


@login_required
@require_GET
def vendedores_list(request):
    if not _table_exists("Elion_vendedor"):
        return JsonResponse({"linhas": []})
    return JsonResponse({"linhas": list(_filtrar_empresa_generico(request, Vendedor.objects.all(), model=Vendedor).values("id", "nome", "email", "telefone", "comissao_padrao").order_by("nome"))})


@login_required
@require_POST
def vendedores_salvar(request):
    data = _parse_json_request(request)
    obj = _get_objeto_empresa_or_404(request, Vendedor, data["id"]) if data.get("id") else Vendedor()
    _atribuir_empresa_ativa(request, obj)
    obj.nome = data.get("nome", "")
    obj.email = data.get("email") or None
    obj.telefone = data.get("telefone") or data.get("contato", "")
    obj.comissao_padrao = _to_decimal(data.get("comissao_padrao"))
    erro = _save_model_or_error(obj)
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return _json_ok({"id": obj.id})


@login_required
@require_GET
def artigos_list(request):
    if not _table_exists("Elion_artigo"):
        return JsonResponse({"linhas": []})
    linhas = []
    for item in _filtrar_empresa_generico(request, Artigo.objects.select_related("categoria", "armazem"), model=Artigo).order_by("nome"):
        linhas.append({
            "id": item.id,
            "codigo": item.codigo,
            "nome": item.nome,
            "tipo": item.tipo,
            "categoria": item.categoria.nome if item.categoria else "",
            "armazem": item.armazem.nome if item.armazem else "",
            "preco_venda": float(item.preco_venda or 0),
            "preco": float(item.preco_venda or 0),
            "iva": float(getattr(item, "taxa_iva_padrao", 23) or 23),
            "unidade_medida": getattr(item, "unidade_medida", "UN") or "UN",
            "stock_atual": float(item.stock_atual or 0),
            "localizacao_prateleira": item.localizacao_prateleira or "",
        })
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def artigos_salvar(request):
    data = _parse_json_request(request)
    obj = _get_objeto_empresa_or_404(request, Artigo, data["id"]) if data.get("id") else Artigo()
    _atribuir_empresa_ativa(request, obj)
    obj.codigo = data.get("codigo", "")
    obj.nome = data.get("nome", "")
    obj.tipo = data.get("tipo") or "produto"
    obj.preco_venda = _to_decimal(data.get("preco_venda"))
    obj.preco_custo = _to_decimal(data.get("preco_custo"))
    obj.taxa_iva_padrao = _to_decimal(data.get("taxa_iva_padrao"), "23.00")
    obj.unidade_medida = (data.get("unidade_medida") or "UN")[:20]
    obj.stock_atual = _to_decimal(data.get("stock_atual"))
    obj.stock_minimo = _to_decimal(data.get("stock_minimo"))
    obj.localizacao_prateleira = data.get("localizacao_prateleira", "")
    if data.get("categoria_id"):
        obj.categoria_id = data.get("categoria_id")
    if data.get("armazem_id"):
        obj.armazem_id = data.get("armazem_id")
    erro = _save_model_or_error(obj, "Já existe um artigo com este código.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return _json_ok({"id": obj.id})


@login_required
@require_GET
def series_documentos_list(request):
    if not _table_exists("Elion_seriedocumento"):
        return JsonResponse({"linhas": []})
    empresa = _empresa_ativa(request)
    qs = SerieDocumento.objects.all()
    if empresa:
        qs = qs.filter(Q(empresa=empresa) | Q(empresa__isnull=True))
    linhas = list(qs.values("id", "codigo", "descricao", "tipo_documento", "prefixo_documento", "usar_ano", "separador", "casas_numero", "ano", "proximo_numero", "codigo_validacao", "ativa").order_by("codigo", "ano"))
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def series_documentos_salvar(request):
    data = _parse_json_request(request)
    obj = get_object_or_404(SerieDocumento, id=data["id"]) if data.get("id") else SerieDocumento()
    obj.empresa = _empresa_ativa(request)
    obj.codigo = data.get("codigo", "FT")
    obj.descricao = data.get("descricao", "")
    obj.tipo_documento = data.get("tipo_documento") or obj.codigo
    obj.prefixo_documento = (data.get("prefixo_documento") or data.get("codigo") or "").strip()
    obj.usar_ano = bool(data.get("usar_ano", True))
    obj.separador = (data.get("separador") or "/")[:5]
    obj.casas_numero = int(data.get("casas_numero") or 4)
    obj.ano = int(data.get("ano") or timezone.localdate().year)
    obj.proximo_numero = int(data.get("proximo_numero") or 1)
    obj.codigo_validacao = (data.get("codigo_validacao") or "").strip()[:20]
    obj.ativa = bool(data.get("ativa", True))
    erro = _save_model_or_error(obj, "Já existe uma série com este código/ano para a empresa.")
    if erro:
        return JsonResponse({"error": erro}, status=400)
    return _json_ok({"id": obj.id})


def _documento_precisa_nota_credito(documento):
    hoje = timezone.localdate()
    return documento.data_emissao and (documento.data_emissao.month != hoje.month or documento.data_emissao.year != hoje.year)


def _enriquecer_documentos_contexto(itens):
    for item in itens:
        precisa_nc = _documento_precisa_nota_credito(item)
        item.pode_cancelar = not precisa_nc and item.estado != "anulado"
        item.pode_nota_credito = precisa_nc and item.estado != "anulado"
        item.total_liquido = item.subtotal
    return itens


@login_required
def documentos_venda_list(request, modo="faturacao"):
    if not _table_exists("Elion_documentovenda"):
        return render(request, "Elion/documentos_venda.html", {"linhas": [], "series": [], "clientes": [], "artigos": [], "modo": modo})
    empresa = _empresa_ativa(request)
    qs = DocumentoVenda.objects.select_related("cliente", "serie")
    if empresa:
        qs = qs.filter(Q(empresa=empresa) | Q(empresa__isnull=True))
    if modo == "orcamentos":
        qs = qs.filter(serie__tipo_documento="OR")
    elif modo == "faturacao":
        qs = qs.exclude(serie__tipo_documento="OR")
    linhas = list(qs.order_by("-data_emissao", "-id")[:100])
    _enriquecer_documentos_contexto(linhas)
    series_qs = SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True).order_by("codigo")
    if modo == "orcamentos":
        series_qs = series_qs.filter(tipo_documento="OR")
    elif modo == "faturacao":
        series_qs = series_qs.filter(tipo_documento__in=["FT", "FS", "FR"])
        if not series_qs.exists():
            fallback = SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True).exclude(tipo_documento="OR").order_by("codigo")
            series_qs = fallback
    series_com_preview = [{"obj": s, "preview": _preview_numero_serie(s)} for s in series_qs]
    artigos_lista = list(_filtrar_empresa_generico(request, Artigo.objects.filter(ativo=True), model=Artigo).order_by("nome")[:200].values("id", "codigo", "nome", "preco_venda", "taxa_iva_padrao", "unidade_medida"))
    for item in artigos_lista:
        item["preco"] = float(item.pop("preco_venda") or 0)
        item["iva"] = float(item.pop("taxa_iva_padrao") or 23)
    context = {
        "linhas": linhas,
        "series": series_qs,
        "series_com_preview": series_com_preview,
        "pode_emitir_faturacao": _usuario_tem_permissao(request, "faturacao", "create"),
        "pode_exportar_faturacao": _usuario_tem_permissao(request, "faturacao", "export"),
        "pode_configurar_series": _usuario_tem_permissao(request, "series", "edit"),
        "clientes": _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente).order_by("nome"),
        "artigos": _filtrar_empresa_generico(request, Artigo.objects.filter(ativo=True), model=Artigo).order_by("nome")[:200],
        "artigos_json": json.dumps(artigos_lista, ensure_ascii=False),
        "hoje": timezone.localdate(),
        "empresa_ativa": empresa,
        "series_orcamento": SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento="OR").order_by("codigo"),
        "series_fatura": SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento__in=["FT", "FS", "FR"]).order_by("codigo"),
        "modo": modo,
        "titulo_pagina": "Orçamentos" if modo == "orcamentos" else "Faturação",
        "subtitulo_pagina": "Crie, guarde rascunhos, cancele ou converta em fatura com séries independentes." if modo == "orcamentos" else "Emita documentos comerciais com séries próprias, produtos e serviços.",
    }
    return render(request, "Elion/documentos_venda.html", context)


@login_required
def orcamentos_list(request):
    return documentos_venda_list(request, modo="orcamentos")


@login_required
def faturacao_list(request):
    return documentos_venda_list(request, modo="faturacao")


@login_required
@require_POST
def documentos_venda_salvar(request):
    data = _parse_json_request(request)
    empresa = _empresa_ativa(request)
    serie = get_object_or_404(SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True)), id=data.get("serie_id"))
    cliente = get_object_or_404(_filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente), id=data.get("cliente_id"))
    documento = DocumentoVenda(
        empresa=empresa,
        cliente=cliente,
        vendedor_id=data.get("vendedor_id") or None,
        serie=serie,
        template_id=data.get("template_id") or None,
        centro_custo_id=data.get("centro_custo_id") or None,
        tipo_documento=serie.tipo_documento,
        estado=(data.get("estado") or ("rascunho" if serie.tipo_documento == "OR" else "emitido")),
        data_emissao=_to_date(data.get("data_emissao")) or timezone.localdate(),
        data_vencimento=_to_date(data.get("data_vencimento")),
        observacoes=data.get("observacoes", ""),
        criado_por=request.user,
    )
    erro = _save_model_or_error(documento)
    if erro:
        return JsonResponse({"error": erro}, status=400)
    linhas = data.get("linhas") or []
    for linha in linhas:
        artigo = None
        if linha.get("artigo_id"):
            artigo = _filtrar_empresa_generico(request, Artigo.objects.all(), model=Artigo).filter(id=linha.get("artigo_id")).first()
        item = DocumentoVendaLinha(
            documento=documento,
            artigo=artigo,
            descricao=linha.get("descricao") or (artigo.nome if artigo else "Item"),
            quantidade=_to_decimal(linha.get("quantidade"), "1.00"),
            preco_unitario=_to_decimal(linha.get("preco_unitario")),
            taxa_iva=_to_decimal(linha.get("taxa_iva"), str(getattr(artigo, "taxa_iva_padrao", "23.00") if artigo else "23.00")),
        )
        item.save()
    documento.recalcular_totais()
    return _json_ok({"id": documento.id, "numero_documento": documento.numero_documento})


@login_required
def documento_venda_pdf(request, id):
    documento = get_object_or_404(DocumentoVenda.objects.select_related("empresa", "cliente", "serie", "template"), id=id)
    linhas = list(documento.linhas.select_related("artigo").all())
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{documento.numero_documento or "documento"}.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        largura, altura = A4
        y = altura - 20 * mm
        empresa = documento.empresa
        if empresa and getattr(empresa, "logo", None):
            try:
                p.drawImage(empresa.logo.path, 15 * mm, y - 15 * mm, width=28 * mm, height=15 * mm, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        p.setFont("Helvetica-Bold", 15)
        p.drawString(50 * mm, y, documento.numero_documento or "Documento")
        y -= 8 * mm
        p.setFont("Helvetica", 9)
        if empresa:
            p.drawString(15 * mm, y, f"Empresa: {empresa.nome} | NIF: {empresa.nif}")
            y -= 5 * mm
            p.drawString(15 * mm, y, f"Morada: {empresa.morada or '-'} | Contacto: {empresa.contato or '-'}")
            y -= 5 * mm
        p.drawString(15 * mm, y, f"Cliente: {documento.cliente.nome} | NIF: {documento.cliente.nif}")
        y -= 5 * mm
        p.drawString(15 * mm, y, f"Data emissão: {documento.data_emissao:%d/%m/%Y} | Vencimento: {documento.data_vencimento.strftime('%d/%m/%Y') if documento.data_vencimento else '-'}")
        y -= 10 * mm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(15 * mm, y, "Descrição")
        p.drawString(115 * mm, y, "Qtd.")
        p.drawString(135 * mm, y, "P.Unit")
        p.drawString(165 * mm, y, "Total")
        y -= 4 * mm
        p.line(15 * mm, y, 195 * mm, y)
        y -= 6 * mm
        p.setFont("Helvetica", 9)
        for linha in linhas:
            if y < 35 * mm:
                p.showPage()
                y = altura - 20 * mm
                p.setFont("Helvetica", 9)
            p.drawString(15 * mm, y, (linha.descricao or "")[:55])
            p.drawRightString(128 * mm, y, f"{linha.quantidade:.2f}")
            p.drawRightString(156 * mm, y, f"{linha.preco_unitario:.2f} €")
            p.drawRightString(194 * mm, y, f"{linha.total:.2f} €")
            y -= 5 * mm
        y -= 5 * mm
        p.line(120 * mm, y, 195 * mm, y)
        y -= 6 * mm
        p.drawRightString(194 * mm, y, f"Subtotal: {documento.subtotal:.2f} €")
        y -= 5 * mm
        p.drawRightString(194 * mm, y, f"IVA: {documento.total_iva:.2f} €")
        y -= 5 * mm
        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(194 * mm, y, f"Total: {documento.total:.2f} €")
        y -= 10 * mm
        p.setFont("Helvetica", 8)
        p.drawString(15 * mm, y, f"Estado: {documento.get_estado_display()} | Série: {documento.serie.codigo}/{documento.serie.ano}")
        p.showPage()
        p.save()
        return response
    except Exception:
        html = render_to_string("Elion/documento_venda_print.html", {"documento": documento, "linhas": linhas})
        return HttpResponse(html)


@login_required
@require_POST
def documento_venda_cancelar(request, id):
    data = _parse_json_request(request)
    documento = get_object_or_404(DocumentoVenda.objects.select_related("serie", "cliente", "empresa", "template", "centro_custo", "vendedor"), id=id)
    motivo = (data.get("motivo") or "").strip()
    if _documento_precisa_nota_credito(documento):
        serie_nc = SerieDocumento.objects.filter(Q(empresa=documento.empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento="NC").order_by("codigo").first()
        if not serie_nc:
            return JsonResponse({"error": "O documento já passou o mês de emissão. Configure uma série de Nota de Crédito para continuar."}, status=400)
        novo = DocumentoVenda.objects.create(
            empresa=documento.empresa, cliente=documento.cliente, vendedor=documento.vendedor,
            serie=serie_nc, template=documento.template, centro_custo=documento.centro_custo,
            tipo_documento=serie_nc.tipo_documento, estado="emitido", data_emissao=timezone.localdate(),
            data_vencimento=None,
            observacoes=(f"Nota de crédito gerada para o documento {documento.numero_documento}. Motivo: {motivo}"),
            criado_por=request.user,
        )
        for linha in documento.linhas.all():
            DocumentoVendaLinha.objects.create(documento=novo, artigo=linha.artigo, descricao=linha.descricao, quantidade=linha.quantidade, preco_unitario=linha.preco_unitario, taxa_iva=linha.taxa_iva)
        novo.recalcular_totais()
        return _json_ok({"id": novo.id, "estado": novo.estado, "numero_documento": novo.numero_documento, "tipo": "nota_credito"})
    documento.estado = "anulado"
    documento.observacoes = ((documento.observacoes or "") + f"\nDocumento anulado. Motivo: {motivo}").strip()
    documento.save(update_fields=["estado", "observacoes", "atualizado_em"])
    return _json_ok({"id": documento.id, "estado": documento.estado, "tipo": "anulado"})


@login_required
@require_POST
def documento_venda_duplicar(request, id):
    origem = get_object_or_404(DocumentoVenda.objects.select_related("serie", "cliente", "empresa", "template", "centro_custo", "vendedor"), id=id)
    novo = DocumentoVenda.objects.create(
        empresa=origem.empresa, cliente=origem.cliente, vendedor=origem.vendedor, serie=origem.serie, template=origem.template, centro_custo=origem.centro_custo,
        tipo_documento=origem.tipo_documento, estado="rascunho", data_emissao=timezone.localdate(), data_vencimento=origem.data_vencimento,
        observacoes=((origem.observacoes or "") + f"\nDocumento duplicado de {origem.numero_documento}").strip(), criado_por=request.user,
    )
    for linha in origem.linhas.all():
        DocumentoVendaLinha.objects.create(documento=novo, artigo=linha.artigo, descricao=linha.descricao, quantidade=linha.quantidade, preco_unitario=linha.preco_unitario, taxa_iva=linha.taxa_iva)
    novo.recalcular_totais()
    return _json_ok({"id": novo.id, "numero_documento": novo.numero_documento})


@login_required
@require_POST
def documento_venda_converter_fatura(request, id):
    data = _parse_json_request(request)
    origem = get_object_or_404(DocumentoVenda.objects.select_related("serie", "cliente", "empresa", "template", "centro_custo", "vendedor"), id=id)
    serie_destino = get_object_or_404(SerieDocumento, id=data.get("serie_id"))
    novo = DocumentoVenda.objects.create(
        empresa=origem.empresa,
        cliente=origem.cliente,
        vendedor=origem.vendedor,
        serie=serie_destino,
        template=origem.template,
        centro_custo=origem.centro_custo,
        tipo_documento=serie_destino.tipo_documento,
        estado="emitido",
        data_emissao=timezone.localdate(),
        data_vencimento=origem.data_vencimento,
        observacoes=(origem.observacoes or "") + f"\nConvertido do orçamento {origem.numero_documento}",
        criado_por=request.user,
    )
    for linha in origem.linhas.all():
        DocumentoVendaLinha.objects.create(
            documento=novo,
            artigo=linha.artigo,
            descricao=linha.descricao,
            quantidade=linha.quantidade,
            preco_unitario=linha.preco_unitario,
            taxa_iva=linha.taxa_iva,
        )
    novo.recalcular_totais()
    origem.estado = "anulado"
    origem.save(update_fields=["estado", "atualizado_em"])
    return _json_ok({"id": novo.id, "numero_documento": novo.numero_documento})


@login_required
def extrato_clientes(request):
    if not _table_exists("Elion_documentovenda"):
        return render(request, "Elion/extratos.html", {"titulo": "Extrato de Clientes", "linhas": []})
    cliente_id = request.GET.get("cliente_id")
    qs = DocumentoVenda.objects.select_related("cliente").order_by("cliente__nome", "data_emissao", "id")
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    saldo = Decimal("0.00")
    linhas = []
    for item in qs:
        debito = item.total
        credito = item.total - item.saldo_aberto
        saldo += debito - credito
        linhas.append({"entidade": item.cliente.nome, "data": item.data_emissao, "documento": item.numero_documento, "descricao": item.get_tipo_documento_display() if hasattr(item, 'get_tipo_documento_display') else item.tipo_documento, "debito": debito, "credito": credito, "saldo": saldo, "vencimento": item.data_vencimento, "estado": item.get_estado_display()})
    return render(request, "Elion/extratos.html", {"titulo": "Extrato de Clientes", "linhas": linhas, "clientes": _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente).order_by("nome")})


@login_required
def extrato_fornecedores(request):
    saldo = Decimal("0.00")
    linhas = []
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor")).order_by("fornecedor__nome", "data_emissao", "id")
    fornecedor_id = request.GET.get("fornecedor_id")
    if fornecedor_id:
        qs = qs.filter(fornecedor_id=fornecedor_id)
    for item in qs:
        debito = item.valor_fatura
        credito = item.total
        saldo += debito - credito
        linhas.append({"entidade": item.fornecedor.nome if item.fornecedor else "", "data": item.data_emissao, "documento": item.numero_fatura, "descricao": "Lançamento fornecedor", "debito": debito, "credito": credito, "saldo": saldo, "vencimento": item.data_vencimento, "estado": item.status_pagamento})
    return render(request, "Elion/extratos.html", {"titulo": "Extrato de Fornecedores", "linhas": linhas, "fornecedores": _fornecedores_empresa_qs(request).order_by("nome")})


clientes_list = _require_perm("clientes", "view")(clientes_list)
clientes_busca = _require_perm("clientes", "view")(clientes_busca)
clientes_salvar = _require_perm("clientes", "edit")(clientes_salvar)
cliente_excluir = _require_perm("clientes", "delete")(cliente_excluir)
vendedores_list = _require_perm("vendedores", "view")(vendedores_list)
vendedores_salvar = _require_perm("vendedores", "edit")(vendedores_salvar)
artigos_list = _require_perm("artigos", "view")(artigos_list)
artigos_salvar = _require_perm("artigos", "edit")(artigos_salvar)
series_documentos_list = _require_perm("series", "view")(series_documentos_list)
series_documentos_salvar = _require_perm("series", "edit")(series_documentos_salvar)
documentos_venda_list = _require_perm("faturacao", "view")(documentos_venda_list)
documentos_venda_salvar = _require_perm("faturacao", "create")(documentos_venda_salvar)
documento_venda_cancelar = _require_perm("faturacao", "edit")(documento_venda_cancelar)
documento_venda_converter_fatura = _require_perm("faturacao", "create")(documento_venda_converter_fatura)
documento_venda_pdf = _require_perm("faturacao", "export")(documento_venda_pdf)
extrato_clientes = _require_perm("extrato_clientes", "view")(extrato_clientes)
extrato_fornecedores = _require_perm("extrato_fornecedores", "view")(extrato_fornecedores)


def _gerar_codigo_recuperacao():
    return f"{random.randint(100000, 999999)}"


def _buscar_usuario_por_login_email_telefone(valor):
    valor = (valor or "").strip()
    if not valor:
        return None

    user = User.objects.filter(username__iexact=valor).first()
    if user:
        return user

    user = User.objects.filter(email__iexact=valor).first()
    if user:
        return user

    perfil = UsuarioSistema.objects.filter(
        Q(contato__iexact=valor) | Q(email_recuperacao__iexact=valor) | Q(nome__iexact=valor)
    ).select_related("user").first()
    if perfil and getattr(perfil, "user", None):
        return perfil.user
    return None


def _enviar_codigo_email(user, code):
    destino = (user.email or "").strip()
    if not destino:
        perfil = UsuarioSistema.objects.filter(user=user).first()
        destino = ((getattr(perfil, "email_recuperacao", "") or "").strip())
    if not destino:
        return False, "Este utilizador não tem email registado."
    try:
        send_mail(
            "Código de recuperação de senha",
            f"O seu código de recuperação é: {code}. Este código expira em 10 minutos.",
            None,
            [destino],
            fail_silently=False,
        )
        return True, "Código enviado por email com sucesso."
    except Exception:
        return False, "Não foi possível enviar o email. Configure o SMTP do sistema."


def _enviar_codigo_sms(user, code):
    perfil = UsuarioSistema.objects.filter(user=user).first()
    telefone = (getattr(perfil, "contato", "") or "").strip() if perfil else ""
    if not telefone:
        return False, "Este utilizador não tem telefone registado."
    # Integração real de SMS depende de API externa.
    return True, f"Código gerado para envio por telemóvel para {telefone}."


@require_POST
def forgot_password_request(request):
    identificador = (request.POST.get("identificador") or "").strip()
    canal = (request.POST.get("canal") or "email").strip().lower()

    user = _buscar_usuario_por_login_email_telefone(identificador)
    if not user:
        return JsonResponse({"ok": False, "message": "Utilizador não encontrado."})

    PasswordResetCode.objects.filter(user=user, used=False).update(used=True)
    code = _gerar_codigo_recuperacao()
    PasswordResetCode.objects.create(
        user=user,
        code=code,
        channel=canal,
        expires_at=timezone.now() + timedelta(minutes=10),
    )

    if canal == "telefone":
        ok, message = _enviar_codigo_sms(user, code)
    else:
        ok, message = _enviar_codigo_email(user, code)

    if not ok:
        return JsonResponse({"ok": False, "message": message})

    return JsonResponse({"ok": True, "message": message})


@require_POST
def forgot_password_confirm(request):
    identificador = (request.POST.get("identificador") or "").strip()
    code = (request.POST.get("codigo") or "").strip()
    nova_senha = (request.POST.get("nova_senha") or "").strip()
    confirmar_senha = (request.POST.get("confirmar_senha") or "").strip()

    if not nova_senha or not confirmar_senha:
        return JsonResponse({"ok": False, "message": "Informe a nova senha e a confirmação."})
    if nova_senha != confirmar_senha:
        return JsonResponse({"ok": False, "message": "As senhas não coincidem."})
    if len(nova_senha) < 6:
        return JsonResponse({"ok": False, "message": "A nova senha deve ter pelo menos 6 caracteres."})

    user = _buscar_usuario_por_login_email_telefone(identificador)
    if not user:
        return JsonResponse({"ok": False, "message": "Utilizador não encontrado."})

    reset = PasswordResetCode.objects.filter(
        user=user, code=code, used=False, expires_at__gte=timezone.now()
    ).order_by("-created_at").first()

    if not reset:
        return JsonResponse({"ok": False, "message": "Código inválido ou expirado."})

    user.set_password(nova_senha)
    user.save()

    reset.used = True
    reset.save(update_fields=["used"])
    PasswordResetCode.objects.filter(user=user, used=False).update(used=True)

    return JsonResponse({"ok": True, "message": "Senha alterada com sucesso."})




def _documentos_por_tipo_contexto(request, tipo_documento, titulo, subtitulo, rotina_perm):
    empresa = _empresa_ativa(request)
    qs = DocumentoVenda.objects.select_related("cliente", "serie") if _table_exists("Elion_documentovenda") else DocumentoVenda.objects.none()
    if empresa and _table_exists("Elion_documentovenda"):
        qs = qs.filter(Q(empresa=empresa) | Q(empresa__isnull=True))
    if _table_exists("Elion_documentovenda"):
        qs = qs.filter(serie__tipo_documento=tipo_documento)
    linhas = list(qs.order_by("-data_emissao", "-id")[:100]) if _table_exists("Elion_documentovenda") else []
    if linhas:
        _enriquecer_documentos_contexto(linhas)
    series_qs = SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento=tipo_documento).order_by("codigo") if _table_exists("Elion_seriedocumento") else SerieDocumento.objects.none()
    return {
        "linhas": linhas,
        "series": series_qs,
        "series_com_preview": [{"obj": s, "preview": _preview_numero_serie(s)} for s in series_qs],
        "pode_emitir_faturacao": _usuario_tem_permissao(request, rotina_perm, "create"),
        "pode_exportar_faturacao": _usuario_tem_permissao(request, rotina_perm, "export"),
        "pode_configurar_series": _usuario_tem_permissao(request, "series", "edit"),
        "clientes": _filtrar_empresa_generico(request, Cliente.objects.all(), model=Cliente).order_by("nome"),
        "artigos": _filtrar_empresa_generico(request, Artigo.objects.filter(ativo=True), model=Artigo).order_by("nome")[:200],
        "empresa_ativa": empresa,
        "series_orcamento": SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento="OR").order_by("codigo") if _table_exists("Elion_seriedocumento") else [],
        "series_fatura": SerieDocumento.objects.filter(Q(empresa=empresa) | Q(empresa__isnull=True), ativa=True, tipo_documento__in=["FT", "FS", "FR"]).order_by("codigo") if _table_exists("Elion_seriedocumento") else [],
        "modo": "documentos_tipo",
        "titulo_pagina": titulo,
        "subtitulo_pagina": subtitulo,
        "tipo_documento_filtro": tipo_documento,
    }


@login_required
def documentos_tipo_view(request, tipo_documento, titulo, subtitulo, rotina_perm):
    return render(request, "Elion/documentos_venda.html", _documentos_por_tipo_contexto(request, tipo_documento, titulo, subtitulo, rotina_perm))


@login_required
def centro_rotinas_view(request, titulo, subtitulo, grupos, destaque_legal=None):
    return render(request, "Elion/centro_rotinas.html", {
        "titulo": titulo,
        "subtitulo": subtitulo,
        "grupos": grupos,
        "empresa_ativa": _empresa_ativa(request),
        "destaque_legal": destaque_legal or [],
        "permissoes_usuario_json": json.dumps(_permissoes_usuario(request), ensure_ascii=False),
    })


@login_required
def crm_centro_view(request):
    grupos = [
        {
            "titulo": "Cadastros principais",
            "rotinas": [
                {"label": "Cliente", "descricao": "Base comercial para faturação, orçamentos e recibos.", "href": "/painel/?secao=cliente-section", "rotina": "clientes"},
                {"label": "Fornecedor", "descricao": "Cadastro e consulta com NIF e nome.", "href": "/painel/?secao=fornecedor-section", "rotina": "fornecedor"},
                {"label": "Funcionário", "descricao": "Equipa, contactos e apoio operacional.", "href": "/painel/?secao=funcionario-section", "rotina": "funcionario"},
                {"label": "Frota", "descricao": "Viaturas, apólices e inspeções.", "href": "/painel/?secao=frota-section", "rotina": "frota"},
                {"label": "Combustível", "descricao": "Tabela de combustíveis e consumos associados.", "href": "/painel/?secao=combustivel-section", "rotina": "combustivel"},
            ],
        }
    ]
    return centro_rotinas_view(request, "CRM", "Estrutura comercial e operacional alinhada com ERP moderno.", grupos)


@login_required
def fornecedores_centro_view(request):
    grupos = [
        {
            "titulo": "Compras e contas a pagar",
            "rotinas": [
                {"label": "Lançamentos de Faturas", "descricao": "Registo de faturas de fornecedor com vencimentos e estados.", "href": "/painel/?secao=lancamentos-section", "rotina": "lancamentos"},
                {"label": "Consultar pagamentos", "descricao": "Consulta consolidada dos lançamentos e liquidações.", "href": "/painel/?secao=consulta-section", "rotina": "consulta"},
                {"label": "Pagamentos a Fornecedor", "descricao": "Baixas e regularizações de contas a pagar.", "href": "/painel/?secao=baixa-faturas-section", "rotina": "baixa_faturas"},
                {"label": "Notas de Crédito", "descricao": "Rotina preparada para regularizações de fornecedor.", "href": "/relatorios/extrato-fornecedores/", "rotina": "notas_credito_fornecedor"},
                {"label": "Notas de Devolução", "descricao": "Acompanhamento documental de devoluções a fornecedor.", "href": "/relatorios/extrato-fornecedores/", "rotina": "notas_devolucao_fornecedor"},
            ],
        }
    ]
    return centro_rotinas_view(request, "Fornecedores", "Fluxo de compras, pagamentos e regularizações.", grupos)


@login_required
def transporte_centro_view(request):
    grupos = [
        {
            "titulo": "Gestão de manutenção e transporte",
            "rotinas": [
                {"label": "Relatório de Manutenção", "descricao": "Consulta, filtros por período e matrícula e análise dos gastos por viatura.", "href": "/painel/?secao=relatorio-manutencao-section", "rotina": "relatorio_manutencao"},
                {"label": "Documento de Frota", "descricao": "Situação documental da frota com seguro, inspeção e dias para vencimento.", "href": "/painel/?secao=relatorio-documentos-section", "rotina": "relatorio_documentos"},
                {"label": "Fichas de Serviço", "descricao": "Serviços prestados, intervenções e deslocações.", "href": "/transporte/fichas-servico/", "rotina": "fichas_servico"},
            ],
        }
    ]
    return centro_rotinas_view(request, "Transporte", "Gestão operacional da frota, manutenção e situação documental.", grupos)


@login_required
def liquidacoes_centro_view(request):
    grupos = [
        {
            "titulo": "Liquidações e regularizações",
            "rotinas": [
                {"label": "Recibos", "descricao": "Emissão e consulta de recibos.", "href": "/liquidacoes/recibos/", "rotina": "recibos_liquidacao"},
                {"label": "Notas de Liquidação", "descricao": "Rotina preparada para liquidações comerciais.", "href": "/liquidacoes/notas-liquidacao/", "rotina": "notas_liquidacao"},
                {"label": "Notas de Crédito", "descricao": "Regularização de valores faturados.", "href": "/liquidacoes/notas-credito/", "rotina": "notas_credito"},
                {"label": "Devoluções Pagamento", "descricao": "Controlo de devoluções e acertos de pagamento.", "href": "/liquidacoes/devolucoes-pagamento/", "rotina": "devolucoes_pagamento"},
            ],
        }
    ]
    return centro_rotinas_view(request, "Liquidações", "Cobrança, regularização e devoluções.", grupos)


@login_required
def autoridade_tributaria_view(request):
    grupos = [
        {
            "titulo": "Fiscal e ficheiros legais",
            "rotinas": [
                {"label": "Séries e templates", "descricao": "Séries, modelos e preparação documental.", "href": "/configuracoes/series/", "rotina": "series"},
                {"label": "Configuração do IVA de Caixa", "descricao": "Preparação do regime IVA de Caixa.", "href": "/configuracoes/iva/", "rotina": "iva_caixa"},
                {"label": "Ficheiro SAF-T(PT)", "descricao": "Exportação legal SAF-T(PT).", "href": "/configuracoes/saft/", "rotina": "saft"},
                {"label": "Inventário de existências", "descricao": "Inventário legal e comunicação de existências.", "href": "/configuracoes/inventario-existencias/", "rotina": "inventario_existencias"},
                {"label": "Numeração e séries", "descricao": "Orçamento, rascunho e faturação por série.", "href": "/configuracoes/series/", "rotina": "numeracao_series"},
                {"label": "Configuração do IVA", "descricao": "Taxas e enquadramento fiscal base.", "href": "/configuracoes/iva/", "rotina": "configuracao_iva"},
                {"label": "Impostos, taxas e retenções", "descricao": "Taxas, retenções e parametrização fiscal.", "href": "/configuracoes/impostos/", "rotina": "impostos"},
                {"label": "SAF-T, inventário e ficheiros legais", "descricao": "Centro único de conformidade documental.", "href": "/configuracoes/exportacoes/", "rotina": "saft_inventario_ficheiros"},
                {"label": "Resumo de stock e inventário legal", "descricao": "Resumo operacional de existências por artigo e armazém.", "href": "/stocks/centro/", "rotina": "resumo_stock_legal"},
                {"label": "Artigos e stocks", "descricao": "Artigos, categorias, armazéns e preparação para importações.", "href": "/stocks/centro/", "rotina": "artigos_stocks"},
            ],
        }
    ]
    destaque = [
        "Estrutura alinhada com séries/ATCUD, SAF-T(PT), inventário e obrigações fiscais em Portugal.",
        "Menus inspirados em ERP portugueses com separação clara entre vendas, compras, stocks e fiscal.",
    ]
    return centro_rotinas_view(request, "A. Tributária", "Configuração fiscal e legal centralizada.", grupos, destaque)


@login_required
def recibos_venda_view(request):
    return documentos_tipo_view(request, "RC", "Recibos", "Consulta e emissão de recibos comerciais por série.", "recibos_venda")


@login_required
def notas_liquidacao_venda_view(request):
    return documentos_tipo_view(request, "NL", "Notas de Liquidação", "Liquidações comerciais separadas por série e empresa.", "notas_liquidacao_venda")


@login_required
def devolucao_cliente_view(request):
    return documentos_tipo_view(request, "DV", "Notas de Devolução Cliente", "Registo de devoluções de cliente com documento próprio.", "devolucao_cliente")


@login_required
def guias_remessa_view(request):
    return documentos_tipo_view(request, "GR", "Guias de Remessa", "Expedição e entrega de mercadoria com série dedicada.", "guias_remessa")


@login_required
def guias_transporte_view(request):
    return documentos_tipo_view(request, "GT", "Guias de Transporte", "Documentos de transporte e circulação de bens.", "guias_transporte")


@login_required
def fichas_servico_view(request):
    return documentos_tipo_view(request, "FSV", "Fichas de Serviço", "Registo documental de serviços e intervenções.", "fichas_servico")


@login_required
def recibos_liquidacao_view(request):
    return documentos_tipo_view(request, "RC", "Recibos", "Liquidações por recibo com controlo por empresa.", "recibos_liquidacao")


@login_required
def notas_liquidacao_view(request):
    return documentos_tipo_view(request, "NL", "Notas de Liquidação", "Regularização e liquidação documental.", "notas_liquidacao")


@login_required
def notas_credito_view(request):
    return documentos_tipo_view(request, "NC", "Notas de Crédito", "Regularização de faturação e devoluções de valor.", "notas_credito")


@login_required
def devolucoes_pagamento_view(request):
    return documentos_tipo_view(request, "DVP", "Devoluções Pagamento", "Controlo documental de devoluções de pagamento.", "devolucoes_pagamento")


@login_required
def series_configuracao(request):
    empresa = _empresa_ativa(request)
    qs = SerieDocumento.objects.all()
    if empresa:
        qs = qs.filter(Q(empresa=empresa) | Q(empresa__isnull=True))
    series = []
    for s in qs.order_by("tipo_documento", "codigo", "ano"):
        s.preview = _preview_numero_serie(s)
        series.append(s)
    return render(request, "Elion/series_configuracao.html", {"series": series, "empresa_ativa": empresa})


@login_required
def impostos_configuracao(request):
    return render(request, "Elion/impostos_configuracao.html")


@login_required
def stocks_hub(request):
    artigos = _filtrar_empresa_generico(request, Artigo.objects.select_related("categoria", "armazem"), model=Artigo).order_by("nome")[:100] if _table_exists("Elion_artigo") else []
    return render(request, "Elion/stocks_hub.html", {
        "artigos": artigos,
        "total_artigos": _filtrar_empresa_generico(request, Artigo.objects.all(), model=Artigo).count() if _table_exists("Elion_artigo") else 0,
        "total_categorias": CategoriaArtigo.objects.count() if _table_exists("Elion_categoriaartigo") else 0,
        "total_armazens": Armazem.objects.count() if _table_exists("Elion_armazem") else 0,
    })


def _configuracao_fiscal_atual(request):
    empresa = _empresa_ativa(request)
    if empresa:
        cfg, _ = ConfiguracaoFiscal.objects.get_or_create(empresa=empresa)
        return cfg
    return ConfiguracaoFiscal.objects.get_or_create(empresa=None)[0]


def _numero_sequencial_de_documento(doc):
    bruto = (doc.numero_documento or "").strip()
    if not bruto:
        return str(doc.id)
    partes = re.findall(r"(\d+)", bruto)
    return partes[-1] if partes else bruto


def _atcud_documento(doc):
    codigo = (getattr(doc.serie, "codigo_validacao", "") or "").strip()
    seq = _numero_sequencial_de_documento(doc)
    return f"{codigo}-{seq}" if codigo else seq


def _saft_xml_bytes(request):
    empresa = _empresa_ativa(request)
    cfg = _configuracao_fiscal_atual(request)
    docs = DocumentoVenda.objects.select_related("cliente", "serie", "empresa").prefetch_related("linhas").order_by("data_emissao", "id")
    if empresa:
        docs = docs.filter(Q(empresa=empresa) | Q(empresa__isnull=True))
    root = Element("AuditFile")
    header = SubElement(root, "Header")
    SubElement(header, "AuditFileVersion").text = "1.04_01"
    SubElement(header, "CompanyName").text = empresa.nome if empresa else "Empresa"
    SubElement(header, "CompanyID").text = empresa.nif if empresa else "999999990"
    SubElement(header, "TaxRegistrationNumber").text = empresa.nif if empresa else "999999990"
    SubElement(header, "CurrencyCode").text = cfg.moeda or "EUR"
    SubElement(header, "ProductCompanyTaxID").text = empresa.nif if empresa else "999999990"
    SubElement(header, "SoftwareCertificateNumber").text = cfg.software_certificado or "PENDENTE"
    SubElement(header, "DateCreated").text = timezone.localdate().isoformat()
    source_docs = SubElement(root, "SourceDocuments")
    sales = SubElement(source_docs, "SalesInvoices")
    SubElement(sales, "NumberOfEntries").text = str(docs.count())
    total_debit = sum((d.total or Decimal("0.00") for d in docs), Decimal("0.00"))
    SubElement(sales, "TotalDebit").text = f"{total_debit:.2f}"
    SubElement(sales, "TotalCredit").text = "0.00"
    for doc in docs:
        inv = SubElement(sales, "Invoice")
        SubElement(inv, "InvoiceNo").text = doc.numero_documento or f"DOC-{doc.id}"
        SubElement(inv, "ATCUD").text = _atcud_documento(doc)
        SubElement(inv, "InvoiceDate").text = doc.data_emissao.isoformat() if doc.data_emissao else timezone.localdate().isoformat()
        SubElement(inv, "InvoiceType").text = doc.tipo_documento or "FT"
        SubElement(inv, "CustomerID").text = doc.cliente.nif
        totals = SubElement(inv, "DocumentTotals")
        SubElement(totals, "TaxPayable").text = f"{(doc.total_iva or Decimal('0.00')):.2f}"
        net = totals
        SubElement(net, "NetTotal").text = f"{(doc.subtotal or Decimal('0.00')):.2f}"
        SubElement(net, "GrossTotal").text = f"{(doc.total or Decimal('0.00')):.2f}"
        for idx, linha in enumerate(doc.linhas.all(), start=1):
            l = SubElement(inv, "Line")
            SubElement(l, "LineNumber").text = str(idx)
            SubElement(l, "ProductCode").text = linha.artigo.codigo if linha.artigo_id else f"ITEM-{idx}"
            SubElement(l, "ProductDescription").text = linha.descricao
            SubElement(l, "Quantity").text = f"{(linha.quantidade or Decimal('0.00')):.2f}"
            SubElement(l, "UnitPrice").text = f"{(linha.preco_unitario or Decimal('0.00')):.2f}"
            tax = SubElement(l, "Tax")
            SubElement(tax, "TaxType").text = "IVA"
            SubElement(tax, "TaxCountryRegion").text = "PT"
            SubElement(tax, "TaxCode").text = "NOR" if (linha.taxa_iva or 0) else "ISE"
            SubElement(tax, "TaxPercentage").text = f"{(linha.taxa_iva or Decimal('0.00')):.2f}"
            SubElement(l, "CreditAmount").text = f"{(linha.subtotal or Decimal('0.00')):.2f}"
    return tostring(root, encoding="utf-8", xml_declaration=True)


@login_required
def configuracao_iva_view(request):
    cfg = _configuracao_fiscal_atual(request)
    if request.method == "POST":
        data = request.POST
        cfg.regime_iva = data.get("regime_iva") or "Regime normal"
        cfg.iva_caixa = bool(data.get("iva_caixa"))
        cfg.taxa_normal = _to_decimal(data.get("taxa_normal"), "23.00")
        cfg.taxa_intermedia = _to_decimal(data.get("taxa_intermedia"), "13.00")
        cfg.taxa_reduzida = _to_decimal(data.get("taxa_reduzida"), "6.00")
        cfg.retencao_padrao = _to_decimal(data.get("retencao_padrao"), "0.00")
        cfg.motivo_isencao = (data.get("motivo_isencao") or "M99").strip()[:60]
        cfg.moeda = (data.get("moeda") or "EUR").strip()[:10]
        cfg.software_certificado = (data.get("software_certificado") or "").strip()[:120]
        cfg.exportar_qr_atcud = bool(data.get("exportar_qr_atcud"))
        cfg.observacoes_legais = data.get("observacoes_legais") or ""
        cfg.save()
        messages.success(request, "Configuração do IVA guardada com sucesso.")
        return redirect("configuracao_iva")
    return render(request, "Elion/configuracao_iva.html", {"cfg": cfg})


@login_required
def exportacoes_view(request):
    return render(request, "Elion/exportacoes.html")


@login_required
def saft_view(request):
    cfg = _configuracao_fiscal_atual(request)
    return render(request, "Elion/saft_pt.html", {"cfg": cfg})


@login_required
def saft_exportar_xml(request):
    response = HttpResponse(_saft_xml_bytes(request), content_type="application/xml")
    response["Content-Disposition"] = 'attachment; filename="saft_pt.xml"'
    return response


@login_required
def inventario_existencias_view(request):
    artigos = _filtrar_empresa_generico(request, Artigo.objects.filter(ativo=True), model=Artigo).order_by("nome") if _table_exists("Elion_artigo") else []
    return render(request, "Elion/inventario_existencias.html", {"artigos": artigos})


@login_required
def inventario_exportar_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="inventario_existencias.csv"'
    writer = csv.writer(response, delimiter=';')
    writer.writerow(["Codigo", "Descricao", "Unidade", "Stock", "PrecoVenda", "IVA"])
    for item in _filtrar_empresa_generico(request, Artigo.objects.filter(ativo=True), model=Artigo).order_by("nome"):
        writer.writerow([item.codigo, item.nome, getattr(item, "unidade_medida", "UN"), f"{item.stock_atual:.2f}", f"{item.preco_venda:.2f}", f"{getattr(item, 'taxa_iva_padrao', Decimal('23.00')):.2f}"])
    return response

series_configuracao = _require_perm("series", "view")(series_configuracao)
impostos_configuracao = _require_perm("impostos", "view")(impostos_configuracao)
stocks_hub = _require_perm("artigos", "view")(stocks_hub)
configuracoes_sistema_view = _require_perm("configuracoes_sistema", "view")(configuracoes_sistema_view)
configuracao_iva_view = _require_perm("configuracao_iva", "view")(configuracao_iva_view)
exportacoes_view = _require_perm("exportacoes", "view")(exportacoes_view)
saft_view = _require_perm("saft", "view")(saft_view)
inventario_existencias_view = _require_perm("inventario_existencias", "view")(inventario_existencias_view)

orcamentos_list = _require_perm("orcamentos", "view")(orcamentos_list)
faturacao_list = _require_perm("faturacao", "view")(faturacao_list)
crm_centro_view = _require_perm("clientes", "view")(crm_centro_view)
fornecedores_centro_view = _require_perm("lancamentos", "view")(fornecedores_centro_view)
transporte_centro_view = _require_perm("relatorio_manutencao", "view")(transporte_centro_view)
liquidacoes_centro_view = _require_perm("recibos_liquidacao", "view")(liquidacoes_centro_view)
autoridade_tributaria_view = _require_perm("series", "view")(autoridade_tributaria_view)
recibos_venda_view = _require_perm("recibos_venda", "view")(recibos_venda_view)
notas_liquidacao_venda_view = _require_perm("notas_liquidacao_venda", "view")(notas_liquidacao_venda_view)
devolucao_cliente_view = _require_perm("devolucao_cliente", "view")(devolucao_cliente_view)
guias_remessa_view = _require_perm("guias_remessa", "view")(guias_remessa_view)
guias_transporte_view = _require_perm("guias_transporte", "view")(guias_transporte_view)
fichas_servico_view = _require_perm("fichas_servico", "view")(fichas_servico_view)
recibos_liquidacao_view = _require_perm("recibos_liquidacao", "view")(recibos_liquidacao_view)
notas_liquidacao_view = _require_perm("notas_liquidacao", "view")(notas_liquidacao_view)
notas_credito_view = _require_perm("notas_credito", "view")(notas_credito_view)
devolucoes_pagamento_view = _require_perm("devolucoes_pagamento", "view")(devolucoes_pagamento_view)


@login_required
def lancamentos_editar(request, id):
    lanc = get_object_or_404(Lancamento, id=id)
    data = model_to_dict(lanc)
    return JsonResponse(data)
