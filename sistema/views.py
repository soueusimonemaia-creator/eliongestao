import csv
import json
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction, connection
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from .models import (
    Empresa,
    Fornecedor,
    Funcionario,
    Frota,
    Combustivel,
    Lancamento,
    UsuarioSistema,
    BaixaFatura,
)

try:
    from .models import ItemCombustivel
except Exception:
    try:
        from .models import LancamentoCombustivel as ItemCombustivel
    except Exception:
        ItemCombustivel = None

try:
    from .models import ManutencaoFrota
except Exception:
    ManutencaoFrota = None


def _is_ajax(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _to_decimal(value, default="0.00"):
    if value in [None, ""]:
        value = default
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _to_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    return parse_date(value)


def _json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return {}


def _normalizar_status_fatura(status):
    status = (status or "").strip().lower()
    mapa = {
        "aberta": "Em aberto",
        "em aberto": "Em aberto",
        "em_aberto": "Em aberto",
        "parcial": "Parcial",
        "paga": "Paga",
        "pago": "Paga",
        "fechada": "Paga",
        "fechado": "Paga",
        "quitada": "Paga",
        "quitado": "Paga",
        "a vencer": "Em aberto",
        "vencida": "Em aberto",
    }
    return mapa.get(status, "")


def _empresa_ativa(request):
    empresas = list(_empresas_usuario(request))
    if not empresas:
        return None

    empresa_id = request.session.get("empresa_ativa_id")
    if empresa_id:
        for empresa in empresas:
            if str(empresa.id) == str(empresa_id):
                return empresa
    return empresas[0]


def _empresas_usuario(request):
    try:
        perfil = UsuarioSistema.objects.filter(user=request.user).first()
        if perfil:
            if getattr(perfil, "administrador_geral", False):
                return Empresa.objects.all().order_by("nome")
            if hasattr(perfil, "empresas"):
                return perfil.empresas.all().order_by("nome")
            if hasattr(perfil, "empresa") and getattr(perfil, "empresa_id", None):
                return Empresa.objects.filter(id=perfil.empresa_id).order_by("nome")
    except Exception:
        pass
    return Empresa.objects.all().order_by("nome")


def _filtrar_lancamentos_empresa(request, qs):
    empresa = _empresa_ativa(request)
    if empresa and hasattr(Lancamento, "empresa_id"):
        return qs.filter(empresa=empresa)
    return qs


def _fornecedor_nome(obj):
    fornecedor = getattr(obj, "fornecedor", None)
    return getattr(fornecedor, "nome", "") if fornecedor else ""


def _fornecedor_nif(obj):
    fornecedor = getattr(obj, "fornecedor", None)
    return getattr(fornecedor, "nif", "") if fornecedor else ""


def _total_pago_lancamento(obj):
    total_pago = getattr(obj, "total_pago", None)
    if total_pago not in [None, ""]:
        return _to_decimal(total_pago)

    total_field = getattr(obj, "total", None)
    if total_field not in [None, ""]:
        return _to_decimal(total_field)

    dinheiro = _to_decimal(getattr(obj, "dinheiro", 0))
    cartao = _to_decimal(getattr(obj, "cartao", 0))
    transferencia = _to_decimal(getattr(obj, "transferencia", 0))
    mbway = _to_decimal(getattr(obj, "mbway", 0))
    nota_credito = _to_decimal(getattr(obj, "nota_credito", 0))
    return dinheiro + cartao + transferencia + mbway + nota_credito


def _seguradora_frota(obj):
    if hasattr(obj, "seguradora"):
        return getattr(obj, "seguradora", "") or ""
    if hasattr(obj, "asseguradora"):
        return getattr(obj, "asseguradora", "") or ""
    return ""


def _recalcular_status_lancamento(lancamento):
    dinheiro = _to_decimal(getattr(lancamento, "dinheiro", 0))
    cartao = _to_decimal(getattr(lancamento, "cartao", 0))
    transferencia = _to_decimal(getattr(lancamento, "transferencia", 0))
    mbway = _to_decimal(getattr(lancamento, "mbway", 0))
    nota_credito = _to_decimal(getattr(lancamento, "nota_credito", 0))
    valor_fatura = _to_decimal(getattr(lancamento, "valor_fatura", 0))

    total_pago = dinheiro + cartao + transferencia + mbway + nota_credito
    saldo_aberto = valor_fatura - total_pago
    if saldo_aberto < 0:
        saldo_aberto = Decimal("0.00")

    lancamento.total_pago = total_pago
    if hasattr(lancamento, "total"):
        lancamento.total = total_pago
    lancamento.saldo_aberto = saldo_aberto

    if valor_fatura <= 0 and total_pago > 0:
        lancamento.status_pagamento = "Paga"
    elif valor_fatura > 0 and total_pago >= valor_fatura:
        lancamento.status_pagamento = "Paga"
    elif total_pago > 0:
        lancamento.status_pagamento = "Parcial"
    else:
        lancamento.status_pagamento = "Em aberto"

    return lancamento


def _linhas_financeiras(qs):
    linhas = []
    for obj in qs:
        valor_fatura = _to_decimal(getattr(obj, "valor_fatura", 0))
        total_pago = _total_pago_lancamento(obj)
        saldo_aberto = _to_decimal(getattr(obj, "saldo_aberto", 0))
        if valor_fatura > 0 and saldo_aberto == 0 and total_pago < valor_fatura:
            saldo_aberto = valor_fatura - total_pago
        linhas.append({
            "id": obj.id,
            "data_emissao": obj.data_emissao.strftime("%Y-%m-%d") if getattr(obj, "data_emissao", None) else "",
            "data_vencimento": obj.data_vencimento.strftime("%Y-%m-%d") if getattr(obj, "data_vencimento", None) else "",
            "data_pagamento": obj.data_pagamento.strftime("%Y-%m-%d") if getattr(obj, "data_pagamento", None) else "",
            "numero_fatura": getattr(obj, "numero_fatura", "") or "",
            "fornecedor": _fornecedor_nome(obj),
            "nif": _fornecedor_nif(obj),
            "valor_fatura": float(valor_fatura),
            "total_pago": float(total_pago),
            "saldo_aberto": float(max(Decimal("0.00"), saldo_aberto)),
            "status_pagamento": getattr(obj, "status_pagamento", "") or "",
        })
    return linhas


def _totais_financeiros(qs):
    total_valor = Decimal("0.00")
    total_pago = Decimal("0.00")
    total_saldo = Decimal("0.00")

    for obj in qs:
        valor_fatura = _to_decimal(getattr(obj, "valor_fatura", 0))
        pago = _total_pago_lancamento(obj)
        saldo = _to_decimal(getattr(obj, "saldo_aberto", 0))
        if valor_fatura > 0 and saldo == 0 and pago < valor_fatura:
            saldo = valor_fatura - pago
        total_valor += valor_fatura
        total_pago += pago
        total_saldo += max(Decimal("0.00"), saldo)

    return {
        "valor_fatura": float(total_valor),
        "total_pago": float(total_pago),
        "saldo_aberto": float(total_saldo),
    }


def _faturas_alerta(request):
    hoje = date.today()
    limite = hoje + timedelta(days=20)

    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    ).filter(
        data_vencimento__isnull=False
    ).exclude(
        status_pagamento__iexact="Paga"
    ).filter(
        data_vencimento__range=[hoje, limite]
    ).order_by("data_vencimento", "id")

    dados = []
    for item in qs:
        dias = (item.data_vencimento - hoje).days if item.data_vencimento else 0
        dados.append({
            "id": item.id,
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": _fornecedor_nome(item),
            "data": item.data_vencimento.strftime("%d/%m/%Y") if item.data_vencimento else "",
            "status": item.status_pagamento or "Em aberto",
            "dias": dias,
            "valor_fatura": float(_to_decimal(getattr(item, "valor_fatura", 0))),
            "total_pago": float(_total_pago_lancamento(item)),
            "saldo_aberto": float(_to_decimal(getattr(item, "saldo_aberto", 0))),
        })
    return dados


def _faturas_vencidas(request):
    hoje = date.today()

    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    ).filter(
        data_vencimento__isnull=False,
        data_vencimento__lt=hoje
    ).exclude(
        status_pagamento__iexact="Paga"
    ).order_by("data_vencimento", "id")

    dados = []
    for item in qs:
        dias_atraso = (hoje - item.data_vencimento).days if item.data_vencimento else 0
        dados.append({
            "id": item.id,
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": _fornecedor_nome(item),
            "data": item.data_vencimento.strftime("%d/%m/%Y") if item.data_vencimento else "",
            "status": item.status_pagamento or "Em aberto",
            "dias_atraso": dias_atraso,
            "valor_fatura": float(_to_decimal(getattr(item, "valor_fatura", 0))),
            "total_pago": float(_total_pago_lancamento(item)),
            "saldo_aberto": float(_to_decimal(getattr(item, "saldo_aberto", 0))),
        })
    return dados


def _frota_alerta(request):
    hoje = date.today()
    limite = hoje + timedelta(days=30)

    qs = Frota.objects.all().order_by("matricula")
    dados = []

    for item in qs:
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)

        dias_seguro = (seguro - hoje).days if seguro else None
        dias_inspecao = (inspecao - hoje).days if inspecao else None

        status_seguro = "OK"
        if seguro:
            if seguro < hoje:
                status_seguro = "Vencido"
            elif seguro <= limite:
                status_seguro = "A vencer"

        status_inspecao = "OK"
        if inspecao:
            if inspecao < hoje:
                status_inspecao = "Vencido"
            elif inspecao <= limite:
                status_inspecao = "A vencer"

        status_geral = "OK"
        if status_seguro == "Vencido" or status_inspecao == "Vencido":
            status_geral = "Vencido"
        elif status_seguro == "A vencer" or status_inspecao == "A vencer":
            status_geral = "A vencer"

        dados.append({
            "id": item.id,
            "matricula": getattr(item, "matricula", "") or "",
            "seguradora": _seguradora_frota(item),
            "seguro": seguro.strftime("%d/%m/%Y") if seguro else "",
            "inspecao": inspecao.strftime("%d/%m/%Y") if inspecao else "",
            "dias_seguro": dias_seguro,
            "dias_inspecao": dias_inspecao,
            "status_seguro": status_seguro,
            "status_inspecao": status_inspecao,
            "status": status_geral,
            "css_seguro": "danger" if status_seguro in ["Vencido", "A vencer"] else "success",
            "css_inspecao": "danger" if status_inspecao in ["Vencido", "A vencer"] else "success",
            "css_status": "danger" if status_geral in ["Vencido", "A vencer"] else "success",
        })

    return dados


def _render_dashboard(request, secao_ativa="dashboard"):
    context = {
        "secao_ativa": secao_ativa.replace("-section", ""),
        "empresa_ativa": _empresa_ativa(request),
        "empresas_usuario": _empresas_usuario(request),
        "alertas_faturas": _faturas_alerta(request),
        "faturas_vencidas": _faturas_vencidas(request),
        "alertas_frota": _frota_alerta(request),
        "combustiveis_menu": json.dumps(
            list(Combustivel.objects.values("id", "nome").order_by("nome"))
        ),
        "empresas_json": json.dumps(
            list(_empresas_usuario(request).values("id", "nome"))
        ),
        "frota_json": json.dumps(
            list(Frota.objects.values("id", "matricula").order_by("matricula"))
        ),
    }
    return render(request, "Elion/dashboard.html", context)


def login_view(request):
    if request.user.is_authenticated:
        return redirect("/painel/")

    if request.method == "POST":
        username = (
            request.POST.get("username")
            or request.POST.get("nome de usuário")
            or request.POST.get("usuario")
        )
        password = request.POST.get("password") or request.POST.get("senha")
        empresa_id = request.POST.get("empresa")

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            if empresa_id:
                request.session["empresa_ativa_id"] = empresa_id
            return redirect(request.GET.get("next") or "/painel/")
        messages.error(request, "Utilizador ou senha inválidos.")

    return render(request, "Elion/login.html", {
        "empresas": Empresa.objects.all().order_by("nome")
    })


@login_required
def logout_view(request):
    logout(request)
    return redirect("/")


@login_required
def dashboard_view(request):
    return _render_dashboard(request, "dashboard")


@login_required
@require_POST
def trocar_empresa_ativa(request):
    data = _json_body(request)
    empresa_id = data.get("empresa_id")
    if empresa_id:
        request.session["empresa_ativa_id"] = str(empresa_id)
    return JsonResponse({"ok": True})


@login_required
@require_GET
def fornecedores_list(request):
    linhas = list(Fornecedor.objects.values("id", "nif", "nome").order_by("nome"))
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def fornecedores_salvar(request):
    data = _json_body(request)
    obj = get_object_or_404(Fornecedor, id=data["id"]) if data.get("id") else Fornecedor()

    obj.nif = data.get("nif", "")
    obj.nome = data.get("nome", "")
    obj.save()

    return JsonResponse({
        "ok": True,
        "id": obj.id,
        "nome": obj.nome,
        "nif": obj.nif,
    })


@login_required
@require_GET
def fornecedores_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = Fornecedor.objects.all()

    if q:
        filtro = Q(nome__icontains=q) | Q(nif__icontains=q)
        if q.isdigit():
            filtro |= Q(id=int(q))
        qs = qs.filter(filtro)

    linhas = list(qs.order_by("nome").values("id", "nome", "nif")[:20])
    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def funcionarios_list(request):
    linhas = list(
        Funcionario.objects.values("id", "nome", "contato", "email").order_by("nome")
    )
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def funcionarios_salvar(request):
    data = _json_body(request)
    obj = get_object_or_404(Funcionario, id=data["id"]) if data.get("id") else Funcionario()

    obj.nome = data.get("nome", "")
    obj.contato = data.get("contato", "")
    obj.email = data.get("email", "")
    obj.save()

    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def frota_list(request):
    linhas = []
    for item in Frota.objects.all().order_by("matricula"):
        linhas.append({
            "id": item.id,
            "matricula": getattr(item, "matricula", "") or "",
            "seguro": item.seguro.strftime("%Y-%m-%d") if getattr(item, "seguro", None) else "",
            "seguradora": _seguradora_frota(item),
            "inspecao": item.inspecao.strftime("%Y-%m-%d") if getattr(item, "inspecao", None) else "",
        })
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def frota_salvar(request):
    data = _json_body(request)
    obj = get_object_or_404(Frota, id=data["id"]) if data.get("id") else Frota()

    obj.matricula = data.get("matricula", "")
    obj.seguro = _to_date(data.get("seguro"))
    obj.inspecao = _to_date(data.get("inspecao"))

    if hasattr(obj, "seguradora"):
        obj.seguradora = data.get("seguradora", "")
    elif hasattr(obj, "asseguradora"):
        obj.asseguradora = data.get("seguradora", "")

    obj.save()
    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def frota_busca(request):
    q = (request.GET.get("q") or "").strip()
    qs = Frota.objects.all()

    if q:
        qs = qs.filter(matricula__icontains=q)

    linhas = list(qs.order_by("matricula").values("id", "matricula")[:20])
    return JsonResponse({"linhas": linhas})


@login_required
@require_GET
def combustiveis_list(request):
    linhas = list(Combustivel.objects.values("id", "nome").order_by("nome"))
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def combustiveis_salvar(request):
    data = _json_body(request)
    obj = get_object_or_404(Combustivel, id=data["id"]) if data.get("id") else Combustivel()

    obj.nome = data.get("nome", "")
    obj.save()
    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def empresas_list(request):
    linhas = list(
        Empresa.objects.values("id", "nif", "nome", "cidade", "contato").order_by("nome")
    )
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def empresas_salvar(request):
    obj = get_object_or_404(Empresa, id=request.POST.get("id")) if request.POST.get("id") else Empresa()

    obj.nif = request.POST.get("nif", "")
    obj.nome = request.POST.get("nome", "")
    obj.morada = request.POST.get("morada", "")
    obj.caixa_postal = request.POST.get("caixa_postal", "")
    obj.cidade = request.POST.get("cidade", "")
    obj.contato = request.POST.get("contato", "")
    obj.email = request.POST.get("email", "")

    if request.POST.get("remover_logo") == "1":
        obj.logo = None

    if request.FILES.get("logo"):
        obj.logo = request.FILES["logo"]

    obj.save()
    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def usuarios_list(request):
    linhas = []
    for item in UsuarioSistema.objects.select_related("user").all().order_by("nome"):
        linhas.append({
            "id": item.id,
            "nome": getattr(item, "nome", ""),
            "contato": getattr(item, "contato", ""),
            "email_recuperacao": getattr(item, "email_recuperacao", ""),
            "username": item.user.username if getattr(item, "user", None) else "",
            "empresa": item.empresas.first().id if hasattr(item, "empresas") and item.empresas.exists() else "",
            "empresa_nome": item.empresas.first().nome if hasattr(item, "empresas") and item.empresas.exists() else "",
            "administrador_geral": bool(getattr(item, "administrador_geral", False)),
        })
    return JsonResponse({"linhas": linhas})


@login_required
@require_POST
def usuarios_salvar(request):
    data = _json_body(request)

    if data.get("id"):
        perfil = get_object_or_404(UsuarioSistema, id=data["id"])
        user = perfil.user
    else:
        user = User()

    user.username = data.get("username", user.username if user.pk else "")
    if data.get("password"):
        user.set_password(data["password"])
    elif not user.pk:
        user.set_password("123456")
    user.save()

    if data.get("id"):
        perfil = get_object_or_404(UsuarioSistema, id=data["id"])
    else:
        perfil, _ = UsuarioSistema.objects.get_or_create(user=user)

    perfil.nome = data.get("nome", "")
    perfil.contato = data.get("contato", "")
    perfil.email_recuperacao = data.get("email_recuperacao", "")

    if hasattr(perfil, "empresa_id"):
        perfil.empresa_id = data.get("empresa") or None

    perfil.administrador_geral = bool(data.get("administrador_geral"))
    user.is_staff = perfil.administrador_geral or user.is_superuser
    user.save(update_fields=["username", "password", "is_staff"]) if user.pk else user.save()

    perfil.save()

    if hasattr(perfil, "empresas"):
        if perfil.administrador_geral:
            perfil.empresas.set(Empresa.objects.all())
        else:
            empresa_id = data.get("empresa")
            if empresa_id:
                perfil.empresas.set(Empresa.objects.filter(id=empresa_id))
            else:
                perfil.empresas.clear()

    return JsonResponse({"ok": True, "id": perfil.id})


@login_required
@require_POST
def fornecedor_excluir(request, id):
    get_object_or_404(Fornecedor, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def funcionario_excluir(request, id):
    get_object_or_404(Funcionario, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def frota_excluir(request, id):
    get_object_or_404(Frota, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def combustivel_excluir(request, id):
    get_object_or_404(Combustivel, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def empresa_excluir(request, id):
    get_object_or_404(Empresa, id=id).delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def usuario_excluir(request, id):
    perfil = get_object_or_404(UsuarioSistema, id=id)
    user = getattr(perfil, "user", None)
    if user and (user.username or "").strip().casefold() == "maia":
        return JsonResponse({"error": "O utilizador Maia é protegido e não pode ser excluído."}, status=403)
    perfil.delete()
    if user:
        user.delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def usuario_alterar_senha(request, id=None):
    data = _json_body(request)
    alvo = request.user

    if id:
        perfil = get_object_or_404(UsuarioSistema, id=id)
        alvo = perfil.user
        if request.user != alvo and not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({"error": "Sem permissão para alterar a senha deste utilizador."}, status=403)

    senha = (data.get("password") or "").strip()
    if not senha:
        return JsonResponse({"error": "Informe a nova senha."}, status=400)

    alvo.set_password(senha)
    alvo.save(update_fields=["password"])
    return JsonResponse({"ok": True})


@login_required
@require_POST
@transaction.atomic
def lancamentos_salvar(request):
    data = _json_body(request)

    if data.get("id"):
        obj = get_object_or_404(Lancamento, id=data["id"])
    else:
        obj = Lancamento()

    empresa = _empresa_ativa(request)
    if hasattr(obj, "empresa") and empresa:
        obj.empresa = empresa

    fornecedor_id = data.get("fornecedor_id")
    obj.fornecedor = Fornecedor.objects.filter(id=fornecedor_id).first() if fornecedor_id else None

    obj.data_emissao = _to_date(data.get("data_emissao")) or date.today()
    obj.numero_fatura = data.get("numero_fatura", "")
    obj.dinheiro = _to_decimal(data.get("dinheiro"))
    obj.cartao = _to_decimal(data.get("cartao"))
    obj.valor_fatura = _to_decimal(data.get("valor_fatura"))
    obj.data_vencimento = _to_date(data.get("data_vencimento")) or obj.data_emissao

    if hasattr(obj, "parcelas"):
        try:
            obj.parcelas = int(data.get("parcelas") or 1)
        except Exception:
            obj.parcelas = 1

    if hasattr(obj, "valor_parcela"):
        obj.valor_parcela = _to_decimal(data.get("valor_parcela"))

    if hasattr(obj, "transferencia") and obj.pk is None:
        obj.transferencia = _to_decimal("0")

    if hasattr(obj, "mbway") and obj.pk is None:
        obj.mbway = _to_decimal("0")

    if hasattr(obj, "nota_credito") and obj.pk is None:
        obj.nota_credito = _to_decimal("0")

    _recalcular_status_lancamento(obj)
    obj.save()

    if ItemCombustivel is not None:
        ItemCombustivel.objects.filter(lancamento=obj).delete()
        for item in data.get("itens_combustivel", []):
            ItemCombustivel.objects.create(
                lancamento=obj,
                combustivel_id=item.get("combustivel_id") or None,
                frota_id=item.get("frota_id") or None,
                km_inicio=_to_decimal(item.get("km_inicio")),
                km_final=_to_decimal(item.get("km_final")),
                km_total=_to_decimal(item.get("km_total")),
                litro=_to_decimal(item.get("litro")),
                valor_litro=_to_decimal(item.get("valor_litro")),
                valor_total=_to_decimal(item.get("valor_total")),
            )

    if ManutencaoFrota is not None:
        ManutencaoFrota.objects.filter(lancamento=obj).delete()
        for item in data.get("itens_manutencao", []):
            kwargs = {
                "lancamento": obj,
                "frota_id": item.get("frota_id") or None,
                "km_inicio": _to_decimal(item.get("km_inicio")),
                "km_final": _to_decimal(item.get("km_final")),
                "km_total": _to_decimal(item.get("km_total")),
                "valor": _to_decimal(item.get("valor")),
                "observacao": item.get("observacao", ""),
            }
            ManutencaoFrota.objects.create(**kwargs)

    return JsonResponse({"ok": True, "id": obj.id})


@login_required
@require_GET
def lancamento_detalhe(request, id):
    obj = get_object_or_404(Lancamento, id=id)

    itens_combustivel = []
    if ItemCombustivel is not None:
        for item in ItemCombustivel.objects.filter(lancamento=obj).select_related("combustivel", "frota"):
            itens_combustivel.append({
                "combustivel_id": item.combustivel_id,
                "combustivel_nome": getattr(item.combustivel, "nome", ""),
                "frota_id": item.frota_id,
                "km_inicio": float(_to_decimal(item.km_inicio)),
                "km_final": float(_to_decimal(item.km_final)),
                "km_total": float(_to_decimal(item.km_total)),
                "litro": float(_to_decimal(item.litro)),
                "valor_litro": float(_to_decimal(item.valor_litro)),
                "valor_total": float(_to_decimal(item.valor_total)),
            })

    itens_manutencao = []
    if ManutencaoFrota is not None:
        for item in ManutencaoFrota.objects.filter(lancamento=obj).select_related("frota"):
            itens_manutencao.append({
                "frota_id": item.frota_id,
                "matricula": getattr(item.frota, "matricula", "") if getattr(item, "frota", None) else "",
                "km_inicio": float(_to_decimal(item.km_inicio)),
                "km_final": float(_to_decimal(item.km_final)),
                "km_total": float(_to_decimal(item.km_total)),
                "valor": float(_to_decimal(item.valor)),
                "observacao": getattr(item, "observacao", ""),
            })

    return JsonResponse({
        "id": obj.id,
        "data_emissao": obj.data_emissao.strftime("%Y-%m-%d") if obj.data_emissao else "",
        "numero_fatura": obj.numero_fatura or "",
        "fornecedor_id": obj.fornecedor_id,
        "fornecedor": _fornecedor_nome(obj),
        "dinheiro": float(_to_decimal(getattr(obj, "dinheiro", 0))),
        "cartao": float(_to_decimal(getattr(obj, "cartao", 0))),
        "valor_fatura": float(_to_decimal(getattr(obj, "valor_fatura", 0))),
        "parcelas": getattr(obj, "parcelas", 1),
        "valor_parcela": float(_to_decimal(getattr(obj, "valor_parcela", 0))),
        "data_vencimento": obj.data_vencimento.strftime("%Y-%m-%d") if obj.data_vencimento else "",
        "status_pagamento": obj.status_pagamento or "",
        "total_pago": float(_total_pago_lancamento(obj)),
        "saldo_aberto": float(_to_decimal(getattr(obj, "saldo_aberto", 0))),
        "itens_combustivel": itens_combustivel,
        "itens_manutencao": itens_manutencao,
    })


@login_required
@require_POST
def lancamento_excluir(request, id):
    obj = get_object_or_404(Lancamento, id=id)
    obj.delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
@transaction.atomic
def lancamento_baixa(request, id):
    obj = get_object_or_404(Lancamento, id=id)
    data = _json_body(request)

    obj.dinheiro = _to_decimal(getattr(obj, "dinheiro", 0)) + _to_decimal(data.get("dinheiro"))
    obj.cartao = _to_decimal(getattr(obj, "cartao", 0)) + _to_decimal(data.get("cartao"))

    if hasattr(obj, "mbway"):
        obj.mbway = _to_decimal(getattr(obj, "mbway", 0)) + _to_decimal(data.get("mbway"))

    if hasattr(obj, "transferencia"):
        obj.transferencia = _to_decimal(getattr(obj, "transferencia", 0)) + _to_decimal(data.get("transferencia"))

    if hasattr(obj, "nota_credito"):
        obj.nota_credito = _to_decimal(getattr(obj, "nota_credito", 0)) + _to_decimal(data.get("nota_credito"))

    proxima_data = _to_date(data.get("proxima_data_vencimento"))
    if proxima_data and hasattr(obj, "data_vencimento"):
        obj.data_vencimento = proxima_data

    if hasattr(obj, "observacao_baixa"):
        obj.observacao_baixa = data.get("observacao", "")

    _recalcular_status_lancamento(obj)
    obj.save()

    total_baixa = _to_decimal(data.get("total_baixa"))
    if total_baixa > Decimal("0.00") and _baixa_table_available():
        try:
            BaixaFatura.objects.create(
                empresa=getattr(obj, "empresa", None),
                lancamento=obj,
                usuario=request.user,
                data_baixa=timezone.localdate(),
                fornecedor_snapshot=_fornecedor_nome(obj),
                numero_fatura_snapshot=obj.numero_fatura or "",
                dinheiro=_to_decimal(data.get("dinheiro")),
                cartao=_to_decimal(data.get("cartao")),
                transferencia=_to_decimal(data.get("transferencia")),
                mbway=_to_decimal(data.get("mbway")),
                nota_credito=_to_decimal(data.get("nota_credito")),
                total_baixa=total_baixa,
                saldo_resultante=_to_decimal(obj.saldo_aberto),
                observacao=data.get("observacao", ""),
            )
        except (OperationalError, ProgrammingError):
            pass

    return JsonResponse({
        "ok": True,
        "status_pagamento": obj.status_pagamento,
        "total_pago": float(_total_pago_lancamento(obj)),
        "saldo_aberto": float(_to_decimal(obj.saldo_aberto)),
    })


@login_required
@require_GET
def consulta_lancamentos(request):
    qs = _filtrar_lancamentos_empresa(
        request,
        Lancamento.objects.select_related("fornecedor").all()
    )

    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()

    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)

    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)

    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)

    qs = qs.order_by("-data_emissao", "-id")

    return JsonResponse({"linhas": _linhas_financeiras(qs)})


def _filtrar_lancamentos_financeiros(request, qs=None):
    qs = _filtrar_lancamentos_empresa(
        request,
        qs or Lancamento.objects.select_related("fornecedor", "empresa").all()
    )

    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()
    status = _normalizar_status_fatura(request.GET.get("status"))
    periodo_tipo = (request.GET.get("periodo_tipo") or "vencimento").strip().lower()

    campo_data = "data_vencimento"
    if periodo_tipo == "emissao":
        campo_data = "data_emissao"
    elif periodo_tipo == "pagamento":
        campo_data = "data_pagamento"

    if data_inicio and data_fim and data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    if data_inicio:
        qs = qs.filter(**{f"{campo_data}__gte": data_inicio})
    if data_fim:
        qs = qs.filter(**{f"{campo_data}__lte": data_fim})

    if fornecedor:
        filtro = Q(fornecedor__nome__icontains=fornecedor) | Q(fornecedor__nif__icontains=fornecedor)
        if fornecedor.isdigit():
            filtro |= Q(fornecedor__id=int(fornecedor))
        qs = qs.filter(filtro)

    if fatura:
        qs = qs.filter(numero_fatura__icontains=fatura)

    if status:
        qs = qs.filter(status_pagamento__iexact=status)

    return qs.order_by("data_vencimento", "data_emissao", "id"), {
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "fornecedor": fornecedor,
        "fatura": fatura,
        "status": status,
        "periodo_tipo": periodo_tipo,
    }


@login_required
def relatorio_baixa_faturas(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "baixa-faturas")

    qs, filtros = _filtrar_lancamentos_financeiros(request)
    qs = qs.filter(
        Q(numero_fatura__isnull=False) & ~Q(numero_fatura="") | Q(valor_fatura__gt=0) | Q(data_vencimento__isnull=False)
    ).distinct()

    return JsonResponse({
        "cabecalho": "BAIXA DE FATURAS",
        "linhas": _linhas_financeiras(qs),
        "totais": _totais_financeiros(qs),
        "filtros": filtros,
    })


@login_required
def relatorio_financeiro(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-financeiro")

    qs, filtros = _filtrar_lancamentos_financeiros(request)
    return JsonResponse({
        "cabecalho": "RELATÓRIO FINANCEIRO",
        "linhas": _linhas_financeiras(qs),
        "totais": _totais_financeiros(qs),
        "filtros": filtros,
    })


def _baixa_table_available():
    try:
        return BaixaFatura._meta.db_table in connection.introspection.table_names()
    except Exception:
        return False


def _linhas_relatorio_baixa(request):
    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    fatura = (request.GET.get("fatura") or "").strip()

    linhas = []
    total_baixado = Decimal("0.00")

    if _baixa_table_available():
        try:
            qs = BaixaFatura.objects.select_related("empresa", "usuario", "lancamento", "lancamento__fornecedor")

            empresa = _empresa_ativa(request)
            if empresa is not None:
                qs = qs.filter(empresa=empresa)
            if data_inicio:
                qs = qs.filter(data_baixa__gte=data_inicio)
            if data_fim:
                qs = qs.filter(data_baixa__lte=data_fim)
            if fornecedor:
                qs = qs.filter(
                    Q(fornecedor_snapshot__icontains=fornecedor)
                    | Q(lancamento__fornecedor__nome__icontains=fornecedor)
                    | Q(lancamento__fornecedor__nif__icontains=fornecedor)
                )
            if fatura:
                qs = qs.filter(
                    Q(numero_fatura_snapshot__icontains=fatura)
                    | Q(lancamento__numero_fatura__icontains=fatura)
                )

            for item in qs.order_by("-data_baixa", "-id"):
                valor_baixado = _to_decimal(getattr(item, "total_baixa", 0))
                total_baixado += valor_baixado
                linhas.append({
                    "id": item.id,
                    "empresa": getattr(item.empresa, "nome", "") if getattr(item, "empresa", None) else "",
                    "fornecedor": item.fornecedor_snapshot or _fornecedor_nome(item.lancamento),
                    "numero_fatura": item.numero_fatura_snapshot or getattr(item.lancamento, "numero_fatura", ""),
                    "data_baixa": item.data_baixa.strftime("%Y-%m-%d") if item.data_baixa else "",
                    "forma_pagamento": _forma_pagamento_baixa(item),
                    "usuario": getattr(item.usuario, "username", "") if getattr(item, "usuario", None) else "",
                    "valor_baixado": float(valor_baixado),
                    "saldo_resultante": float(_to_decimal(getattr(item, "saldo_resultante", 0))),
                })
        except (OperationalError, ProgrammingError):
            pass

    if not linhas:
        qs, _ = _filtrar_lancamentos_financeiros(request)
        qs = qs.filter(Q(status_pagamento__iexact="Parcial") | Q(status_pagamento__iexact="Paga"))

        for item in qs.order_by("-data_pagamento", "-data_emissao", "-id"):
            valor_baixado = _total_pago_lancamento(item)
            if valor_baixado <= Decimal("0.00"):
                continue
            total_baixado += valor_baixado
            linhas.append({
                "id": item.id,
                "empresa": getattr(getattr(item, "empresa", None), "nome", "") or "",
                "fornecedor": _fornecedor_nome(item),
                "numero_fatura": getattr(item, "numero_fatura", "") or "",
                "data_baixa": item.data_pagamento.strftime("%Y-%m-%d") if getattr(item, "data_pagamento", None) else (item.data_emissao.strftime("%Y-%m-%d") if getattr(item, "data_emissao", None) else ""),
                "forma_pagamento": _forma_pagamento_baixa(item),
                "usuario": getattr(getattr(item, "criado_por", None), "username", "") or "",
                "valor_baixado": float(valor_baixado),
                "saldo_resultante": float(_to_decimal(getattr(item, "saldo_aberto", 0))),
            })

    filtros = {
        "data_inicio": data_inicio.strftime("%Y-%m-%d") if data_inicio else "",
        "data_fim": data_fim.strftime("%Y-%m-%d") if data_fim else "",
        "fornecedor": fornecedor,
        "fatura": fatura,
    }
    return linhas, total_baixado, filtros


def _forma_pagamento_baixa(item):
    formas = []
    if _to_decimal(getattr(item, "dinheiro", 0)) > 0:
        formas.append("Dinheiro")
    if _to_decimal(getattr(item, "cartao", 0)) > 0:
        formas.append("Cartão")
    if _to_decimal(getattr(item, "transferencia", 0)) > 0:
        formas.append("Transferência")
    if _to_decimal(getattr(item, "mbway", 0)) > 0:
        formas.append("MBWay")
    if _to_decimal(getattr(item, "nota_credito", 0)) > 0:
        formas.append("Nota de Crédito")
    return ", ".join(formas)


def _safe_div(numerador, denominador, casas="0.00"):
    numerador = _to_decimal(numerador)
    denominador = _to_decimal(denominador)
    if denominador == 0:
        return Decimal(casas)
    return numerador / denominador


def _aplicar_estilo_excel(ws):
    preenchimento = PatternFill("solid", fgColor="D9EAF7")
    borda = Border(
        left=Side(style="thin", color="D5DCE5"),
        right=Side(style="thin", color="D5DCE5"),
        top=Side(style="thin", color="D5DCE5"),
        bottom=Side(style="thin", color="D5DCE5"),
    )
    for celula in ws[1]:
        celula.font = Font(bold=True, color="1E3A5F")
        celula.fill = preenchimento
        celula.border = borda
        celula.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for celula in row:
            celula.border = borda
            celula.alignment = Alignment(vertical="top")


def _auto_ajustar_colunas(ws):
    for coluna in ws.columns:
        tamanho = 0
        letra = coluna[0].column_letter
        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            tamanho = max(tamanho, len(valor))
        ws.column_dimensions[letra].width = min(max(tamanho + 2, 12), 35)


def _linhas_relatorio_combustivel(request):
    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    matricula = (request.GET.get("matricula") or "").strip()
    fornecedor = (request.GET.get("fornecedor") or "").strip()
    combustivel = (request.GET.get("combustivel") or "").strip()

    if ItemCombustivel is None:
        return [], [], {
            "valor_total": 0.0,
            "litros_total": 0.0,
            "km_total": 0.0,
            "media_km_l": 0.0,
            "media_euro_km": 0.0,
            "preco_medio_litro": 0.0,
        }

    qs = ItemCombustivel.objects.select_related(
        "lancamento", "lancamento__fornecedor", "lancamento__empresa", "frota", "combustivel"
    ).all()

    if hasattr(Lancamento, 'empresa_id'):
        empresa = _empresa_ativa(request)
        if empresa:
            qs = qs.filter(lancamento__empresa=empresa)

    if data_inicio:
        qs = qs.filter(lancamento__data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(lancamento__data_emissao__lte=data_fim)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)
    if fornecedor:
        qs = qs.filter(Q(lancamento__fornecedor__nome__icontains=fornecedor) | Q(lancamento__fornecedor__nif__icontains=fornecedor))
    if combustivel:
        qs = qs.filter(combustivel__nome__icontains=combustivel)

    linhas = []
    resumo_map = {}
    total_valor = Decimal("0.00")
    total_litros = Decimal("0.00")
    total_km = Decimal("0.00")

    for item in qs.order_by("-lancamento__data_emissao", "-id"):
        lanc = getattr(item, 'lancamento', None)
        litros = _to_decimal(getattr(item, 'litro', 0))
        valor_litro = _to_decimal(getattr(item, 'valor_litro', 0))
        valor_total = _to_decimal(getattr(item, 'valor_total', 0))
        if valor_total <= 0 and litros > 0 and valor_litro > 0:
            valor_total = litros * valor_litro
        km_inicio = Decimal(getattr(item, 'km_inicio', 0) or 0)
        km_final = Decimal(getattr(item, 'km_final', 0) or 0)
        km_total = Decimal(getattr(item, 'km_total', 0) or 0)
        if km_total <= 0 and km_final > km_inicio:
            km_total = km_final - km_inicio
        media_km_l = _safe_div(km_total, litros)
        media_euro_km = _safe_div(valor_total, km_total, "0.000")

        total_valor += valor_total
        total_litros += litros
        total_km += km_total

        matric = getattr(getattr(item, 'frota', None), 'matricula', '') or '-'
        acumulado = resumo_map.setdefault(matric, {
            'matricula': matric, 'abastecimentos': 0, 'litros': Decimal('0.00'), 'km': Decimal('0.00'), 'valor': Decimal('0.00')
        })
        acumulado['abastecimentos'] += 1
        acumulado['litros'] += litros
        acumulado['km'] += km_total
        acumulado['valor'] += valor_total

        linhas.append({
            'data': lanc.data_emissao.strftime('%Y-%m-%d') if getattr(lanc, 'data_emissao', None) else '',
            'empresa': getattr(getattr(lanc, 'empresa', None), 'nome', '') or '',
            'fornecedor': _fornecedor_nome(lanc) if lanc else '',
            'fatura': getattr(lanc, 'numero_fatura', '') or '',
            'matricula': matric,
            'combustivel': getattr(getattr(item, 'combustivel', None), 'nome', '') or '',
            'km_inicio': float(km_inicio),
            'km_final': float(km_final),
            'km_total': float(km_total),
            'litros': float(litros),
            'valor_litro': float(valor_litro),
            'valor_total': float(valor_total),
            'media_km_l': float(media_km_l),
            'media_euro_km': float(media_euro_km),
        })

    resumo_matriculas = []
    for item in resumo_map.values():
        litros = item['litros']
        km = item['km']
        valor = item['valor']
        resumo_matriculas.append({
            'matricula': item['matricula'],
            'abastecimentos': item['abastecimentos'],
            'litros': float(litros),
            'km': float(km),
            'valor': float(valor),
            'media_km_l': float(_safe_div(km, litros)),
            'media_euro_km': float(_safe_div(valor, km, '0.000')),
            'preco_medio_litro': float(_safe_div(valor, litros, '0.000')),
        })
    resumo_matriculas.sort(key=lambda x: x['matricula'])

    resumo = {
        'valor_total': float(total_valor),
        'litros_total': float(total_litros),
        'km_total': float(total_km),
        'media_km_l': float(_safe_div(total_km, total_litros)),
        'media_euro_km': float(_safe_div(total_valor, total_km, '0.000')),
        'preco_medio_litro': float(_safe_div(total_valor, total_litros, '0.000')),
    }
    return linhas, resumo_matriculas, resumo


@login_required
def relatorio_faturas(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-faturas")

    linhas, total_baixado, filtros = _linhas_relatorio_baixa(request)
    return JsonResponse({
        "cabecalho": "RELATÓRIO DE BAIXA DE FATURAS",
        "linhas": linhas,
        "totais": {"total_baixado": float(total_baixado)},
        "filtros": filtros,
    })


@login_required
def relatorio_frota(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-frota")

    matricula = (request.GET.get("matricula") or "").strip()
    qs = Frota.objects.all()

    if matricula:
        qs = qs.filter(matricula__icontains=matricula)

    linhas = []
    hoje = date.today()
    limite = hoje + timedelta(days=30)

    for item in qs.order_by("matricula"):
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)
        dias_seguro = (seguro - hoje).days if seguro else None
        dias_inspecao = (inspecao - hoje).days if inspecao else None

        status = "OK"
        if seguro and seguro < hoje:
            status = "Vencido"
        elif inspecao and inspecao < hoje:
            status = "Vencido"
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            status = "A vencer"

        linhas.append({
            "matricula": item.matricula,
            "seguradora": _seguradora_frota(item),
            "seguro": seguro.strftime("%Y-%m-%d") if seguro else "",
            "dias_seguro": dias_seguro,
            "inspecao": inspecao.strftime("%Y-%m-%d") if inspecao else "",
            "dias_inspecao": dias_inspecao,
            "status": status,
        })

    return JsonResponse({"linhas": linhas})


@login_required
def relatorio_manutencao(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-manutencao")

    if ManutencaoFrota is None:
        return JsonResponse({"linhas": []})

    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    matricula = (request.GET.get("matricula") or "").strip()

    qs = ManutencaoFrota.objects.select_related("frota", "lancamento").all()

    if data_inicio:
        qs = qs.filter(lancamento__data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(lancamento__data_emissao__lte=data_fim)
    if matricula:
        qs = qs.filter(frota__matricula__icontains=matricula)

    linhas = []
    for item in qs.order_by("-lancamento__data_emissao", "-id"):
        data_ref = getattr(item.lancamento, "data_emissao", None)
        linhas.append({
            "data": data_ref.strftime("%Y-%m-%d") if data_ref else "",
            "matricula": getattr(item.frota, "matricula", "") if getattr(item, "frota", None) else "",
            "km_inicio": float(_to_decimal(item.km_inicio)),
            "km_final": float(_to_decimal(item.km_final)),
            "km_total": float(_to_decimal(item.km_total)),
            "valor": float(_to_decimal(item.valor)),
            "observacao": getattr(item, "observacao", ""),
        })

    return JsonResponse({"linhas": linhas})


@login_required
def relatorio_caixa(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-caixa")

    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))

    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor").all())

    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)

    linhas = []
    total_dinheiro = Decimal("0.00")
    total_cartao = Decimal("0.00")
    total_geral = Decimal("0.00")

    for item in qs.order_by("-data_emissao", "-id"):
        dinheiro = _to_decimal(getattr(item, "dinheiro", 0))
        cartao = _to_decimal(getattr(item, "cartao", 0))
        transferencia = _to_decimal(getattr(item, "transferencia", 0))
        mbway = _to_decimal(getattr(item, "mbway", 0))
        nota_credito = _to_decimal(getattr(item, "nota_credito", 0))
        total_pago = dinheiro + cartao + transferencia + mbway + nota_credito

        total_dinheiro += dinheiro
        total_cartao += cartao
        total_geral += total_pago

        linhas.append({
            "id": item.id,
            "data_emissao": item.data_emissao.strftime("%Y-%m-%d") if item.data_emissao else "",
            "data": item.data_emissao.strftime("%Y-%m-%d") if item.data_emissao else "",
            "fatura": item.numero_fatura or f"Lançamento #{item.id}",
            "descricao": item.numero_fatura or f"Lançamento #{item.id}",
            "fornecedor": item.fornecedor.nome if getattr(item, "fornecedor", None) else "",
            "dinheiro": float(dinheiro),
            "cartao": float(cartao),
            "status_pagamento": getattr(item, "status_pagamento", ""),
            "total": float(total_pago),
        })

    return JsonResponse({
        "linhas": linhas,
        "totais": {
            "dinheiro": float(total_dinheiro),
            "cartao": float(total_cartao),
            "total": float(total_geral),
            "geral": float(total_geral),
        }
    })


@login_required
def relatorio_combustivel(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-combustivel")

    linhas, resumo_matriculas, resumo = _linhas_relatorio_combustivel(request)
    return JsonResponse({
        "linhas": linhas,
        "resumo_matriculas": resumo_matriculas,
        "resumo": resumo,
    })


@login_required
def relatorio_documentos(request):
    if not _is_ajax(request):
        return _render_dashboard(request, "relatorio-documentos")

    matricula = (request.GET.get("matricula") or "").strip()
    status = (request.GET.get("status") or "").strip().lower()

    qs = Frota.objects.all()
    if matricula:
        qs = qs.filter(matricula__icontains=matricula)

    hoje = date.today()
    limite = hoje + timedelta(days=30)

    linhas = []
    for item in qs.order_by("matricula"):
        seguro = getattr(item, "seguro", None)
        inspecao = getattr(item, "inspecao", None)

        estado = "OK"
        if seguro and seguro < hoje:
            estado = "Vencido"
        elif inspecao and inspecao < hoje:
            estado = "Vencido"
        elif (seguro and seguro <= limite) or (inspecao and inspecao <= limite):
            estado = "A vencer"

        if status:
            if status == "ok" and estado != "OK":
                continue
            if status == "a vencer" and estado != "A vencer":
                continue
            if status == "vencido" and estado != "Vencido":
                continue

        linhas.append({
            "matricula": item.matricula,
            "seguradora": _seguradora_frota(item),
            "seguro": seguro.strftime("%Y-%m-%d") if seguro else "",
            "inspecao": inspecao.strftime("%Y-%m-%d") if inspecao else "",
            "status": estado,
        })

    return JsonResponse({"linhas": linhas})


@login_required
def exportar_financeiro_excel(request):
    qs, _ = _filtrar_lancamentos_financeiros(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Financeiro"
    ws.append(["Data Emissão", "Vencimento", "Pagamento", "Fatura", "Fornecedor", "NIF", "Valor Fatura", "Total Pago", "Saldo", "Status"])

    for item in qs:
        ws.append([
            item.data_emissao.strftime("%d/%m/%Y") if getattr(item, "data_emissao", None) else "",
            item.data_vencimento.strftime("%d/%m/%Y") if getattr(item, "data_vencimento", None) else "",
            item.data_pagamento.strftime("%d/%m/%Y") if getattr(item, "data_pagamento", None) else "",
            getattr(item, "numero_fatura", "") or "",
            _fornecedor_nome(item),
            _fornecedor_nif(item),
            float(_to_decimal(getattr(item, "valor_fatura", 0))),
            float(_total_pago_lancamento(item)),
            float(_to_decimal(getattr(item, "saldo_aberto", 0))),
            getattr(item, "status_pagamento", "") or "",
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_financeiro.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_faturas_excel(request):
    linhas, total_baixado, _ = _linhas_relatorio_baixa(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Baixa de Faturas"
    ws.append(["Empresa", "Fornecedor", "Número Fatura", "Data Baixa", "Forma Pagamento", "Utilizador", "Valor Baixado", "Saldo Restante"])

    for item in linhas:
        ws.append([
            item.get("empresa", ""),
            item.get("fornecedor", ""),
            item.get("numero_fatura", ""),
            item.get("data_baixa", ""),
            item.get("forma_pagamento", ""),
            item.get("usuario", ""),
            item.get("valor_baixado", 0),
            item.get("saldo_resultante", 0),
        ])

    ws.append([])
    ws.append(["", "", "", "", "Total Baixado", float(total_baixado), "", ""])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_baixa_faturas.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_frota_excel(request):
    return HttpResponse("Exportação frota ainda a ajustar no backend.", content_type="text/plain")


@login_required
def exportar_manutencao_excel(request):
    return HttpResponse("Exportação manutenção ainda a ajustar no backend.", content_type="text/plain")


@login_required
def exportar_caixa_excel(request):
    data_inicio = _to_date(request.GET.get("data_inicio"))
    data_fim = _to_date(request.GET.get("data_fim"))
    qs = _filtrar_lancamentos_empresa(request, Lancamento.objects.select_related("fornecedor").all())
    if data_inicio:
        qs = qs.filter(data_emissao__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_emissao__lte=data_fim)

    wb = Workbook()
    ws = wb.active
    ws.title = "Folha de Caixa"
    ws.append(["Folha de Caixa"])
    ws.append(["Período", f"{data_inicio.strftime('%d/%m/%Y') if data_inicio else '-'} a {data_fim.strftime('%d/%m/%Y') if data_fim else '-'}"])
    ws.append([])
    ws.append(["ID", "Data", "Fatura", "Fornecedor", "Dinheiro (€)", "Cartão (€)", "Transferência (€)", "MBWay (€)", "Nota de Crédito (€)", "Total Pago (€)", "Estado"])

    total_dinheiro = Decimal("0.00")
    total_cartao = Decimal("0.00")
    total_transferencia = Decimal("0.00")
    total_mbway = Decimal("0.00")
    total_nota_credito = Decimal("0.00")
    total_geral = Decimal("0.00")

    for item in qs.order_by("-data_emissao", "-id"):
        dinheiro = _to_decimal(getattr(item, "dinheiro", 0))
        cartao = _to_decimal(getattr(item, "cartao", 0))
        transferencia = _to_decimal(getattr(item, "transferencia", 0))
        mbway = _to_decimal(getattr(item, "mbway", 0))
        nota_credito = _to_decimal(getattr(item, "nota_credito", 0))
        total_pago = dinheiro + cartao + transferencia + mbway + nota_credito

        total_dinheiro += dinheiro
        total_cartao += cartao
        total_transferencia += transferencia
        total_mbway += mbway
        total_nota_credito += nota_credito
        total_geral += total_pago

        ws.append([
            item.id,
            item.data_emissao.strftime("%d/%m/%Y") if item.data_emissao else "",
            item.numero_fatura or f"Lançamento #{item.id}",
            _fornecedor_nome(item),
            float(dinheiro),
            float(cartao),
            float(transferencia),
            float(mbway),
            float(nota_credito),
            float(total_pago),
            getattr(item, "status_pagamento", "") or "",
        ])

    ws.append([])
    ws.append(["Resumo", "", "", "", float(total_dinheiro), float(total_cartao), float(total_transferencia), float(total_mbway), float(total_nota_credito), float(total_geral), ""])

    for row in ws.iter_rows(min_row=4, min_col=5, max_col=10):
        for cell in row:
            cell.number_format = '#,##0.00 [$€-pt-PT]'
    ws[1][0].font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells('A1:K1')
    ws['A2'].font = Font(bold=True)
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="folha_caixa_profissional.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_combustivel_excel(request):
    linhas, resumo_matriculas, resumo = _linhas_relatorio_combustivel(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Combustível"
    ws.append(["Relatório de Combustível"])
    ws.merge_cells('A1:N1')
    ws['A1'].font = Font(bold=True, size=14, color="1E3A5F")
    ws.append(["Valor total", resumo.get('valor_total', 0), "Litros totais", resumo.get('litros_total', 0), "Km totais", resumo.get('km_total', 0), "Média km/l", resumo.get('media_km_l', 0), "Média €/km", resumo.get('media_euro_km', 0), "Preço médio €/l", resumo.get('preco_medio_litro', 0)])
    ws.append([])
    ws.append(["Data", "Empresa", "Fornecedor", "Fatura", "Matrícula", "Combustível", "Km início", "Km final", "Km total", "Litros", "€/l", "Valor", "Km/l", "€/km"])
    for item in linhas:
        ws.append([
            item.get('data', ''), item.get('empresa', ''), item.get('fornecedor', ''), item.get('fatura', ''),
            item.get('matricula', ''), item.get('combustivel', ''), item.get('km_inicio', 0), item.get('km_final', 0),
            item.get('km_total', 0), item.get('litros', 0), item.get('valor_litro', 0), item.get('valor_total', 0),
            item.get('media_km_l', 0), item.get('media_euro_km', 0),
        ])
    for row in ws.iter_rows(min_row=4, min_col=10, max_col=14):
        for cell in row:
            if cell.column in (11,12,14):
                cell.number_format = '#,##0.000 [$€-pt-PT]' if cell.column != 12 else '#,##0.00 [$€-pt-PT]'
            else:
                cell.number_format = '#,##0.00'
    _aplicar_estilo_excel(ws)
    _auto_ajustar_colunas(ws)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="relatorio_combustivel.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_documentos_excel(request):
    return HttpResponse("Exportação documentos ainda a ajustar no backend.", content_type="text/plain")